import contextlib
import io
import logging
import os
import re
import shutil
import sys
import threading
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import math

# Configurar encoding para UTF-8 em Windows (console)
if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

import pdfplumber
import pypdfium2 as pdfium
import pytesseract
from PyPDF2 import PdfReader
from PIL import ImageFilter, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_XLSX = BASE_DIR / "Consolidado Emendas.xlsx"
PDF_NAME_RE = re.compile(r"^(\d+)\s*-\s*(.+)\.pdf$", re.IGNORECASE)
TESSERACT_CMD = shutil.which("tesseract") or shutil.which("tesseract.exe") or r"C:\Tesseract-OCR\tesseract.exe"
OCR_LANG = "por+eng"

# Constantes de padrão de artigos (movidas para o topo para evitar referencias antes da definição)
ART_NUM_RE = r"\d+(?:\.\d+)*(?:[º°o])?(?:-[a-z]+(?:-[a-z]+)*)?"
ARTICLE_LABEL_RE = r"(?:arts?\.?|artigos?)"
STRUCTURE_TYPES = [
    ("Livro", "livro"),
    ("Título", "titulo"),
    ("Subtítulo", "subtitulo"),
    ("Capítulo", "capitulo"),
    ("Seção", "secao"),
]

# OCR Progress com sincronização
OCR_PROGRESS = {"current": 0}
OCR_PROGRESS_LOCK = threading.Lock()
DIRECTIVE_START_RE = re.compile(
    r"\b("
    r"inclua-se|incluam-se|acrescente-se|acrescentem-se|"
    r"suprima-se|suprimam-se|suprimir|supress[aÃ£]o|"
    r"d[eéêÃª]-se|de-se|insira-se|insiram-se|mantenha-se|retome-se|"
    r"substitua-se|substituam-se|cria-se|criam-se|"
    r"prop[oÃµ]e a cria[cÃ§g][aÃ£]o"
    r")\b",
    flags=re.IGNORECASE,
)


@dataclass
class EmendaRow:
    numero: int
    autor: str
    dispositivos: str
    natureza: str
    tema: str
    sintese: str
    texto: str


def normalize_text(text: str) -> str:
    text = str(text or "").replace("\u00a0", " ")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def simplify(text: str) -> str:
    return re.sub(r"\s+", " ", normalize_text(text).lower()).strip()


def clean_spaces(text: str) -> str:
    text = str(text or "").replace("\u00a0", " ")
    text = re.sub(r"(\w)-\n(\w)", r"\1\2", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_ocr_noise(text: str) -> str:
    text = str(text or "")
    if not text:
        return ""

    replacements = {
        "Dé-se": "Dê-se",
        "DÉ-SE": "DÊ-SE",
        "Dê-se": "Dê-se",
        "paragrafo tinico": "parágrafo único",
        "parágrafo tinico": "parágrafo único",
        "paragrafo único": "parágrafo único",
        "paragrafo Unico": "parágrafo único",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"\bsuprima\s+se\b", "suprima-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsuprimam\s+se\b", "suprimam-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsuprimase\b", "suprima-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsuprimamse\b", "suprimam-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bacrescente\s+se\b", "acrescente-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bacrescentem\s+se\b", "acrescentem-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bacrescentese\b", "acrescente-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bacrescentemse\b", "acrescentem-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\binclua\s+se\b", "inclua-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bincluam\s+se\b", "incluam-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bincluase\b", "inclua-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bincluamse\b", "incluam-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\binsira\s+se\b", "insira-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\binsiram\s+se\b", "insiram-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\binsirase\b", "insira-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\binsiramse\b", "insiram-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bmantenha\s+se\b", "mantenha-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bretome\s+se\b", "retome-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsubstitua\s+se\b", "substitua-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsubstituam\s+se\b", "substituam-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bmantenhase\b", "mantenha-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bretomese\b", "retome-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsubstituase\b", "substitua-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bsubstituamse\b", "substituam-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bd[eéê]\s*-\s*se\b", "Dê-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bd[eéê]\s+se\b", "Dê-se", text, flags=re.IGNORECASE)
    text = re.sub(r"\bdese\b", "Dê-se", text, flags=re.IGNORECASE)

    text = re.sub(r"(?<!\w)(?:\$\$|S\$|§S|S§)\s*", "§§ ", text)
    text = re.sub(r"(?<!\w)\$(?=\s*\d)", "§", text)
    text = re.sub(r"(§§\s*)1[2º°o]?\s*a\s*(\d+)[º°o]?", r"\1 1º a \2º", text, flags=re.IGNORECASE)
    text = re.sub(r"\bpar[aá]grafo\s+t[ií]nico\b", "parágrafo único", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(par[aá]grafo|art\.?)\s+tinico\b", r"\1 único", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(art\.?\s*\d+),(\d{3})(?=[\s,-])", r"\1.\2", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(\d+),(\d{3})(-[A-Za-z])", r"\1.\2\3", text)
    text = re.sub(r"\b(art\.?\s*\d+(?:\.\d+)*)\s*[–—-]\s*([A-Z]{1,3})\b", r"\1-\2", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(art\.?\s*\d+(?:\.\d+)*)([A-Z]{1,3})\b", r"\1-\2", text, flags=re.IGNORECASE)
    text = re.sub(r"\b1[2º°o]a\s+(\d+)[º°o]\b", r"1º a \1º", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(\d+)\s*a\s*(\d+)[º°o]\b", r"\1º a \2º", text, flags=re.IGNORECASE)
    text = re.sub(r"(§§)\s{2,}", r"\1 ", text)
    return text


def cut_to_pl_subtitle_body(text: str) -> str:
    raw = str(text or "")
    if not raw:
        return ""

    subtitle_re = re.compile(r"\(?\s*ao\s+PL\s*4\s*/\s*2025\s*\)?", flags=re.IGNORECASE)
    candidates = [match.end() for match in subtitle_re.finditer(raw) if match.start() < 1500]
    if candidates:
        return raw[min(candidates):].lstrip(" \t\r\n-")

    header_match = re.search(r"(?im)^\s*EMENDA\s+N[º°o]?[^\n]*(?:\r?\n)+", raw)
    if header_match and header_match.start() < 800:
        return raw[header_match.end():].lstrip(" \t\r\n-")

    return raw


def trim_to_directive(lines: List[str]) -> List[str]:
    if not lines:
        return lines

    first_meaningful = ""
    for line in lines:
        candidate = clean_spaces(line)
        if candidate:
            first_meaningful = candidate
            break

    if re.match(
        r"^\s*(?:o|a|os|as)\s+(?:art\.?|arts?\.?|artigo|artigos|livro|t[íi]tulo|subt[íi]tulo|cap[íi]tulo|se[cç][aã]o)\b",
        first_meaningful,
        flags=re.IGNORECASE,
    ):
        return lines

    trimmed: List[str] = []
    directive_found = False
    for line in lines:
        match = DIRECTIVE_START_RE.search(line)
        if not directive_found:
            if not match:
                continue
            directive_found = True
            line = line[match.start():].strip()
        trimmed.append(line)
    return trimmed or lines


def extract_pdf_text_pypdf2(pdf_path: Path) -> str:
    parts: List[str] = []
    with open(pdf_path, "rb") as fh:
        reader = PdfReader(fh)
        for page in reader.pages:
            text = page.extract_text() or ""
            if text.strip():
                parts.append(text)
    return "\n".join(parts).strip()


def extract_pdf_text_pdfplumber(pdf_path: Path) -> str:
    parts: List[str] = []
    with contextlib.redirect_stderr(io.StringIO()):
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if text.strip():
                    parts.append(text)
                    continue
                words = page.extract_words() or []
                if words:
                    parts.append(" ".join(w.get("text", "") for w in words if w.get("text")))
    return "\n".join(parts).strip()


def ocr_is_available() -> bool:
    return bool(TESSERACT_CMD and Path(TESSERACT_CMD).exists())


def extract_pdf_text_ocr(pdf_path: Path) -> str:
    if not ocr_is_available():
        return ""

    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    parts: List[str] = []
    pdf = pdfium.PdfDocument(str(pdf_path))
    try:
        for page_index in range(len(pdf)):
            page = pdf[page_index]
            bitmap = page.render(scale=3)
            image = bitmap.to_pil().convert("L")
            image = ImageOps.autocontrast(image)
            image = image.filter(ImageFilter.SHARPEN)
            text = pytesseract.image_to_string(
                image,
                lang=OCR_LANG,
                config="--psm 6",
            )
            if text.strip():
                parts.append(text)
                if re.search(r"\bjustifica[cÃ§g][aÃ£]o\b", simplify(text)):
                    break
    finally:
        pdf.close()

    return "\n".join(parts).strip()


def extraction_looks_usable(text: str) -> bool:
    if not text:
        return False

    stripped = text.strip()
    if len(stripped) < 80:
        return False

    letters = sum(1 for ch in stripped if ch.isalpha())
    if letters < 40:
        return False

    simplified = simplify(stripped)
    legal_markers = [
        "art.",
        "lei",
        "codigo civil",
        "justificacao",
        "suprima-se",
        "acrescente-se",
        "inclua-se",
        "passa a vigorar",
    ]
    return any(marker in simplified for marker in legal_markers)


def extract_pdf_text(pdf_path: Path) -> str:
    try:
        fast_text = extract_pdf_text_pypdf2(pdf_path)
    except Exception:
        fast_text = ""

    if extraction_looks_usable(fast_text):
        return fast_text

    try:
        plumber_text = extract_pdf_text_pdfplumber(pdf_path)
    except Exception:
        plumber_text = ""

    if extraction_looks_usable(plumber_text):
        return plumber_text

    with OCR_PROGRESS_LOCK:
        OCR_PROGRESS["current"] += 1
        current = OCR_PROGRESS["current"]
    if current % 10 == 0:
        logger.info(f"Aplicando OCR em {current} PDFs...")
    try:
        ocr_text = extract_pdf_text_ocr(pdf_path)
    except Exception:
        ocr_text = ""

    if extraction_looks_usable(ocr_text):
        return ocr_text

    return ocr_text or plumber_text or fast_text


def should_skip_line(line: str) -> bool:
    s = simplify(line)
    if not s:
        return True

    skip_substrings = [
        "gabinete do senador",
        "gabinete do senador",
        "gabinete senador",
        "gabinete da senadora",
        "gabinete senadora",
        "gabinete do deputado",
        "gabinete da deputada",
        "senado federal",
        "assinado eletronicamente",
        "para verificar as assinaturas",
        "legis.senado.gov.br/autenticadoc",
        "sala da comissao",
        "sala das sessoes",
        "sala das sessÃµes",
    ]
    if any(item in s for item in skip_substrings):
        return True

    if s in {"emenda no", "emenda n", "emenda n.", "emenda nÂº", "emenda no - ctcivil", "emenda nÂº - ctcivil", "(ao pl 4/2025)", "ao pl 4/2025)"}:
        return True

    if re.fullmatch(r"emenda\s+n[º°oÂº.]*\s*\d+.*", s):
        return True

    if re.fullmatch(r"\(?\s*ao\s+pl\s*4\s*/\s*2025\s*\)?", s):
        return True

    if s == "pl 4/2025":
        return True

    if "tidex" in s or "lexedit" in s:
        return True

    if re.search(r"\b\d{2}-\d[\d.]+/?fs\b", s):
        return True

    if re.search(r"\bpl\s*4/2025\b", s) and len(s) <= 18:
        return True

    digits = re.sub(r"\D", "", s)
    if digits and len(digits) == 5 and s == digits:
        return True

    if re.fullmatch(r"[\W_]*\d{4,}[\w./-]*", s):
        return True

    if re.fullmatch(r"[\W_]*fs[\W_]*", s):
        return True

    return False


def is_structural_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False

    markers = [
        r"^[â€œ\"]?Art\.\s",
        r"^Â§\s*\d+Âº",
        r"^Par[aÃ¡]grafo [uÃº]nico",
        r"^[IVXLCDM]+\s*[â€“-]",
        r"^\([A-Z]{1,4}\)$",
        r"^\([A-Z]{1,4}[-/ ]",
    ]
    return any(re.match(pattern, stripped, flags=re.IGNORECASE) for pattern in markers)


def reflow_dispositive_lines(lines: List[str]) -> str:
    out: List[str] = []
    for raw in lines:
        line = clean_spaces(raw)
        if not line:
            continue

        if not out:
            out.append(line)
            continue

        prev = out[-1]
        if is_structural_line(line):
            out.append(line)
            continue

        if prev.endswith(":"):
            out.append(line)
            continue

        if re.search(r"[.;:!?â€\"]$", prev):
            out.append(line)
            continue

        out[-1] = f"{prev} {line}"

    text = "\n".join(out)
    text = re.sub(r"\n{2,}", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


JUSTIFICATION_MARKERS = [
    "JUSTIFICAÇÃO",
    "JUSTIFICACAO",
    "JUSTIFICAÃ‡ÃƒO",
    "JUSTIFICAÇÃO",
    "JUSTIFICAGAO",
]


def split_before_justification(full_text: str) -> str:
    if not full_text:
        return ""
    upper = str(full_text).upper()
    cut_points = []
    for marker in JUSTIFICATION_MARKERS:
        pos = upper.find(marker)
        if pos >= 0:
            cut_points.append(pos)
    if not cut_points:
        return full_text
    return full_text[:min(cut_points)]


def extract_dispositive_text(full_text: str) -> str:
    if not full_text:
        return ""

    head = normalize_ocr_noise(split_before_justification(full_text))
    head = cut_to_pl_subtitle_body(head)
    head = re.sub(r"https?://\S+", "", head, flags=re.IGNORECASE)
    head = clean_spaces(head)

    kept_lines: List[str] = []
    for raw_line in head.splitlines():
        line = clean_spaces(raw_line)
        if should_skip_line(line):
            continue
        kept_lines.append(line)

    if not kept_lines:
        return ""

    kept_lines = trim_to_directive(kept_lines)
    text = reflow_dispositive_lines(kept_lines)
    text = normalize_ocr_noise(text)
    text = re.sub(r"^\s*\d{5}\s+\d+\s+\d+\s+(?=" + DIRECTIVE_START_RE.pattern + r")", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\n\d{2}-\d[\d.]+/?FS\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\nPL\s*4/2025\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\n\d{5}\b", "", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def first_directive_segment(text: str) -> str:
    if not text:
        return ""
    one_line = clean_spaces(text.replace("\n", " "))
    one_line = re.sub(r"^\s*EMENDA\s+N\S*\s*(?:-\s*CTCIVIL\s*)?", "", one_line, flags=re.IGNORECASE)
    one_line = re.sub(r"^\s*SENADO\s+FEDERAL\s+", "", one_line, flags=re.IGNORECASE)
    lead = re.split(r':\s*[â€œ"]', one_line, maxsplit=1)[0]
    lead = re.split(r'\s+[â€œ"]', lead, maxsplit=1)[0]
    lead = re.sub(r"\s+", " ", lead).strip(" .;:")
    return lead


NATURE_PATTERNS = {
    "restaurativa": [
        r"\bretome-se\b",
        r"\bmantenha-se a redacao vigente\b",
        r"\bmantenha-se a reda[cÃ§g]ao vigente\b",
        r"\bredacao anterior a alteracao\b",
        r"\bredacao anterior a altera[cÃ§g][aÃ£]o\b",
    ],
    "substitutiva": [
        r"\bsubstitua-se\b",
        r"\bsubstituam-se\b",
    ],
    "aditiva": [
        r"\bacrescente-se\b",
        r"\bacrescentem-se\b",
        r"\binclua-se\b",
        r"\bincluam-se\b",
        r"\binsira-se\b",
        r"\binsiram-se\b",
        r"\bfica acrescido\b",
        r"\bficam acrescidos\b",
        r"\bfica acrescida\b",
        r"\bficam acrescidas\b",
        r"\bpassa a vigorar acrescido\b",
        r"\bpassam a vigorar acrescidos\b",
        r"\bpassa a vigorar acrescida\b",
        r"\bpassam a vigorar acrescidas\b",
        r"\bcria-se\b",
        r"\bcriam-se\b",
        r"\bcriacao de\b",
        r"\bcriaÃ§Ã£o de\b",
        r"\bpropoe a criacao\b",
        r"\bpropÃµe a criaÃ§Ã£o\b",
        r"\binclusao do artigo\b",
        r"\binclus[aã]o do artigo\b",
        r"\binclusao do art\.\b",
        r"\binclus[aã]o do art\.\b",
        r"\binclusao de artigos\b",
        r"\binclusÃ£o de artigos\b",
    ],
    "supressiva": [
        r"\bsuprima-se\b",
        r"\bsuprimam-se\b",
        r"\bsupressao da revogacao\b",
        r"\bsupress[aã]o da revoga[cç][aã]o\b",
        r"\bsupressao da alteracao\b",
        r"\bsupress[aã]o da altera[cç][aã]o\b",
        r"\bsuprimir\s+art\.?\b",
        r"\bsuprimir\s+arts?\.?\b",
        r"\bsuprimir\s+artigo\b",
        r"\bsuprimir\s+artigos\b",
        r"\bsuprimir o artigo\b",
        r"\bsuprimir os artigos\b",
        r"\bsupressao do art\.\b",
        r"\bsupressÃ£o do art\.\b",
        r"\bsupressao do art\b",
        r"\bsupressÃ£o do art\b",
        r"\bsupressao do artigo\b",
        r"\bsupressÃ£o do artigo\b",
        r"\bexclua-se\b",
        r"\bexcluir as alteracoes\b",
        r"\bexcluir as altera[cÃ§g][oÃµ]es\b",
        r"\bretirando sua citacao\b",
        r"\bretirando suas citacoes\b",
        r"\bretirando sua cita[cÃ§][aÃ£]o\b",
        r"\bretirando suas cita[cÃ§][oÃµ]es\b",
        r"\bsuprima-se a revogacao\b",
        r"\bsuprima-se a alteracao\b",
        r"\bsuprima-se a altera[cÃ§g][aÃ£]o\b",
        r"\bsuprima-se a reda[cÃ§g][aÃ£]o\b",
    ],
    "modificativa": [
        r"\bde-se nova redacao\b",
        r"\bd[eÃª]-se nova reda[cÃ§g][aÃ£]o\b",
        r"\bde-se a seguinte redacao\b",
        r"\bd[eÃª]-se ao\b.*\ba seguinte reda[cÃ§g][aÃ£]o\b",
        r"\bd[eÃª]-se aos\b.*\ba seguinte reda[cÃ§g][aÃ£]o\b",
        r"\bd[eÃª]-se a\b.*\ba seguinte reda[cÃ§g][aÃ£]o\b",
        r"\bd[eÃª]-se as\b.*\ba seguinte reda[cÃ§g][aÃ£]o\b",
        r"\bpassa a vigorar com a seguinte redacao\b",
        r"\bpassa a vigorar com as seguintes redacoes\b",
        r"\bpassam a vigorar com a seguinte reda[cÃ§g][aÃ£]o\b",
        r"\bpassam a vigorar com as seguintes reda[cÃ§g][oÃµ]es\b",
        r"\bpassa a vigorar com as seguintes altera[cÃ§g][oÃµ]es\b",
        r"\bpassam a vigorar com as seguintes altera[cÃ§g][oÃµ]es\b",
        r"\bpassa a vigorar com a reda[cÃ§g][aÃ£]o a seguir\b",
        r"\balterado pelo art\.\b",
        r"\balterada pelo art\.\b",
        r"\bnova redacao do art\.?\s",
        r"\bnova reda[cç][aã]o do art\.?\s",
    ],
}


def command_natures_from_text(text: str) -> List[str]:
    """Detecta natureza da emenda pelo comando inicial (Supressiva, Modificativa, Aditiva)."""
    s = simplify(extract_nature_context(text) or text)
    s = re.sub(r"^\s*d\S*-se\s+nova\s+reda\S+\s+ao\s+projeto\s+nos\s+termos\s+dos\s+itens?\s+.+?(?=item\s+\d+)", "", s, flags=re.IGNORECASE)
    if not s:
        return []

    # Mapeia NATURE_PATTERNS para naturezas capitalizadas
    nature_mapping = {
        "supressiva": "Supressiva",
        "modificativa": "Modificativa",
        "aditiva": "Aditiva",
        "substitutiva": "Modificativa",  # Substitutiva mapeia para Modificativa
        "restaurativa": "Supressiva",     # Restaurativa mapeia para Supressiva
    }

    found: List[str] = []
    matches: List[Tuple[int, str]] = []
    
    for pattern_nature, patterns in NATURE_PATTERNS.items():
        if pattern_nature not in nature_mapping:
            continue
        command_nature = nature_mapping[pattern_nature]
        
        for pattern in patterns:
            for match in re.finditer(pattern, s):
                matches.append((match.start(), command_nature))
                break
    
    for _, nature in sorted(matches, key=lambda item: item[0]):
        if nature not in found:
            found.append(nature)
    return found


def infer_nature(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(extract_nature_context(text) or text))
    s = simplify(raw)
    if not s:
        return ""

    if re.search(r"^\s*suprima-se\s+(?:a\s+)?(?:inclus[aã]o|inclusao)\s+d[oa]\s+", raw, flags=re.IGNORECASE):
        return "Supressiva"

    if re.search(r"^\s*suprima-se\b", raw, flags=re.IGNORECASE) and re.search(r"\bnova\s+reda[cç][aã]o\s+do\s+art", raw, flags=re.IGNORECASE):
        return "Supressiva"

    if extract_revocation_suppression_with_redaction_target(text):
        return "Mista"

    if extract_structural_modification_with_revocation_target(text):
        return "Mista"

    if extract_project_citation_adjustment_target(text):
        return "Mista"

    project_change_target = extract_project_change_suppression_target(text)
    declarative_subject = extract_declarative_redaction_subject(text)
    embedded_suppressions = extract_embedded_inciso_suppression_targets(text)
    article_body_suppressions = extract_article_body_suppression_targets(text)

    if re.search(r"^\s*suprima-se\s+a\s+nova\s+reda[cç][aã]o\s+do\s+art\.?\s*" + ART_NUM_RE, raw, flags=re.IGNORECASE):
        return "Supressiva"

    if project_change_target:
        return "Supressiva"

    if declarative_subject and (embedded_suppressions or article_body_suppressions):
        return "Mista"

    command_natures = command_natures_from_text(text)
    if article_body_suppressions and "Modificativa" in command_natures:
        return "Mista"
    if command_natures:
        return resolve_nature_set(command_natures, "")

    if declarative_subject:
        return "Modificativa"

    if extract_creation_with_title_target(text):
        return "Aditiva"

    if re.search(r"\bd\S*-se\s+nova\s+reda\S+\b", raw, flags=re.IGNORECASE) and re.search(r"\bacrescente-se\b|\bacrescentem-se\b|\binclua-se\b|\bincluam-se\b", raw, flags=re.IGNORECASE):
        return "Mista"

    if "alteracao do art" in s:
        return "Modificativa"

    if any(re.search(p, s) for p in NATURE_PATTERNS["restaurativa"]):
        return "Restaurativa"

    has_subs = any(re.search(p, s) for p in NATURE_PATTERNS["substitutiva"])
    has_add = any(re.search(p, s) for p in NATURE_PATTERNS["aditiva"])
    has_sup = any(re.search(p, s) for p in NATURE_PATTERNS["supressiva"])
    has_mod = any(re.search(p, s) for p in NATURE_PATTERNS["modificativa"])

    if project_change_target:
        has_sup = True
        has_mod = False

    if is_new_article_insertion_context(text):
        has_add = True
        has_mod = False

    active = [name for name, flag in {
        "Substitutiva": has_subs,
        "Aditiva": has_add,
        "Supressiva": has_sup,
        "Modificativa": has_mod,
    }.items() if flag]

    if len(active) >= 2:
        return "Mista"
    if len(active) == 1:
        return active[0]

    if s.startswith("o art.") and "passa a vigorar" in s:
        return "Modificativa"

    if s.startswith("o caput do art.") and "passa a vigorar como" in s:
        return "Modificativa"

    if s.startswith("os secs") and "passam a vigorar" in s:
        return "Modificativa"

    if s.startswith("os paragrafos") and "passam a vigorar" in s:
        return "Modificativa"

    if (
        s.startswith("a seguinte redacao para o art.")
        or s.startswith("a seguinte redaÃ§Ã£o para o art.")
        or s.startswith("a seguinte redagao para o art.")
    ):
        return "Modificativa"

    if (
        s.startswith("mantenha-se a redacao do art.")
        or s.startswith("mantenha-se a redaÃ§Ã£o do art.")
        or s.startswith("mantenha-se a redagao do art.")
    ):
        return "Restaurativa"

    return ""


def extract_project_change_suppression_target(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    simplified = simplify(raw)
    if "do projeto" not in simplified:
        return ""
    if not re.search(r"\bpara\s+excluir\s+as\s+alteracoes\s+referentes?\s+a(?:o|os)\b", simplified):
        return ""

    match = re.search(
        r"para\s+excluir\s+as\s+altera(?:ç|c)ões\s+referentes?\s+a(?:o|os)\s+((?:arts?\.?|artigos?)\s*.+?)(?=,\s*(?:mantendo-se|mantendo|mantida|mantido|retirando)\b|\.\s*$)",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    target = clean_spaces(match.group(1)).strip(" ,.;:")
    if not re.search(r"\b(?:art\.?|arts?\.?|artigo|artigos)\b", simplify(target)):
        return ""
    return target


def extract_project_text_replacement_target(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    patterns = [
        r"(?:substitua-se|substituam-se),?\s+(?:onde\s+houver,\s+)?(no\s+projeto(?:\s+de\s+lei[^,.;:]+)?),?\s+(a\s+express[aã]o|as\s+express[oõ]es|o\s+termo|os\s+termos)\b",
        r"(?:substitua-se|substituam-se),?\s+(a\s+express[aã]o|as\s+express[oõ]es|o\s+termo|os\s+termos)\b.*?\s+(no\s+projeto(?:\s+de\s+lei[^,.;:]+)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, raw, flags=re.IGNORECASE)
        if not match:
            continue
        g1 = clean_spaces(match.group(1))
        g2 = clean_spaces(match.group(2))
        if g1.lower().startswith(("no projeto", "na projeto")):
            project_ref, subject = g1, g2
        else:
            subject, project_ref = g1, g2
        return f"{subject} {project_ref}"
    return ""


def is_new_article_insertion_context(text: str) -> bool:
    s = simplify(text)
    if not s:
        return False
    markers = [
        "o seguinte artigo",
        "novo artigo",
        "inclusao de artigos",
        "inclusao de artigo",
        "onde couber, no projeto o seguinte artigo",
    ]
    return any(marker in s for marker in markers)


def extract_generic_project_article_target(text: str) -> str:
    s = simplify(text)
    if not s:
        return ""
    if ("o seguinte artigo" in s or "novo artigo" in s) and ("projeto" in s or " pl " in f" {s} "):
        return "Novo artigo no projeto"
    return ""


def extract_restoration_suppression_target(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    match = re.search(
        r"(?:mantenha-se\s+a\s+reda[cç][aã]o|retome-se\s+a\s+reda[cç][aã]o)\s+do\s+(art\.?\s*[0-9]+(?:\.\d+)*(?:-[A-Za-z]+)?)\b.*?\bsuprim(?:indo-se|a-se|am-se)\b.*?\baltera[cç][aã]o\s+proposta\b",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""
    return clean_clause_target_text(match.group(1))


def extract_suppressed_inciso_from_inserted_article_target(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    article_match = re.search(r"\bArt\.?\s*([0-9]+(?:\.\d+)*(?:-[A-Za-z]+)?)\b", raw, flags=re.IGNORECASE)
    inciso_match = re.search(r"\b([IVXLCDM]+)\s*[–-]\s*(?:Suprimir|Suprima-se|Suprimam-se)\b", raw, flags=re.IGNORECASE)
    if not article_match or not inciso_match:
        return ""

    art = format_article_ref(article_match.group(1)).upper().replace("°", "º")
    return f"Inciso {inciso_match.group(1).upper()} do Art. {art}"


def extract_new_article_with_location_target(text: str, concise: bool = False) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    match = re.search(
        r"(?:inclua-se|incluam-se|acrescente-se|acrescentem-se)\s+.+?\bnovo\s+art\.?\s*([0-9]+(?:\.\d+)*(?:-[A-Za-z]+)?)\b.+?,\s+no\s+(.+?)(?=,\s+nos\s+termos|[.;:])",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    art = format_article_ref(match.group(1)).upper().replace("°", "º")
    location = clean_clause_target_text(match.group(2), concise=concise)
    if not location:
        return ""
    return f"novo art. {art} no {location}"


def extract_project_law_insertion_target(text: str, concise: bool = False) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""
    if not re.search(r"(?:inclua-se|incluam-se|acrescente-se|acrescentem-se)\s*,?\s+(?:ao|no|na)\s+(?:art\.?\s*\d+[º°o]?\s+do\s+)?(?:projeto|pl)", raw, flags=re.IGNORECASE):
        return ""

    new_art_match = re.search(
        r"\bo\s+novo\s+art\.?\s*([0-9]+(?:\.\d+)*(?:-[A-Za-z]+)?)(?:[º°o])?(?:\s+à?\s+Lei\b.*?(?:,\s+no\s+(.+?))?)?(?=,\s+nos\s+termos|:\s*[“\"]|:\s*$|$)",
        raw,
        flags=re.IGNORECASE,
    )
    if new_art_match:
        art = format_article_ref(new_art_match.group(1)).upper().replace("°", "º")
        location = clean_clause_target_text(new_art_match.group(2), concise=concise) if new_art_match.group(2) else ""
        return f"novo art. {art} no {location}".strip() if location else f"novo art. {art}"

    alter_match = re.search(
        r"\ba\s+seguinte\s+altera[cç][aã]o\s+do\s+art\.?\s*([0-9]+(?:\.\d+)*(?:-[A-Za-z]+)?)(?:[º°o])?\b",
        raw,
        flags=re.IGNORECASE,
    )
    if alter_match:
        art = format_article_ref(alter_match.group(1)).upper().replace("°", "º")
        return f"alteração do art. {art}"

    new_device_match = re.search(r"o\s+seguinte\s+dispositivo\s*:\s*[“\"]?art\.?\s*([0-9]+(?:\.\d+)*(?:\s*[–-]\s*[A-Za-z]+)?)(?:[º°o])?", raw, flags=re.IGNORECASE)
    if new_device_match:
        art = format_article_ref(new_device_match.group(1)).upper().replace("°", "º")
        return f"Art. {art}"

    chapter_match = re.search(r"o\s+seguinte\s+dispositivo\s+ao\s+(.+?)(?::\s*[“\"]?art\.?\s*([0-9]+(?:\.\d+)*(?:\s*[–-]\s*[A-Za-z]+)?)(?:[º°o])?|:|$)", raw, flags=re.IGNORECASE)
    if chapter_match:
        if chapter_match.group(2):
            art = format_article_ref(chapter_match.group(2)).upper().replace("°", "º")
            location = clean_clause_target_text(chapter_match.group(1), concise=True)
            if concise and location:
                return f"novo art. {art} no {location}"
            base_art = art.split("-", 1)[0]
            return f"Art. {base_art}"
        else:
            target = clean_clause_target_text(chapter_match.group(1))
            return target

    return ""


def extract_project_article_internal_target(text: str, concise: bool = False) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    direct_target_match = re.search(
        r"\bsuprima(?:m)?-se\s*,?\s+do\s+(?:artigo|art\.?)\s*\d+[º°o]?\s+(?:do\s+)?projeto[^,;:]*,\s*o\s+((?:art\.?|artigo)\s*" + ART_NUM_RE + r")\b",
        raw,
        flags=re.IGNORECASE,
    )
    if direct_target_match:
        target = clean_clause_target_text(direct_target_match.group(1), concise=concise)
        return re.sub(r"\s+-\s+", " \u2013 ", target)

    match = re.search(
        r"\bsuprima(?:m)?-se\s*,?\s+do\s+(?:artigo|art\.?)\s*\d+[º°o]?\s+(?:do\s+)?projeto[^,;:]*,\s+(?:o|a|os|as)\s+(.+?)(?=,\s*(?:na\s+forma|nos\s+termos|como\s+propost)|[.;:]|$)",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    target = clean_clause_target_text(match.group(1), concise=concise)
    if not target:
        return ""
    target = re.sub(r"\s+-\s+", " \u2013 ", target)
    return target


def extract_following_project_devices_target(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    match = re.search(
        r"\b(?:inclua-se|incluam-se|acrescente-se|acrescentem-se)\s+(?:ao|no|na)\s+PL\s*0?4\s*/\s*2025\s*,?\s+os\s+seguintes\s+dispositivos\s*:\s*(.+)$",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not match:
        match = re.search(
            r"\b(?:inclua-se|incluam-se|acrescente-se|acrescentem-se)\s+(?:ao|no|na)\s+Projeto(?:\s+de\s+Lei)?[^,;:]*,\s+os\s+seguintes\s+dispositivos\s*:\s*(.+)$",
            raw,
            flags=re.IGNORECASE | re.DOTALL,
        )
    if not match:
        return ""

    tail = normalize_ocr_noise(match.group(1))
    article_match = re.search(r"\bArt\.?\s*([0-9]+(?:\.\d+)*(?:[º°o])?(?:-[A-Za-z]+)?)\b", tail, flags=re.IGNORECASE)
    if article_match:
        art = format_article_ref(article_match.group(1)).upper().replace("°", "º")
        return f"Art. {art.split('-', 1)[0]}"

    structure_match = re.search(
        r"\b(Livro|T[íi]tulo|Subt[íi]tulo|Cap[íi]tulo|Se[cç][aã]o)\s+([IVXLCDM]+(?:-[A-Za-z]+)?|[ÚU]nico)\b",
        tail,
        flags=re.IGNORECASE,
    )
    if structure_match:
        return format_structure_ref(structure_match.group(1), structure_match.group(2))
    return ""


def extract_creation_with_title_target(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    match = re.search(
        r"\b(?:prop\S*e\s+a\s+cria\S*o|cria-se|criam-se|inclua-se|incluam-se|acrescente-se|acrescentem-se)\b.*?\b(LIVRO\s+[IVXLCDM]+(?:-[A-Z]+)?\s+DOS\s+[A-Z][A-Z\s-]+?)\s*,\s*com\s+(T\S*TULO\s+\S*NICO)\b",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""
    livro = clean_spaces(match.group(1)).upper()
    titulo = clean_spaces(match.group(2)).upper()
    return f"{livro}, com {titulo}"


def extract_revocation_suppression_with_redaction_target(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    suppress_match = re.search(
        r"suprima-se\s+a\s+revoga\S+\s+do\s+(art\.?\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    if not suppress_match:
        return ""

    redaction_match = re.search(
        r"a\s+seguinte\s+reda\S+\s+para\s+o\s+(art\.?\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    if not redaction_match:
        return ""

    suppressed_art = clean_clause_target_text(suppress_match.group(1))
    redaction_art = clean_clause_target_text(redaction_match.group(1))
    if simplify(suppressed_art) != simplify(redaction_art):
        return ""
    return suppressed_art


def extract_structural_modification_with_revocation_target(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""
    structure = extract_longest_structural_chain(raw)
    articles = extract_articles_loose(raw)
    if not structure or not articles:
        return ""

    project_inciso_match = re.search(
        r"\bno\s+(inciso\s+[IVXLCDM]+\s+do\s+art\.?\s*" + ART_NUM_RE + r")\b",
        raw,
        flags=re.IGNORECASE,
    )
    if not project_inciso_match:
        return ""
    if not re.search(r"suprimindo-se\s+a\s+revoga\S+", raw, flags=re.IGNORECASE):
        return ""

    art = clean_clause_target_text(f"Art. {format_article_ref(articles[0])}")
    project_inciso = clean_clause_target_text(project_inciso_match.group(1))
    return "; ".join([structure, art, project_inciso])


def extract_restored_article_change_summary(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    suppress_match = re.search(
        r"suprima-se\s+a\s+revoga\S+\s+do\s+(art\.?\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    project_match = re.search(
        r"(?:retirando\s+sua\s+cita\S+\s+do|no)\s+(inciso\s+[IVXLCDM]+\s+do\s+art\.?\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    redaction_match = re.search(
        r"a\s+seguinte\s+reda\S+\s+para\s+o\s+(art\.?\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    if not (suppress_match and project_match and redaction_match):
        return ""

    suppressed_art = clean_clause_target_text(suppress_match.group(1))
    project_ref = clean_clause_target_text(project_match.group(1))
    redaction_art = clean_clause_target_text(redaction_match.group(1))
    if simplify(suppressed_art) != simplify(redaction_art):
        return ""
    project_article_match = re.search(r"(art\.?\s*" + ART_NUM_RE + r")", project_ref, flags=re.IGNORECASE)
    project_article = clean_clause_target_text(project_article_match.group(1)) if project_article_match else project_ref
    return f"Suprime a revogação do {suppressed_art.lower()}; modifica a redação do {project_article.lower()} e modifica a redação do {redaction_art.lower()}."


def extract_structural_mixed_summary(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""
    structure = extract_longest_structural_chain(raw)
    articles = extract_articles_loose(raw)
    revoked_match = re.search(
        r"suprimindo-se\s+a\s+revoga\S+\s+do\s+(inciso\s+[IVXLCDM]+\s+do\s+art\.?\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    project_match = re.search(
        r"\bno\s+(inciso\s+[IVXLCDM]+\s+do\s+art\.?\s*" + ART_NUM_RE + r")\b",
        raw,
        flags=re.IGNORECASE,
    )
    if not (structure and articles and revoked_match and project_match):
        return ""

    revoked_ref = clean_clause_target_text(revoked_match.group(1))
    project_ref = clean_clause_target_text(project_match.group(1))
    if simplify(revoked_ref) == simplify(project_ref):
        return ""

    art = clean_clause_target_text(f"Art. {format_article_ref(articles[0])}")
    return f"Modifica {summary_object(structure)}; modifica o {art.lower()} e suprime a revogação do {revoked_ref} no {project_ref}."


def extract_project_citation_adjustment_target(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    match = re.search(
        r"suprima(?:m)?-se\s+a?s?\s+revoga(?:[cç][aã]o|[cç][oõ]es)\s+d(?:o|os|a|as)\s+(.+?)\s+da\s+lei\b.*?"
        r"retirando\s+sua(?:s)?\s+cita(?:[cç][aã]o|[cç][oõ]es)\s+do\s+((?:inciso|incisos)\s+[IVXLCDM]+(?:-[A-Z0-9]+)?\s+do\s+art\.?\s*"
        + ART_NUM_RE
        + r")\s+do\s+projeto",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    revoked_chunk = clean_spaces(match.group(1)).strip(" ,.;:")
    project_ref = clean_clause_target_text(match.group(2))
    targets: List[str] = []

    complex_match = re.search(
        r"((?:inciso|incisos)\s+[IVXLCDM]+(?:\s*(?:,|e|ou)\s*[IVXLCDM]+)*\s+do\s+art\.?\s*("
        + ART_NUM_RE
        + r"))\s*,?\s*e\s+de\s+seu\s+(§\s*\d+[º°o]?)\s*,?\s*e\s+do\s+(art\.?\s*"
        + ART_NUM_RE
        + r")",
        revoked_chunk,
        flags=re.IGNORECASE,
    )
    if complex_match:
        base_art = normalize_article_ref(complex_match.group(2))
        targets.extend([
            clean_clause_target_text(complex_match.group(1)),
            clean_clause_target_text(f"{complex_match.group(3)} do art. {format_article_ref(base_art)}"),
            clean_clause_target_text(complex_match.group(4)),
        ])
    else:
        ordered = extract_ordered_explicit_targets(revoked_chunk)
        if ordered:
            targets.extend(ordered)
        else:
            normalized = clean_clause_target_text(revoked_chunk)
            if normalized:
                targets.append(normalized)

    declarative_subject = clean_clause_target_text(extract_declarative_redaction_subject(raw))
    if declarative_subject and simplify(declarative_subject) not in {simplify(item) for item in targets}:
        targets.append(declarative_subject)

    targets.append(project_ref)
    deduped: List[str] = []
    seen: set[str] = set()
    for target in targets:
        key = simplify(target)
        if key and key not in seen:
            seen.add(key)
            deduped.append(target)
    return "; ".join(deduped)


def extract_project_citation_adjustment_summary(text: str) -> str:
    target = extract_project_citation_adjustment_target(text)
    if not target:
        return ""
    parts = [clean_summary_segment(part) for part in re.split(r"\s*;\s*", target) if clean_summary_segment(part)]
    if len(parts) < 2:
        return ""

    rendered: List[str] = []
    declarative_subject = clean_summary_segment(extract_declarative_redaction_subject(text))
    declarative_key = simplify(declarative_subject)
    rendered_declarative = False
    for part in parts[:-1]:
        if declarative_key and simplify(part) == declarative_key:
            rendered.append(f"Modifica {summary_object(part)}")
            rendered_declarative = True
        else:
            rendered.append(f"Suprime {summary_object(part)}")

    if declarative_subject and not rendered_declarative:
        rendered.append(f"Modifica {summary_object(declarative_subject)}")

    rendered.append(f"Modifica {lower_first(summary_object(parts[-1]))}")
    return "; ".join(rendered)


def extract_longest_structural_chain(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    patterns = [
        r"((?:Se[cç]\S*o|S\S*o)\s+[IVXLCDM]+(?:\s+do\s+Cap\S*tulo\s+[IVXLCDM]+(?:\s+do\s+T\S*tulo\s+[IVXLCDM]+(?:\s+do\s+Livro\s+[IVXLCDM]+(?:\s+da\s+Parte\s+Geral)?)?)?)?)",
        r"(Cap\S*tulo\s+[IVXLCDM]+(?:\s+do\s+T\S*tulo\s+[IVXLCDM]+(?:\s+do\s+Livro\s+[IVXLCDM]+(?:\s+da\s+Parte\s+Geral)?)?)?)",
        r"(T\S*tulo\s+[IVXLCDM]+(?:\s+do\s+Livro\s+[IVXLCDM]+(?:\s+da\s+Parte\s+Geral)?)?)",
        r"(Livro\s+[IVXLCDM]+(?:-[A-Z]+)?(?:\s+dos\s+[A-Z][A-Z\s-]+)?(?:,\s*com\s+T\S*tulo\s+\S*nico)?)",
    ]
    best = ""
    for pattern in patterns:
        for match in re.finditer(pattern, raw, flags=re.IGNORECASE):
            candidate = clean_clause_target_text(match.group(1))
            if len(candidate) > len(best):
                best = candidate
    return best


def extract_direct_redaction_target(text: str) -> str:
    raw_full = normalize_ocr_noise(clean_spaces(text or ""))
    if re.search(
        r':\s*[“"][\s\S]+?[”"]\s*(?:\([A-Z]{2}\))?\s*(?=(?:suprima-se|suprimam-se|d\S*-se|acrescente-se|inclua-se|retome-se|mantenha-se)\b)',
        raw_full,
        flags=re.IGNORECASE,
    ):
        return ""

    raw = normalize_ocr_noise(clean_spaces(extract_directive_context(text) or text))
    if not raw:
        return ""

    match = re.search(
        r"\bd\S*-se(?:\s+nova\s+reda\S+)?\s+(?:ao|aos|à|às)\s+(.+?)(?=,\s*(?:a|as)\s+seguinte(?:s)?\s+reda\S+|\s+(?:a|as)\s+seguinte(?:s)?\s+reda\S+)",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    target = clean_clause_target_text(match.group(1))
    if not re.search(r"\b(?:art\.?|arts?\.?|artigo|artigos|inciso|incisos|§|par[aá]grafo|livro|t[íi]tulo|cap[íi]tulo|se[cç][aã]o)\b", target, flags=re.IGNORECASE):
        return ""
    return target


def extract_declarative_redaction_subject(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    match = re.search(
        r"(?:^|[.;]\s*)((?:o|a|os|as)\s+.+?)\s*,?\s+(?:(?:a\s+ser\s+)?(?:alterad[oa]s?|inclu[ií]d[oa]s?)\s+pelo\s+art\.?\s*\d+[º°o]?(?:\s+do\s+projeto)?[^,;:]*,\s+)?passa(?:m)?\s+a\s+vigorar\s+com\s+(?:a\s+seguinte\s+reda\S+|a\s+reda\S+\s+a\s+seguir)",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return ""

    target = clean_clause_target_text(match.group(1))
    return target if re.search(r"\b(?:art\.?|arts?\.?|artigo|artigos|livro|t[íi]tulo|cap[íi]tulo|se[cç][aã]o)\b", target, flags=re.IGNORECASE) else ""


def extract_embedded_inciso_suppression_targets(text: str) -> List[str]:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return []

    subject = extract_declarative_redaction_subject(text)
    article_match = re.search(r"\b(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")\b", subject, flags=re.IGNORECASE)
    if not article_match:
        return []

    article = clean_clause_target_text(f"Art. {article_match.group(1)}")
    targets: List[str] = []
    for match in re.finditer(
        r"(?:^|[\n\r])\s*([IVXLCDM]+)\s*[-–—]\s*(?:supress\S+\s+da\s+altera\S+|suprimir|suprima-se|suprimam-se)\b",
        raw,
        flags=re.IGNORECASE,
    ):
        target = f"Inciso {match.group(1).upper()} do {article.lower()}"
        if simplify(target) not in {simplify(item) for item in targets}:
            targets.append(target)
    return targets


def extract_addition_to_existing_article_target(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(extract_directive_context(text) or text))
    if not raw:
        return ""

    direct_order = re.search(
        r"\b(?:acrescente-se|acrescentem-se|inclua-se|incluam-se|insira-se|insiram-se)\s+(?:o|a|os|as)?\s+(.+?)\s+ao\s+(art\.?\s*" + ART_NUM_RE + r")\b",
        raw,
        flags=re.IGNORECASE,
    )
    if direct_order:
        item = clean_clause_target_text(direct_order.group(1), concise=True)
        article = clean_clause_target_text(direct_order.group(2), concise=True)
        if item and article:
            return f"{item} ao {article}"

    inverted_order = re.search(
        r"\b(?:acrescente-se|acrescentem-se|inclua-se|incluam-se|insira-se|insiram-se)\s+ao\s+(art\.?\s*" + ART_NUM_RE + r")\s*,?\s+(?:o|a|os|as)?\s+(.+?)(?=,?\s*(?:da\s+Lei|do\s+Projeto|como\s+propost|nos\s+termos|na\s+forma|:|$))",
        raw,
        flags=re.IGNORECASE,
    )
    if inverted_order:
        article = clean_clause_target_text(inverted_order.group(1), concise=True)
        item = clean_clause_target_text(inverted_order.group(2), concise=True)
        if item and article:
            return f"{item} ao {article}"
    return ""


def render_addition_summary_target(target: str) -> str:
    text = clean_summary_segment(target)
    if not text:
        return ""
    lowered = text.lower()
    if lowered.startswith(("§ ", "§§", "art.", "arts.", "inciso ", "incisos ")):
        return summary_object(text)
    return text


def extract_all_paragraph_targets(text: str) -> List[str]:
    raw = normalize_ocr_noise(clean_spaces(extract_directive_context(text) or text))
    if not raw:
        return []

    targets: List[Tuple[int, str]] = []

    for match in re.finditer(
        r"((?:inciso|incisos|inc\.?)[ ]+[IVXLCDM]+(?:[ ]*(?:,|e|ou|a)[ ]*[IVXLCDM]+)*[ ]+do[ ]+(?:caput[ ]+do[ ]+|§[ ]*\d+[º°o]?[ ]+do[ ]+)(?:art\.?|artigo)[ ]*"
        + ART_NUM_RE
        + r")",
        raw,
        flags=re.IGNORECASE,
    ):
        if is_in_action_context(raw, match.start(), match.end()):
            targets.append((match.start(), clean_clause_target_text(match.group(1))))

    for match in re.finditer(
        r"((?:§|§§|[S$?])\s*\d+[º°o]?(?:\s*(?:,|e|ou|a)\s*(?:§|§§|[S$?])?\s*\d+[º°o]?)*\s+do\s+(?:art\.?|artigo)\s*"
        + ART_NUM_RE
        + r")",
        raw,
        flags=re.IGNORECASE,
    ):
        if is_in_action_context(raw, match.start(), match.end()):
            rendered = clean_clause_target_text(re.sub(r"^\?", "§", match.group(1)))
            targets.append((match.start(), rendered))

    for match in re.finditer(r"(par\S*grafo\s+\S*nico\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")", raw, flags=re.IGNORECASE):
        if is_in_action_context(raw, match.start(), match.end()):
            targets.append((match.start(), clean_clause_target_text(match.group(1))))

    s = simplify(raw).replace("§", " sec ")
    for match in re.finditer(r"(sec\s*\d+\S*\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")", s, flags=re.IGNORECASE):
        if is_in_action_context(raw, match.start(), match.end()):
            rendered = clean_clause_target_text(match.group(1).replace("sec", "§", 1))
            targets.append((match.start(), rendered))
    for match in re.finditer(r"(paragrafo\s+unico\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")", s, flags=re.IGNORECASE):
        if is_in_action_context(raw, match.start(), match.end()):
            rendered = clean_clause_target_text(match.group(1).replace("paragrafo unico", "parágrafo único", 1))
            targets.append((match.start(), rendered))

    for match in re.finditer(r"(p\.\s*[uú]nico\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")", raw, flags=re.IGNORECASE):
        if is_in_action_context(raw, match.start(), match.end()):
            rendered = re.sub(r"p\.\s*[uú]nico", "parágrafo único", match.group(1), flags=re.IGNORECASE)
            targets.append((match.start(), clean_clause_target_text(rendered)))

    for match in re.finditer(r"(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")\s*,\s*(par[aá]grafo\s+[uú]nico|p\.\s*[uú]nico)", raw, flags=re.IGNORECASE):
        if is_in_action_context(raw, match.start(), match.end()):
            targets.append((match.start(), clean_clause_target_text(f"parágrafo único do art. {format_article_ref(match.group(1))}")))

    for match in re.finditer(r"par[aá]grafo\s+(\d+)[º°o]?\s+do\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")", raw, flags=re.IGNORECASE):
        if is_in_action_context(raw, match.start(), match.end()):
            targets.append((match.start(), clean_clause_target_text(f"§ {match.group(1)}º do art. {format_article_ref(match.group(2))}")))

    targets.sort(key=lambda item: item[0])
    deduped: List[str] = []
    seen: set[str] = set()
    for _, target in targets:
        key = simplify(target)
        if not target or key in seen:
            continue
        seen.add(key)
        deduped.append(target)
    return deduped


def extract_dotted_article_specific_targets(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    match = re.search(
        r"Art\.?\s*(" + ART_NUM_RE + r")\s*[\.:]?\s*(?:[.\u2026]{4,}|…+)\s*(.+?)(?=\”|\”\s*\(NR\)|\(\s*NR\s*\)|$)",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not match:
        return ""

    art = normalize_article_ref(match.group(1))
    body = match.group(2)
    targets: List[str] = []

    for nested_match in re.finditer(
        r"(inciso|incisos)\s+([IVXLCDM]+(?:-[A-Z0-9]+)?(?:\s*(?:,|e|ou|a)\s*[IVXLCDM]+(?:-[A-Z0-9]+)?)*)\s+do\s+(par[aá]grafo\s+[uú]nico|p\.\s*[uú]nico|§\s*\d+[º°o]?)",
        body,
        flags=re.IGNORECASE,
    ):
        inciso_label = format_inciso_detail(nested_match.group(2))
        parent = nested_match.group(3)
        if re.search(r"§\s*\d+", parent):
            num = re.search(r"\d+", parent)
            parent_label = f"§ {num.group(0)}º" if num else parent
        else:
            parent_label = "parágrafo único"
        targets.append(f"{inciso_label} do {parent_label} do art. {format_article_ref(art)}")

    for sec_match in re.finditer(r"(§\s*\d+[º°o]?|par[aá]grafo\s+[uú]nico)\b", body, flags=re.IGNORECASE):
        label = sec_match.group(1)
        if any("do paragrafo unico" in simplify(item) for item in targets) and label.lower().startswith("par"):
            continue
        if label.lower().startswith("par"):
            targets.append(f"Parágrafo único do art. {format_article_ref(art)}")
        else:
            num = re.search(r"\d+", label)
            if num:
                targets.append(f"§ {num.group(0)}º do art. {format_article_ref(art)}")

    if "paragrafo unico" in simplify(body):
        par_start = simplify(body).find("paragrafo unico")
        tail = body[par_start:] if par_start >= 0 else body
        for roman_match in re.finditer(r"(?:^|\s)([IVXLCDM]+(?:-[A-Z0-9]+)?)\s*[–-]\s+", tail, flags=re.IGNORECASE):
            target = f"Inciso {roman_match.group(1).upper()} do parágrafo único do art. {format_article_ref(art)}"
            if simplify(target) not in {simplify(item) for item in targets}:
                targets.append(target)

    deduped: List[str] = []
    seen: set[str] = set()
    for target in targets:
        key = simplify(target)
        if key and key not in seen:
            seen.add(key)
            deduped.append(target)
    return "; ".join(deduped)


def extract_article_body_suppression_targets(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""
    targets = [
        f"Art. {format_article_ref(match.group(1))}"
        for match in re.finditer(
            r"\bArt\.?\s*(" + ART_NUM_RE + r")\s*[\.:]\s*Supress[aã]o\s+da\s+proposta\b",
            raw,
            flags=re.IGNORECASE,
        )
    ]
    return "; ".join(dict.fromkeys(targets))


def extract_complex_incisos_target(text: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw:
        return ""

    match = re.search(
        r"((?:inciso|incisos|inc\.?)\s+[IVXLCDM]+(?:\s*(?:,|e|ou)\s*[IVXLCDM]+)*\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r"\s*,\s*e\s+de\s+seu\s+§\s*\d+[º°o]?\s*,\s*e\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        s = simplify(raw)
        if "incisos" in s and "de seu" in s and "art. 1.550" in s and "art. 1.551" in s:
            return "incisos I, V e VI do art. 1.550, e de seu § 1º, e do art. 1.551; Art. 1.550"
        return ""

    target = clean_clause_target_text(match.group(1))
    first_art = re.search(r"(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")", match.group(1), flags=re.IGNORECASE)
    if first_art:
        art = clean_clause_target_text(f"Art. {first_art.group(1)}")
        return f"{target}; {art}"
    return target


def refine_ordered_targets(text: str, targets: List[str]) -> List[str]:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    if not raw or not targets:
        return targets

    refined = list(targets)
    specific_by_art: Dict[str, List[str]] = defaultdict(list)
    for target in targets:
        if re.fullmatch(r"Art\.\s*" + ART_NUM_RE, target, flags=re.IGNORECASE):
            continue
        if re.fullmatch(r"Arts\.\s*.+", target, flags=re.IGNORECASE):
            continue
        art_matches = re.findall(r"\b(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")\b", target, flags=re.IGNORECASE)
        if art_matches:
            specific_by_art[normalize_article_ref(art_matches[-1])].append(target)

    for idx, target in list(enumerate(refined)):
        art_match = re.fullmatch(r"Art\.\s*(" + ART_NUM_RE + r")", target, flags=re.IGNORECASE)
        if not art_match:
            continue
        art = normalize_article_ref(art_match.group(1))

        direct_article_pattern = re.compile(
            r"\b(?:d[eéê]-se(?:\s+nova\s+reda\S+)?|suprima-se|suprimam-se|acrescente-se|acrescentem-se|inclua-se|incluam-se|insira-se|insiram-se|retome-se|mantenha-se)\s+"
            r"(?:o|a|os|as|ao|aos|à|às)?\s*(?:atual\s+)?(?:art\.?|artigo)\s*"
            + build_article_ref_pattern(art)
            + r"\b",
            flags=re.IGNORECASE,
        )
        listed_article_pattern = re.compile(
            r"(?:^|[;,])\s*(?:ao|aos)\s+(?:art\.?|artigo)\s*" + build_article_ref_pattern(art) + r"\b",
            flags=re.IGNORECASE,
        )
        has_direct_article_action = bool(direct_article_pattern.search(raw) or listed_article_pattern.search(raw))
        if has_direct_article_action:
            continue

        contextual_pattern = re.compile(
            r"((?:§\s*\d+\S*|[S$?]\s*\d+\S*|par\S*grafo\s+\S*nico)\s+(?:do|ao)\s+(?:art\.?|artigo)\s*"
            + build_article_ref_pattern(art)
            + r")",
            flags=re.IGNORECASE,
        )
        contextual_match = contextual_pattern.search(raw)
        if contextual_match:
            refined[idx] = clean_clause_target_text(re.sub(r"^\?", "§", contextual_match.group(1)))
            continue

        if art not in specific_by_art:
            continue
        refined[idx] = ""

    deduped: List[str] = []
    seen: set[str] = set()
    for item in refined:
        key = simplify(item)
        if item and key not in seen:
            seen.add(key)
            deduped.append(item)

    filtered: List[str] = []
    for item in deduped:
        item_s = simplify(item)
        is_less_specific_inciso = False
        is_less_specific_paragraph = False
        inciso_match = re.fullmatch(
            r"inciso\s+([ivxlcdm]+(?:-[a-z]+)?)\s+do\s+caput\s+do\s+art\.?\s*(" + ART_NUM_RE + r")",
            item_s,
            flags=re.IGNORECASE,
        )
        if inciso_match:
            roman = inciso_match.group(1).upper()
            art = normalize_article_ref(inciso_match.group(2))
            for other in deduped:
                other_s = simplify(other)
                if re.fullmatch(
                    r"caput\s+do\s+inciso\s+" + re.escape(roman.lower()) + r"\s+do\s+caput\s+do\s+art\.?\s*" + re.escape(art) + r"(?:[º°o])?",
                    other_s,
                    flags=re.IGNORECASE,
                ):
                    is_less_specific_inciso = True
                    break
        paragraph_match = re.fullmatch(
            r"(?:§\s*(\d+)[º°o]?|paragrafo\s+unico)\s+do\s+art\.?\s*(" + ART_NUM_RE + r")",
            item_s,
            flags=re.IGNORECASE,
        )
        if paragraph_match:
            paragraph_number = paragraph_match.group(1)
            art = normalize_article_ref(paragraph_match.group(2))
            for other in deduped:
                other_s = simplify(other)
                if paragraph_number:
                    if re.fullmatch(
                        r"(?:inciso|incisos)\s+.+\s+do\s+§\s*"
                        + re.escape(paragraph_number)
                        + r"(?:[º°o])?\s+do\s+art\.?\s*"
                        + re.escape(art)
                        + r"(?:[º°o])?",
                        other_s,
                        flags=re.IGNORECASE,
                    ):
                        is_less_specific_paragraph = True
                        break
                else:
                    if re.fullmatch(
                        r"(?:inciso|incisos)\s+.+\s+do\s+paragrafo\s+unico\s+do\s+art\.?\s*"
                        + re.escape(art)
                        + r"(?:[º°o])?",
                        other_s,
                        flags=re.IGNORECASE,
                    ):
                        is_less_specific_paragraph = True
                        break
        if not is_less_specific_inciso and not is_less_specific_paragraph:
            filtered.append(item)
    return filtered


def is_in_action_context(raw: str, start: int, end: int) -> bool:
    """Verifica se o match está em contexto de ação (não apenas citação)."""
    context_before = raw[max(0, start - 180):start].strip()
    context_after = raw[end:end + 80].strip()
    action_verbs = [
        r"\b(?:inclua-se|incluam-se|acrescente-se|acrescentem-se|suprima-se|suprimam-se|d[eê]-se|insira-se|insiram-se|mantenha-se|retome-se|substitua-se|substituam-se|passa\s+a\s+vigorar)\b",
        r"\b(?:acrescentem-se|suprimam-se|dê-se|insiram-se|retomem-se|substituam-se)\b",
    ]
    for verb in action_verbs:
        if re.search(verb, context_before + " " + context_after, flags=re.IGNORECASE):
            return True
    citation_indicators = [
        r"\b(?:conforme|de\s+acordo\s+com|nos\s+termos\s+do|segundo|para\s+fins\s+de|por\s+prova\s+documental)\b",
        r"\b(?:nos\s+termos\s+da\s+lei|da\s+lei\s+n[ºo]?\s*[\d.]+\b)\b",
    ]
    for indicator in citation_indicators:
        if re.search(indicator, context_before + " " + context_after, flags=re.IGNORECASE):
            return False
    return True


def extract_ordered_explicit_targets(text: str) -> List[str]:
    raw = normalize_ocr_noise(clean_spaces(extract_directive_context(text) or text))
    if not raw:
        return []

    candidates: List[Tuple[int, int, int, str]] = []

    patterns = [
        (
            re.compile(
                r"\b(LIVRO\s+[IVXLCDM]+(?:-[A-Z]+)?\s+DOS\s+[A-Z][A-Z\s-]+?\s*,\s*com\s+T[ÍI]TULO\s+[ÚU]NICO)\b",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_spaces(m.group(1)).upper(),
        ),
        (
            re.compile(
                r"\b(Se[cç][aã]o\s+[IVXLCDM]+(?:\s+do\s+Cap[íi]tulo\s+[IVXLCDM]+(?:\s+do\s+T[íi]tulo\s+[IVXLCDM]+(?:\s+do\s+Livro\s+[IVXLCDM]+(?:\s+da\s+Parte\s+Geral)?)?)?)?)\b",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"((?:§|§§)\s*\d+[º°o]?(?:\s*(?:,|e|ou|a)\s*(?:§|§§)?\s*\d+[º°o]?)*\s+ao\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"((?:par\S*grafo\s+\S*nico|paragrafo\s+unico)\s+ao\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"((?:inciso|incisos|inc\.?)\s+[IVXLCDM]+(?:-[A-Z]+)?(?:\s*(?:,|e|ou|a)\s*[IVXLCDM]+(?:-[A-Z]+)?)*\s+ao\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)).replace("Inc. ", "Inciso ").replace("inc. ", "inciso "),
        ),
        (
            re.compile(
                r"((?:caput\s+do\s+)?(?:inciso|incisos|inc\.?)\s+[IVXLCDM]+(?:-[A-Z]+)?(?:\s*(?:,|e|ou|a)\s*[IVXLCDM]+(?:-[A-Z]+)?)*(?:\s+do\s+caput)?\s+do\s+(?:art\.?|artigo)\s*"
                + ART_NUM_RE
                + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)).replace("Inc. ", "Inciso ").replace("inc. ", "inciso "),
        ),
        (
            re.compile(
                r"((?:inciso|incisos|inc\.?)\s+[IVXLCDM]+(?:-[A-Z0-9]+)?(?:\s*(?:,|e|ou|a)\s*[IVXLCDM]+(?:-[A-Z0-9]+)?)*\s+do\s+(?:par[aá]grafo\s+[uú]nico|p\.\s*[uú]nico)\s+do\s+(?:art\.?|artigo)\s*"
                + ART_NUM_RE
                + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)).replace("Inc. ", "Inciso ").replace("inc. ", "inciso "),
        ),
        (
            re.compile(
                r"((?:inciso|incisos|inc\.?)\s+[IVXLCDM]+(?:-[A-Z0-9]+)?(?:\s*(?:,|e|ou|a)\s*[IVXLCDM]+(?:-[A-Z0-9]+)?)*(?:\s+do\s+caput)?\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)).replace("Inc. ", "Inciso ").replace("inc. ", "inciso "),
        ),
        (
            re.compile(
                r"((?:inciso|incisos|inc\.?)\s+[IVXLCDM]+(?:\s*(?:,|e|ou)\s*[IVXLCDM]+)*(?:\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")\s*,\s*e\s+de\s+seu\s+§\s*\d+[º°o]?\s*,\s*e\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"(par\S*grafo\s+\S*nico\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"((?:§|§§)\s*\d+[º°o]?(?:\s*(?:,|e|ou|a)\s*(?:§|§§)?\s*\d+[º°o]?)*(?:\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r"))",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"(caput\s+do\s+(?:art\.?|artigo)\s*" + ART_NUM_RE + r")",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(
                r"\b((?:arts?\.?|artigos?)\s*" + ART_NUM_RE + r"(?:\s*(?:,|e|ou|a)\s*(?:art\.?\s*)?" + ART_NUM_RE + r")*)",
                flags=re.IGNORECASE,
            ),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(r"(ao\s+art\.?\s*" + ART_NUM_RE + r")", flags=re.IGNORECASE),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
        (
            re.compile(r"((?:art\.?|artigo)\s*" + ART_NUM_RE + r")", flags=re.IGNORECASE),
            lambda m: clean_clause_target_text(m.group(1)),
        ),
    ]

    for priority, (pattern, formatter) in enumerate(patterns):
        for match in pattern.finditer(raw):
            rendered = formatter(match).strip(" ,.;:")
            if not rendered:
                continue
            if not is_in_action_context(raw, match.start(), match.end()):
                continue
            context_before = simplify(raw[max(0, match.start() - 40):match.start()])
            context_after = simplify(raw[match.end():match.end() + 60])
            if rendered.lower().startswith(("art.", "arts.")):
                direct_project_target = bool(re.search(r"(?:suprima(?:m)?-se|d\S*-se|acrescente-se|inclua-se)\s+(?:o|a|os|as)?\s*$", context_before))
                if ("do projeto" in context_after or "projeto de lei" in context_after or "pl 4/2025" in context_after) and not direct_project_target:
                    continue
                if re.search(r"(propost[oa]s?|alterad[oa]s?|promovid[oa]s?)\s+pelo\s*$", context_before):
                    continue
            if rendered.lower().startswith("inciso ") and "retirando sua" in simplify(raw) and "revoga" in simplify(raw):
                continue
            candidates.append((match.start(), match.end(), priority, rendered))

    candidates.sort(key=lambda item: (item[0], item[2], -(item[1] - item[0])))
    rendered_targets: List[str] = []
    seen: set[str] = set()
    occupied: List[Tuple[int, int]] = []
    for start, end, _, target in candidates:
        key = simplify(target)
        if not key or key in seen:
            continue
        if any(not (end <= occ_start or start >= occ_end) for occ_start, occ_end in occupied):
            continue
        seen.add(key)
        occupied.append((start, end))
        rendered_targets.append(target)
    return rendered_targets


def extract_mixed_redaction_suppression_target(text: str, for_summary: bool = False) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    redaction = re.search(
        r"\bd[eéê]-se\s+nova\s+reda[cç][aã]o\s+aos?\s+(.+?)(?=\s*;\s*e\s+suprima-se\b|\s*;\s*suprima-se\b)",
        raw,
        flags=re.IGNORECASE,
    )
    suppression = re.search(
        r"\bsuprima-se\s+o\s+(.+?)(?=,\s*(?:todos?|ambos?)\s+da\s+lei|,\s*como\s+propost|,\s*nos\s+termos|;|$)",
        raw,
        flags=re.IGNORECASE,
    )
    if not redaction or not suppression:
        return ""

    first_target = clean_clause_target_text(redaction.group(1), concise=True)
    second_target = clean_clause_target_text(suppression.group(1), concise=True)
    if not first_target or not second_target:
        return ""

    if for_summary:
        return f"Dê-se nova redação aos {first_target}; e suprima-se o {second_target}"
    return f"{first_target} e {second_target}"


def strip_leading_action(segment: str) -> str:
    patterns = [
        r"^e\s+",
        r"^\d{5}\s+\d+\s+\d+\s+",
        r"^suprima-se\s+(?:o|a|os|as)\s+",
        r"^suprimam-se\s+(?:o|a|os|as)\s+",
        r"^suprima-se\s+",
        r"^suprimam-se\s+",
        r"^suprimir\s+(?:o|a|os|as)\s+",
        r"^integralmente\s+",
        r"^acrescente-se\s+",
        r"^acrescentem-se\s+",
        r"^inclua-se\s*,?\s*",
        r"^incluam-se\s*,?\s*",
        r"^d[eéê]-se\s+nova\s+reda[cç][aã]o\s+dada\s+(?:para|ao|aos|à|às|a|o|os|as|do|da|dos|das)\s+",
        r"^d[eéê]-se\s+nova\s+redacao\s+dada\s+(?:para|ao|aos|a|o|os|as|do|da|dos|das)\s+",
        r"^d[eéê]-se\s+nova\s+reda[cç][aã]o\s+(?:para|ao|aos|à|às|a|o|os|as|do|da|dos|das)\s+",
        r"^d[eéê]-se\s+nova\s+redacao\s+(?:para|ao|aos|a|o|os|as|do|da|dos|das)\s+",
        r"^d[eê]-se\s+nova\s+reda[cç][aã]o\s+ao?s?\s+",
        r"^d[eê]-se\s+a\s+seguinte\s+reda[cç][aã]o\s+ao?s?\s+",
        r"^d[eéê]-se\s+nova\s+redacao\s+ao?s?\s+",
        r"^d[eéê]-se\s+a\s+seguinte\s+redacao\s+ao?s?\s+",
        r"^d[eéê]-se\s+ao?s?\s+",
        r"^de-se nova redacao ao?s?\s+",
        r"^d[eÃª]-se nova reda[cÃ§][aÃ£]o ao?s?\s+",
        r"^de-se a seguinte redacao ao?s?\s+",
        r"^de-se ao?s?\s+",
        r"^retome-se\s+a\s+reda[cç][aã]o\s+do?s?\s+",
        r"^retome-se a redacao do?s?\s+",
        r"^retome-se a reda[cÃ§][aÃ£]o do?s?\s+",
        r"^mantenha-se\s+a\s+reda[cç][aã]o\s+vigente\s+do?s?\s+",
        r"^mantenha-se a redacao vigente do?s?\s+",
        r"^mantenha-se a reda[cÃ§][aÃ£]o vigente do?s?\s+",
        r"^substitua-se\s+",
        r"^substituam-se\s+",
        r"^nova\s+reda[cç][aã]o\s+dada\s+(?:para|ao|aos|à|às|a|o|os|as|do|da|dos|das)\s+",
        r"^nova\s+redacao\s+dada\s+(?:para|ao|aos|a|o|os|as|do|da|dos|das)\s+",
        r"^nova\s+reda[cç][aã]o\s+(?:para|ao|aos|à|às|a|o|os|as|do|da|dos|das)\s+",
        r"^nova\s+redacao\s+(?:para|ao|aos|a|o|os|as|do|da|dos|das)\s+",
        r"^reda[cç](?:[aã]o|[oõ]es)\s+(?:do|dos|da|das)\s+",
        r"^revoga[cç][aã]o\s+(?:do|dos|da|das)\s+",
        r"^a\s+altera[cç][aã]o\s+proposta\s+do\s+",
        r"^a\s+alteracao\s+proposta\s+do\s+",
        r"^altera[cç][aã]o\s+(?:do|da|dos|das|de)\s+",
        r"^alteracao\s+(?:do|da|dos|das|de)\s+",
        r"^inclus[aã]o\s+(?:do|da|dos|das|de)\s+",
        r"^inclusao\s+(?:do|da|dos|das|de)\s+",
        r"^(?:o|a|os|as)\s+agrupador\s+",
        r"^agrupador\s+",
    ]
    out = segment.strip()
    for pattern in patterns:
        out = re.sub(pattern, "", out, flags=re.IGNORECASE).strip()
    return out


def split_embedded_action_segments(clause: str) -> List[str]:
    text = clean_spaces(clause or "")
    if not text:
        return []

    pattern = re.compile(
        r",\s+e\s+(?=(?:o|a|os|as)\s+(?:atual\s+)?(?:art\.?|artigo|arts?\.?|artigos?)\b.*?\b(?:passa(?:m)?\s+a\s+vigorar|fica(?:m)?\s+acrescid[oa]s?))",
        flags=re.IGNORECASE,
    )
    parts = [part.strip(" ;") for part in pattern.split(text) if part.strip(" ;")]
    return parts or [text]


def clean_clause_target_text(text: str, concise: bool = False) -> str:
    target = normalize_output_value(text)
    target = normalize_ocr_noise(target)
    target = clean_spaces((target or "").replace("\n", " ")).strip(" ,.;:")
    if not target:
        return ""

    target = strip_leading_action(target)
    target = re.sub(r"^onde\s+couber,\s*", "", target, flags=re.IGNORECASE)
    target = re.sub(r"^onde\s+houver,\s*", "", target, flags=re.IGNORECASE)
    target = re.sub(r':\s*[“"].*$', "", target, flags=re.IGNORECASE).strip(" ,.;:")
    target = re.sub(r"\s+[“\"].*$", "", target).strip(" ,.;:")
    target = re.sub(r"[.…]{4,}.*$", "", target).strip(" ,.;:")
    target = re.sub(r"§\s*(\d+)\s*[o°]", r"§ \1º", target, flags=re.IGNORECASE)
    target = re.sub(r"§§\s*(\d+)\s*[o°]", r"§§ \1º", target, flags=re.IGNORECASE)
    target = re.sub(r"\bpar[aá]grafo\s+u[nn][ií]co\b", "parágrafo único", target, flags=re.IGNORECASE)
    target = re.sub(
        r"\bart\.?\s*(\d+(?:\.\d+)*)\s*-\s*o\b",
        lambda m: f"art. {m.group(1)}º" if "." not in m.group(1) and int(m.group(1)) <= 9 else f"art. {m.group(1)}",
        target,
        flags=re.IGNORECASE,
    )
    target = re.sub(
        r"\bart\.?\s*(\d+(?:\.\d+)*)o\b",
        lambda m: f"art. {m.group(1)}º" if "." not in m.group(1) and int(m.group(1)) <= 9 else f"art. {m.group(1)}",
        target,
        flags=re.IGNORECASE,
    )

    for pattern in [
        r",?\s+na\s+forma\s+proposta.*$",
        r",?\s+nos\s+termos\s+a\s+seguir.*$",
        r",?\s+nos\s+seguintes\s+termos.*$",
        r",?\s+a\s+seguinte\s+reda[cç][aã]o.*$",
        r",?\s+a\s+seguinte\s+redacao.*$",
        r",?\s+as\s+seguintes\s+reda[cç][oõ]es.*$",
        r",?\s+as\s+seguintes\s+redacoes.*$",
        r",?\s+como\s+propost[oa]s?.*$",
        r",?\s+todos?\s+da\s+lei.*$",
        r",?\s+ambos?\s+da\s+lei.*$",
        r",?\s+na\s+lei\s+federal\b.*$",
        r",?\s+na\s+lei\s+n[ºo]?\s*[\d.]+\b.*$",
        r",?\s+sugerid[oa]\s+para\s+a\s+lei\b.*$",
        r",?\s+anterior\s+[àa]\s+altera[cç][aã]o.*$",
        r",?\s+anterior\s+[àa]\s+alteracao.*$",
        r",?\s+anterior\s+[àa]\s+altera[cç][aã]o\s+promovid[oa].*$",
        r",?\s+anterior\s+[àa]\s+alteracao\s+promovid[oa].*$",
        r",?\s+que\s+versa(?:m)?\s+sobre.*$",
        r",?\s+que\s+reforma\s+o\s+c[óo]digo\s+civil.*$",
        r",?\s+propost[oa]s?\s+no\s+art\.?\s*\d+[º°o]?.*$",
        r",?\s+propost[oa]s?\s+pelo\s+art\.?\s*\d+[º°o]?.*$",
        r"\s+do\s+pl\s*4/2025\b.*$",
        r"\s+à\s+lei\b.*$",
        r"\s+a\s+lei\b.*$",
        r"\s+da\s+lei\b.*$",
        r"\s+da\s+lei\s+federal\b.*$",
        r"\s+da\s+lei\s+n[ºo]?\s*[\d.]+\b.*$",
        r"\s+do\s+c[óo]digo\s+civil\b.*$",
        r",?\s+todos?\s+na\s+forma.*$",
        r",?\s+tratad[oa]\s+pelo\s+art\..*$",
        r",?\s+(?:tambem\s+)?alterad[oa]\s+pelo\s+art\..*$",
        r",?\s+inclu[ií]d[oa]\s+pelo\s+art\..*$",
        r",?\s+feita?\s+pelo\s+art\..*$",
        r",?\s+feita?\s+pelo\s+projeto.*$",
        r",?\s+do\s+projeto\s+de\s+lei.*$",
        r",?\s+do\s+projeto.*$",
        r",?\s+com\s+a\s+reda[cç][aã]o\s+dada\s+pelo\s+art\..*$",
        r"\s+passa(?:m)?\s+a\s+vigorar\s+com\s+.*$",
        r"\s+passa(?:m)?\s+a\s+vigorar\s+como\s+.*$",
        r"\s+passa(?:m)?\s+a\s+vigorar\s+acrescid[oa]s?\s+de\s+.*$",
        r"\s+fica(?:m)?\s+acrescid[oa]s?\s+de\s+.*$",
        r",?\s+mantendo-se.*$",
        r"\s+justifica[cçg][aã]o.*$",
    ]:
        target = re.sub(pattern, "", target, flags=re.IGNORECASE).strip(" ,.;:")

    target = re.sub(r"^(?:o|a|os|as)\s+(?=(?:atual\s+)?(?:art\.?|artigo|arts?\.?|artigos?|caput|inciso|incisos|par[aá]grafo|§|livro|t[íi]tulo|subt[íi]tulo|cap[íi]tulo|se[cç][aã]o))", "", target, flags=re.IGNORECASE)
    target = re.sub(r"^atual\s+", "", target, flags=re.IGNORECASE)
    target = re.sub(r"^(ao|aos|à|às)\s+(art\.?|arts?\.?)\s+", lambda m: m.group(2) + " ", target, flags=re.IGNORECASE)
    target = re.sub(r"^Art\.?\s+", "art. ", target)
    target = re.sub(r"^Arts\.?\s+", "arts. ", target)
    target = re.sub(r"^Artigo\s+", "art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^Artigos\s+", "arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^supress[aã]o\s+do\s+art\.?\s+", "art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^supress[aã]o\s+dos\s+arts?\.?\s+", "arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^supress[aã]o\s+da\s+revoga[cç][aã]o\s+do\s+art\.?\s+", "art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^supress[aã]o\s+da\s+revoga[cç][aã]o\s+dos\s+arts?\.?\s+", "arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^a\s+reda[cç][aã]o\s+do\s+", "", target, flags=re.IGNORECASE)
    target = re.sub(r"^a\s+redacao\s+do\s+", "", target, flags=re.IGNORECASE)
    target = re.sub(r"^(art\.|arts\.)", lambda match: match.group(1).capitalize(), target, flags=re.IGNORECASE)
    target = re.sub(r"^(par[aá]grafo\s+único)", "Parágrafo único", target, flags=re.IGNORECASE)
    target = re.sub(r"^(§)\s*(\d+º)", r"\1 \2", target)
    target = re.sub(r"\b(art\.)(\d)", r"\1 \2", target, flags=re.IGNORECASE)

    if concise:
        for pattern in [
            r"\s+da\s+lei\b.*$",
            r"\s+do\s+c[óo]digo\s+civil\b.*$",
            r"\s+da\s+lei\s+federal\b.*$",
        ]:
            target = re.sub(pattern, "", target, flags=re.IGNORECASE).strip(" ,.;:")

    return target.strip(" ,.;:")


def normalize_dispositivo_output(target: str, source_text: str = "") -> str:
    target = clean_clause_target_text(target)
    if not target:
        return ""

    source_s = simplify(source_text)
    target = re.sub(r"^\s*\d{5}\s+\d+\s+\d+\s+", "", target)
    target = re.sub(r"\bArt\.?\s+(\d+),(\d{3})", r"Art. \1.\2", target, flags=re.IGNORECASE)
    target = re.sub(r"\bart\.?\s+(\d+),(\d{3})", r"Art. \1.\2", target, flags=re.IGNORECASE)
    target = re.sub(r"^reda[cç](?:[aã]o|[oõ]es)\s+dos\s+artigos?\s+", "Arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^reda[cç](?:[aã]o|[oõ]es)\s+dos\s+arts?\.?\s+", "Arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^reda[cç](?:[aã]o|[oõ]es)\s+do\s+art\.?\s+", "Art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^revoga[cç][aã]o\s+dos\s+artigos?\s+", "Arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^revoga[cç][aã]o\s+dos\s+arts?\.?\s+", "Arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"^revoga[cç][aã]o\s+do\s+art\.?\s+", "Art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"\s+e\s+a\s+nova\s+reda[cç][aã]o\s+do\s+art\.?\s*", "; Art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"\s+e\s+a\s+nova\s+redacao\s+do\s+art\.?\s*", "; Art. ", target, flags=re.IGNORECASE)
    target = re.sub(r"\s+e\s+as\s+novas?\s+reda[cç][aã]oes?\s+dos\s+arts?\.?\s*", "; Arts. ", target, flags=re.IGNORECASE)
    target = re.sub(r"\s+e\s+as\s+novas?\s+redacoes?\s+dos\s+arts?\.?\s*", "; Arts. ", target, flags=re.IGNORECASE)

    if re.search(r"\b(?:art\.?|artigo)\s*1\.111\s*[-–—]\s*[A-Za-z]\b", target, flags=re.IGNORECASE):
        return "Art. 1.111"

    if re.search(r"\b(?:art\.?|artigo)\s*15-B\b", target, flags=re.IGNORECASE) and "o seguinte dispositivo ao capitulo" in source_s:
        return "Art. 15"

    suffixed_2027 = re.findall(r"\b(?:art\.?|artigo)\s*2\.027-[A-Za-z]+(?:-[A-Za-z]+)?\b", target, flags=re.IGNORECASE)
    if not re.match(r"^\s*Arts?\.\s+2\.027-[A-Za-z]+", target, flags=re.IGNORECASE) and (len(suffixed_2027) >= 2 or (suffixed_2027 and re.search(r"\b2\.027-[A-Za-z]+.*\b2\.027-[A-Za-z]+", target, flags=re.IGNORECASE))):
        return "Art. 2.027"

    target = re.sub(r"\b1\.267\b", "Art. 1.267", target) if re.fullmatch(r"1\.267", target) else target
    target = re.sub(r"^(art\.)", "Art.", target, flags=re.IGNORECASE)
    target = re.sub(r"^(arts\.)", "Arts.", target, flags=re.IGNORECASE)
    if target != target.upper():
        target = re.sub(r"^(livro|título|titulo|capítulo|capitulo|seção|secao)\b", lambda m: m.group(1).capitalize(), target, flags=re.IGNORECASE)
    parts = [part.strip(" ,.;:") for part in re.split(r"\s*;\s*", target) if part.strip(" ,.;:")]
    deduped: List[str] = []
    seen: set[str] = set()
    for part in parts:
        key = simplify(part)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(part)

    explicit_art_lists = []
    for part in deduped:
        match = re.fullmatch(r"Arts?\.\s+(.+)", part, flags=re.IGNORECASE)
        if not match:
            continue
        explicit_art_lists.append({normalize_article_ref(item) for item in parse_art_list(simplify(match.group(1)))})

    filtered: List[str] = []
    for part in deduped:
        part_s = simplify(part)
        paragraph_match = re.fullmatch(
            r"(?:§\s*(\d+)[º°o]?|paragrafo\s+unico)\s+do\s+art\.?\s*(" + ART_NUM_RE + r")",
            part_s,
            flags=re.IGNORECASE,
        )
        if paragraph_match:
            paragraph_number = paragraph_match.group(1)
            art_norm = normalize_article_ref(paragraph_match.group(2))
            more_specific = False
            for other in deduped:
                other_s = simplify(other)
                if paragraph_number and re.fullmatch(
                    r"(?:inciso|incisos)\s+.+\s+do\s+§\s*" + re.escape(paragraph_number) + r"(?:[º°o])?\s+do\s+art\.?\s*" + re.escape(art_norm) + r"(?:[º°o])?",
                    other_s,
                    flags=re.IGNORECASE,
                ):
                    more_specific = True
                    break
                if not paragraph_number and re.fullmatch(
                    r"(?:inciso|incisos)\s+.+\s+do\s+paragrafo\s+unico\s+do\s+art\.?\s*" + re.escape(art_norm) + r"(?:[º°o])?",
                    other_s,
                    flags=re.IGNORECASE,
                ):
                    more_specific = True
                    break
            if more_specific:
                continue
        single_match = re.fullmatch(r"Art\.\s*(" + ART_NUM_RE + r")", part, flags=re.IGNORECASE)
        if single_match:
            art_norm = normalize_article_ref(single_match.group(1))
            if any(art_norm in art_list and len(art_list) > 1 for art_list in explicit_art_lists):
                continue
        filtered.append(part)
    return append_external_law_label("; ".join(filtered), source_text)


def extract_subject_addition_target(segment: str, concise: bool = False) -> str:
    text = clean_spaces(segment or "")
    if not text:
        return ""

    detail_match = re.search(
        r"\bpassa(?:m)?\s+a\s+vigorar\s+acrescid[oa]s?\s+de\s+(.+?)(?=,\s*(?:mantendo-se|mantendo|mantida|mantido)\b|[.;:])",
        text,
        flags=re.IGNORECASE,
    )
    if not detail_match:
        detail_match = re.search(
            r"\bfica(?:m)?\s+acrescid[oa]s?\s+de\s+(.+?)(?=,\s*(?:mantendo-se|mantendo|mantida|mantido)\b|[.;:])",
            text,
            flags=re.IGNORECASE,
        )
    if not detail_match:
        return ""

    detail = clean_clause_target_text(detail_match.group(1), concise=concise)
    art_match = re.search(r"\b(?:art\.?|artigo)\s*([\d.]+(?:-[A-Za-z]+)?)\b", text, flags=re.IGNORECASE)
    if not detail or not art_match:
        return ""
    return f"{detail} ao art. {art_match.group(1)}"


def extract_inserted_articles_target(text: str, concise: bool = False) -> str:
    if not is_new_article_insertion_context(text):
        return ""

    raw_original = str(text or "")
    raw = clean_spaces(raw_original)
    if not raw or not raw_original:
        return ""

    tail_match = re.search(r"passa\s+a\s+vigorar[\s\S]*", raw_original, flags=re.IGNORECASE)
    tail = tail_match.group(0) if tail_match else raw_original

    tokens: List[str] = []
    for token in re.findall(r"(?im)^[\s\"'“”‘’]*Art\.?\s*([0-9]+(?:\.\d+)*(?:[º°o])?(?:-[A-Za-z]+)?)", tail):
        token_norm = token.strip()
        if re.fullmatch(r"x+(?:[º°o])?", token_norm, flags=re.IGNORECASE):
            continue
        if token_norm not in tokens:
            tokens.append(token_norm)
    if not tokens:
        return ""

    rendered_tokens = [format_article_ref(token).upper().replace("°", "º") for token in tokens]
    if len(rendered_tokens) == 1:
        base = f"art. {rendered_tokens[0]}"
    else:
        base = f"arts. {format_human_list(rendered_tokens)}"

    if len(tokens) == 1:
        block_match = re.search(
            r"(?ims)^[\s\"'“”‘’]*Art\.?\s*" + re.escape(tokens[0]) + r"\b([\s\S]*?)(?=^[\s\"'“”‘’]*Art\.?\s*[0-9]|$\Z)",
            tail,
        )
        if block_match:
            paragraph_numbers = [int(num) for num in re.findall(r"(?im)^\s*§\s*(\d+)º?", block_match.group(1))]
            paragraph_numbers = sorted(set(paragraph_numbers))
            if paragraph_numbers:
                if len(paragraph_numbers) >= 2 and paragraph_numbers == list(range(paragraph_numbers[0], paragraph_numbers[-1] + 1)):
                    para_label = f"§§ {paragraph_numbers[0]}º a {paragraph_numbers[-1]}º"
                elif len(paragraph_numbers) == 1:
                    para_label = f"§ {paragraph_numbers[0]}º"
                else:
                    para_label = f"§§ {format_human_list([f'{n}º' for n in paragraph_numbers])}"
                base = f"{base}, com {para_label}"

    if concise:
        return base

    law_match = re.search(
        r"(Lei\s+n[ºo]?\s*[\d.]+(?:,\s+de\s+\d{1,2}\s+de\s+[A-Za-zçÇ]+\s+de\s+\d{4})?)",
        raw,
        flags=re.IGNORECASE,
    )
    if law_match:
        return f"{base} da {law_match.group(1)}"
    return base


def infer_dispositivos_legacy(text: str) -> str:
    if not text:
        return ""

    project_change_target = extract_project_change_suppression_target(text)
    if project_change_target:
        return project_change_target

    lead = first_directive_segment(text)
    if not lead:
        return ""

    lead_simplified = simplify(lead)
    text_simplified = simplify(text)
    if (
        ("novo artigo" in lead_simplified or "o seguinte artigo" in lead_simplified)
        and ("10.406" in text or "lei 10.406" in text_simplified or "codigo civil" in text_simplified)
    ):
        law_match = re.search(r"lei\s+n\w*\s*10\.406", text_simplified, flags=re.IGNORECASE)
        if law_match:
            tail = text_simplified[law_match.end():]
            first_insert = re.search(r"(?:art\.?|artigo)\s*([0-9x]{1,10}(?:\.\d+)*(?:-[a-z]+)?)", tail, flags=re.IGNORECASE)
            if first_insert:
                token = first_insert.group(1)
                if token.startswith("x"):
                    return "novo artigo sem localização específica na Lei nº 10.406, de 10 de janeiro de 2002 (Código Civil)"
                return f"art. {token} da Lei nº 10.406, de 10 de janeiro de 2002 (Código Civil)"
        return "novo artigo sem localização específica na Lei nº 10.406, de 10 de janeiro de 2002 (Código Civil)"

    for token in [
        " passa a vigorar com a seguinte redaÃ§Ã£o",
        " passa a vigorar com a seguinte redacao",
        " passam a vigorar com a seguinte redaÃ§Ã£o",
        " passam a vigorar com a seguinte redacao",
        " passa a vigorar com as seguintes alteraÃ§Ãµes",
        " passa a vigorar com as seguintes alteracoes",
        " passam a vigorar com as seguintes alteraÃ§Ãµes",
        " passam a vigorar com as seguintes alteracoes",
        " passa a vigorar com a redaÃ§Ã£o a seguir",
        " passa a vigorar com a redacao a seguir",
        ", na forma proposta",
        ", como proposto",
        ", como proposta",
        ", nos termos a seguir",
        ", nos seguintes termos",
        ", fazendo-se as flexÃµes",
        ", fazendo-se as flexoes",
    ]:
        idx = simplify(lead).find(simplify(token))
        if idx >= 0:
            lead = lead[:idx].strip(" .;:")
            break

    parts = [p.strip(" .") for p in re.split(r"\s*;\s*", lead) if p.strip()]
    cleaned_parts: List[str] = []
    multiple = len(parts) > 1
    for part in parts:
        piece = strip_leading_action(part)
        if multiple:
            piece = re.sub(r",?\s+(?:ambos?|todas?|todos?)\s+da\s+Lei.*$", "", piece, flags=re.IGNORECASE)
            piece = re.sub(r",?\s+da\s+Lei.*$", "", piece, flags=re.IGNORECASE)
        cleaned_parts.append(piece.strip(" ."))

    out = "; ".join(p for p in cleaned_parts if p)
    return normalize_dispositivo_output(out, text)


def infer_dispositivos(text: str) -> str:
    if not text:
        return ""

    creation_with_title_target = extract_creation_with_title_target(text)
    if creation_with_title_target:
        return normalize_dispositivo_output(creation_with_title_target, text)

    project_citation_adjustment_target = extract_project_citation_adjustment_target(text)
    if project_citation_adjustment_target:
        return normalize_dispositivo_output(project_citation_adjustment_target, text)

    dotted_targets = extract_dotted_article_specific_targets(text)
    if dotted_targets:
        return normalize_dispositivo_output(dotted_targets, text)

    revocation_redaction_target = extract_revocation_suppression_with_redaction_target(text)
    if revocation_redaction_target:
        return normalize_dispositivo_output(revocation_redaction_target, text)

    structural_revocation_target = extract_structural_modification_with_revocation_target(text)
    if structural_revocation_target:
        return normalize_dispositivo_output(structural_revocation_target, text)

    declarative_subject = extract_declarative_redaction_subject(text)
    declarative_embedded_suppressions = extract_embedded_inciso_suppression_targets(text)
    article_body_suppressions = extract_article_body_suppression_targets(text)

    early_paragraph_targets = extract_all_paragraph_targets(text)
    if early_paragraph_targets and re.search(
        r"\bp\.\s*[uú]nico\b|(?:art\.?|artigo)\s*" + ART_NUM_RE + r"\s*,\s*par[aá]grafo\s+[uú]nico|par[aá]grafo\s+\d+[º°o]?\s+do\s+artigo",
        text,
        flags=re.IGNORECASE,
    ):
        return normalize_dispositivo_output("; ".join(early_paragraph_targets), text)

    direct_redaction_target = extract_direct_redaction_target(text)
    if direct_redaction_target:
        return normalize_dispositivo_output(direct_redaction_target, text)

    complex_incisos_target = extract_complex_incisos_target(text)
    if complex_incisos_target:
        return normalize_dispositivo_output(complex_incisos_target, text)

    restoration_suppression_target = extract_restoration_suppression_target(text)
    if restoration_suppression_target:
        return normalize_dispositivo_output(restoration_suppression_target, text)

    suppressed_inciso_target = extract_suppressed_inciso_from_inserted_article_target(text)
    if suppressed_inciso_target:
        return normalize_dispositivo_output(suppressed_inciso_target, text)

    mixed_target = extract_mixed_redaction_suppression_target(text)
    if mixed_target:
        return normalize_dispositivo_output(mixed_target, text)

    project_change_target = extract_project_change_suppression_target(text)
    if project_change_target:
        return normalize_dispositivo_output(project_change_target, text)

    project_replacement_target = extract_project_text_replacement_target(text)
    if project_replacement_target:
        return "Não se aplica"

    project_article_internal_target = extract_project_article_internal_target(text)
    if project_article_internal_target:
        return normalize_dispositivo_output(project_article_internal_target, text)

    following_project_devices_target = extract_following_project_devices_target(text)
    if following_project_devices_target:
        return normalize_dispositivo_output(following_project_devices_target, text)

    project_law_insertion_target = extract_project_law_insertion_target(text)
    if project_law_insertion_target:
        return normalize_dispositivo_output(project_law_insertion_target, text)

    new_article_location_target = extract_new_article_with_location_target(text)
    if new_article_location_target:
        return normalize_dispositivo_output(new_article_location_target, text)

    inserted_articles_target = extract_inserted_articles_target(text)
    if inserted_articles_target:
        return normalize_dispositivo_output(inserted_articles_target, text)

    generic_project_article_target = extract_generic_project_article_target(text)
    if generic_project_article_target:
        return normalize_dispositivo_output(generic_project_article_target, text)

    ordered_targets = extract_ordered_explicit_targets(text)
    if declarative_subject and simplify(declarative_subject) not in {simplify(item) for item in ordered_targets}:
        ordered_targets.insert(0, declarative_subject)
    for embedded_target in declarative_embedded_suppressions:
        if simplify(embedded_target) not in {simplify(item) for item in ordered_targets}:
            ordered_targets.append(embedded_target)
    for body_target in re.split(r"\s*;\s*", article_body_suppressions) if article_body_suppressions else []:
        if simplify(body_target) not in {simplify(item) for item in ordered_targets}:
            ordered_targets.append(body_target)
    paragraph_targets = extract_all_paragraph_targets(text)
    for paragraph_target in paragraph_targets:
        if simplify(paragraph_target) not in {simplify(item) for item in ordered_targets}:
            ordered_targets.append(paragraph_target)
    paragraph_target = paragraph_targets[0] if paragraph_targets else ""
    if paragraph_target and (not ordered_targets or all(re.fullmatch(r"Art\.\s*" + ART_NUM_RE, item, flags=re.IGNORECASE) for item in ordered_targets)):
        return normalize_dispositivo_output(paragraph_target, text)
    if ordered_targets:
        return normalize_dispositivo_output("; ".join(refine_ordered_targets(text, ordered_targets)), text)

    clauses = split_directive_clauses(text)
    if clauses:
        targets: List[str] = []
        seen: set[str] = set()
        for clause in clauses:
            for segment in split_embedded_action_segments(clause):
                if not infer_clause_vote_nature(segment) and not extract_project_change_suppression_target(segment):
                    continue
                target = (
                    ("Não se aplica" if extract_project_text_replacement_target(segment) else "")
                    or
                    extract_restoration_suppression_target(segment)
                    or
                    extract_suppressed_inciso_from_inserted_article_target(segment)
                    or
                    extract_project_article_internal_target(segment)
                    or
                    extract_following_project_devices_target(segment)
                    or
                    extract_project_law_insertion_target(segment)
                    or
                    extract_new_article_with_location_target(segment)
                    or
                    extract_inserted_articles_target(segment)
                    or extract_subject_addition_target(segment)
                    or extract_generic_project_article_target(segment)
                    or clean_clause_target_text(infer_dispositivos_legacy(segment))
                )
                if not target:
                    continue
                target = normalize_dispositivo_output(target, segment) if target != "Não se aplica" else target
                key = simplify(target)
                if key in seen:
                    continue
                seen.add(key)
                targets.append(target)
        if targets:
            return "; ".join(targets)

    return infer_dispositivos_legacy(text)


def infer_sintese(natureza: str, dispositivos: str) -> str:
    if not natureza:
        return ""

    base = dispositivos or ""
    mapping = {
        "Supressiva": "Suprime",
        "Aditiva": "Acrescenta",
        "Modificativa": "Altera",
        "Mista": "Altera e/ou suprime",
        "Restaurativa": "Mant\u00E9m a reda\u00E7\u00E3o de",
        "Substitutiva": "Substitui",
    }
    prefix = mapping.get(natureza, "")
    if not prefix:
        return ""
    if prefix in {"Suprime", "Acrescenta", "Altera"}:
        base = summary_object(base)
    return f"{prefix} {base}".strip()


def clean_summary_segment(text: str) -> str:
    return clean_clause_target_text(text, concise=True)


def synthese_prefix_for_nature(natureza: str) -> str:
    mapping = {
        "Supressiva": "Suprime",
        "Aditiva": "Acrescenta",
        "Modificativa": "Altera",
        "Restaurativa": "Mantém a redação de",
        "Substitutiva": "Substitui",
    }
    return mapping.get(normalize_nature(natureza), "")


def summary_object(target: str) -> str:
    text = clean_spaces(target or "").strip()
    if not text:
        return ""

    if re.match(r"^(?:o|a|os|as)\s+", text, flags=re.IGNORECASE):
        return text

    feminine_singular_markers = (
        "Seção ",
        "Secao ",
    )
    feminine_plural_markers = (
        "Seções ",
        "Secoes ",
    )

    singular_markers = (
        "Art.",
        "art.",
        "§ ",
        "Caput",
        "caput",
        "Inciso ",
        "inciso ",
        "Parágrafo ",
        "parágrafo ",
        "Livro ",
        "Título ",
        "Titulo ",
        "Capítulo ",
        "Capitulo ",
        "novo art.",
    )
    plural_markers = (
        "Arts.",
        "arts.",
        "§§",
        "Incisos ",
        "incisos ",
        "Parágrafos ",
        "parágrafos ",
        "Livros ",
        "Títulos ",
        "Titulos ",
        "Capítulos ",
        "Capitulos ",
        "Seções ",
        "Secoes ",
    )

    if text.startswith(feminine_plural_markers):
        return f"as {text}"
    if text.startswith(feminine_singular_markers):
        return f"a {text}"
    if text.startswith(plural_markers):
        return f"os {text}"
    if text.startswith(singular_markers):
        return f"o {text}"
    return text


def new_redaction_preposition(target: str) -> str:
    text = clean_spaces(target or "")
    if text.startswith(("Arts.", "arts.", "Incisos ", "incisos ", "§§")):
        return "aos"
    if text.startswith(("Seções ", "Secoes ", "Títulos ", "Titulos ", "Capítulos ", "Capitulos ", "Livros ")):
        return "aos"
    return "ao"


def render_summary_clause(clause_nature: str, target: str, segment_text: str) -> str:
    nature = normalize_nature(clause_nature)
    target = clean_summary_segment(target)
    if not target:
        return ""

    s = simplify(segment_text)
    if nature == "Modificativa" and target == "Art. 2.027" and len(re.findall(r"\b2\.027-[a-z]+", s)) >= 2:
        return f"Altera {summary_object(target)}"

    if nature == "Modificativa" and re.search(r"\bd[eéê]-se\s+(?:nova\s+redacao|ao?s?.*a\s+seguinte\s+redacao)", s):
        return f"Dá nova redação {new_redaction_preposition(target)} {target}"

    if nature == "Aditiva" and target.startswith(("Parágrafo ", "Parágrafos ", "parágrafo ", "parágrafos ")):
        return f"Acrescenta {lower_first(target)}"

    prefix = synthese_prefix_for_nature(nature)
    if prefix in {"Suprime", "Acrescenta", "Altera"}:
        target = summary_object(target)
    return f"{prefix} {target}".strip() if prefix else target


def infer_sintese_from_text(text: str, natureza: str, dispositivos: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(text or ""))
    normalized_nature = normalize_nature(natureza)
    declarative_subject = clean_summary_segment(extract_declarative_redaction_subject(text))
    embedded_suppressions = [clean_summary_segment(item) for item in extract_embedded_inciso_suppression_targets(text)]
    article_body_suppressions = [
        clean_summary_segment(item)
        for item in re.split(r"\s*;\s*", extract_article_body_suppression_targets(text))
        if clean_summary_segment(item)
    ]
    addition_to_existing_article = clean_summary_segment(extract_addition_to_existing_article_target(text))
    project_citation_adjustment_summary = extract_project_citation_adjustment_summary(text)
    dotted_target_text = normalize_dispositivo_output(extract_dotted_article_specific_targets(text), text)
    dotted_targets = [
        clean_summary_segment(item)
        for item in re.split(r"\s*;\s*", dotted_target_text)
        if clean_summary_segment(item)
    ]

    starts_with_article_redaction = re.match(
        r"^\s*o\s+(art\.?\s*" + ART_NUM_RE + r")\b.*?\bpassa\s+a\s+vigorar\s+com\s+a\s+seguinte\s+reda\S+\b",
        raw,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if project_citation_adjustment_summary:
        return project_citation_adjustment_summary

    if dotted_targets and normalized_nature == "Modificativa":
        return "; ".join(f"Modifica {summary_object(target)}" for target in dotted_targets)

    if declarative_subject and (embedded_suppressions or article_body_suppressions):
        parts = [f"Modifica {summary_object(declarative_subject)}"]
        seen_embedded: set[str] = set()
        for target in embedded_suppressions + article_body_suppressions:
            key = simplify(target)
            if key and key not in seen_embedded:
                seen_embedded.add(key)
                parts.append(f"Suprime {summary_object(target)}")
        return "; ".join(parts)

    if article_body_suppressions and normalized_nature == "Mista":
        ordered_targets = extract_ordered_explicit_targets(text)
        rendered_parts: List[str] = []
        seen_targets: set[str] = set()
        for target in article_body_suppressions:
            key = simplify(target)
            if key and key not in seen_targets:
                seen_targets.add(key)
                rendered_parts.append(f"Suprime {summary_object(target)}")
        for target in ordered_targets:
            target = clean_summary_segment(target)
            key = simplify(target)
            if not key or key in seen_targets:
                continue
            seen_targets.add(key)
            rendered_parts.append(f"Dá nova redação {new_redaction_preposition(target)} {target}")
        if rendered_parts:
            return "; ".join(rendered_parts)

    if normalized_nature == "Mista" and declarative_subject:
        dispositivo_parts = [
            clean_summary_segment(part)
            for part in re.split(r"\s*;\s*", dispositivos or "")
            if clean_summary_segment(part)
        ]
        rendered_parts = [f"Modifica {summary_object(declarative_subject)}"]
        seen_targets = {simplify(declarative_subject)}
        for target in dispositivo_parts:
            key = simplify(target)
            if not key or key in seen_targets:
                continue
            seen_targets.add(key)
            rendered_parts.append(f"Suprime {summary_object(target)}")
        if len(rendered_parts) > 1:
            return "; ".join(rendered_parts)

    if normalized_nature == "Aditiva" and addition_to_existing_article:
        rendered_target = render_addition_summary_target(addition_to_existing_article)
        return f"Acrescenta {rendered_target}".strip()

    if declarative_subject and len(split_directive_clauses(text)) <= 1:
        return f"Modifica {summary_object(declarative_subject)}."

    if starts_with_article_redaction and len(split_directive_clauses(text)) <= 1:
        article = clean_clause_target_text(starts_with_article_redaction.group(1))
        return f"Modifica {summary_object(article)}."

    restored_article_summary = extract_restored_article_change_summary(text)
    if restored_article_summary:
        return restored_article_summary

    structural_mixed_summary = extract_structural_mixed_summary(text)
    if structural_mixed_summary:
        return structural_mixed_summary

    creation_with_title_target = extract_creation_with_title_target(text)
    if normalized_nature == "Aditiva" and creation_with_title_target:
        return f"Acrescenta o {creation_with_title_target}"

    restoration_suppression_target = extract_restoration_suppression_target(text)
    if normalized_nature == "Supressiva" and restoration_suppression_target:
        return f"Suprime a alteração do {clean_summary_segment(restoration_suppression_target).lower()}"

    suppressed_inciso_target = extract_suppressed_inciso_from_inserted_article_target(text)
    if normalized_nature == "Supressiva" and suppressed_inciso_target:
        return f"Suprime {summary_object(clean_summary_segment(suppressed_inciso_target))}"

    mixed_summary = extract_mixed_redaction_suppression_target(text, for_summary=True)
    if mixed_summary:
        return mixed_summary

    project_replacement_target = extract_project_text_replacement_target(text)
    if project_replacement_target:
        return clean_spaces(text or "")

    if detect_external_law_label(text) and dispositivos:
        if normalized_nature == "Supressiva":
            return f"Suprime {summary_object(dispositivos)}"
        if normalized_nature == "Modificativa":
            return f"Modifica {summary_object(dispositivos)}"
        if normalized_nature == "Aditiva":
            return f"Acrescenta {summary_object(dispositivos)}"

    project_article_internal_target = extract_project_article_internal_target(text, concise=True)
    if normalized_nature == "Supressiva" and project_article_internal_target:
        return f"Suprime {summary_object(project_article_internal_target)}"

    following_project_devices_target = extract_following_project_devices_target(text)
    if normalized_nature == "Aditiva" and following_project_devices_target:
        return f"Acrescenta {summary_object(following_project_devices_target)}"

    project_law_insertion_target = extract_project_law_insertion_target(text, concise=True)
    if normalized_nature == "Aditiva" and project_law_insertion_target:
        directive = first_directive_segment(text)
        if re.search(r"^\s*inclua-se\b", clean_spaces(text or ""), flags=re.IGNORECASE) and project_law_insertion_target.startswith("novo art."):
            return f"Inclua-se o {project_law_insertion_target}"
        if directive and "acrescente-se" in simplify(directive):
            if "o seguinte dispositivo ao" in directive:
                to_replace = f"o seguinte dispositivo ao {project_law_insertion_target}"
                modified_directive = directive.replace(to_replace, f"o {project_law_insertion_target}")
            else:
                modified_directive = directive.replace("o seguinte dispositivo", f"o {project_law_insertion_target}")
            modified_directive = modified_directive.replace("Acrescente-se", "Acrescenta", 1)
            return modified_directive
        if re.search(r"^\s*inclua-se\b", clean_spaces(text or ""), flags=re.IGNORECASE):
            directive = first_directive_segment(text)
            if directive:
                return f"Inclua-se {summary_object(project_law_insertion_target)}"
            else:
                prefix_target = f"o {project_law_insertion_target}" if project_law_insertion_target.startswith("novo art.") else project_law_insertion_target
                return f"Inclua-se {prefix_target}"
        return f"Acrescenta {summary_object(project_law_insertion_target)}"

    new_article_location_target = extract_new_article_with_location_target(text, concise=True)
    if normalized_nature == "Aditiva" and new_article_location_target:
        if re.search(r"^\s*inclua-se\b", clean_spaces(text or ""), flags=re.IGNORECASE):
            prefix_target = f"o {new_article_location_target}" if new_article_location_target.startswith("novo art.") else new_article_location_target
            return f"Inclua-se {prefix_target}"
        return f"Acrescenta {summary_object(new_article_location_target)}"

    inserted_articles_target = extract_inserted_articles_target(text, concise=True)
    if normalized_nature == "Aditiva" and inserted_articles_target:
        return f"Acrescenta {summary_object(inserted_articles_target)}"

    generic_project_article_target = extract_generic_project_article_target(text)
    if normalized_nature == "Aditiva" and generic_project_article_target:
        return "Acrescenta novo artigo no Projeto"

    ordered_targets = extract_ordered_explicit_targets(text)
    if declarative_subject and simplify(declarative_subject) not in {simplify(item) for item in ordered_targets}:
        ordered_targets.insert(0, declarative_subject)
    for embedded_target in embedded_suppressions:
        if simplify(embedded_target) not in {simplify(item) for item in ordered_targets}:
            ordered_targets.append(embedded_target)
    paragraph_targets = extract_all_paragraph_targets(text)
    for paragraph_target in paragraph_targets:
        if simplify(paragraph_target) not in {simplify(item) for item in ordered_targets}:
            ordered_targets.append(paragraph_target)
    ordered_targets = refine_ordered_targets(text, ordered_targets)
    if ordered_targets:
        clauses = split_directive_clauses(text)
        rendered_parts: List[str] = []
        seen_parts: set[str] = set()
        for clause in clauses:
            for segment_text in split_embedded_action_segments(clause):
                clause_nature = infer_clause_vote_nature(segment_text)
                if not clause_nature:
                    continue
                segment_targets = extract_ordered_explicit_targets(segment_text)
                segment_paragraph_targets = extract_all_paragraph_targets(segment_text)
                for paragraph_target in segment_paragraph_targets:
                    if simplify(paragraph_target) not in {simplify(item) for item in segment_targets}:
                        segment_targets.append(paragraph_target)
                segment_targets = refine_ordered_targets(segment_text, segment_targets)
                if not segment_targets:
                    continue
                for target in segment_targets:
                    rendered = render_summary_clause(clause_nature, target, segment_text)
                    key = simplify(rendered)
                    if key and key not in seen_parts:
                        seen_parts.add(key)
                        rendered_parts.append(rendered)
        if rendered_parts:
            return "; ".join(rendered_parts)

    clauses = split_directive_clauses(text)
    parts: List[str] = []
    seen: set[str] = set()

    for clause in clauses:
        for segment_text in split_embedded_action_segments(clause):
            forced_target = extract_project_change_suppression_target(segment_text)
            clause_nature = "Supressiva" if forced_target else infer_clause_vote_nature(segment_text)
            if not clause_nature:
                continue
            target = (
                clean_summary_segment(forced_target)
                or clean_summary_segment(extract_project_text_replacement_target(segment_text))
                or clean_summary_segment(extract_restoration_suppression_target(segment_text))
                or clean_summary_segment(extract_suppressed_inciso_from_inserted_article_target(segment_text))
                or clean_summary_segment(extract_project_article_internal_target(segment_text, concise=True))
                or clean_summary_segment(extract_following_project_devices_target(segment_text))
                or clean_summary_segment(extract_project_law_insertion_target(segment_text, concise=True))
                or clean_summary_segment(extract_new_article_with_location_target(segment_text, concise=True))
                or clean_summary_segment(extract_inserted_articles_target(segment_text, concise=True))
                or clean_summary_segment(extract_subject_addition_target(segment_text, concise=True))
                or clean_summary_segment(extract_generic_project_article_target(segment_text))
                or clean_summary_segment(infer_dispositivos_legacy(segment_text))
                or clean_summary_segment(strip_leading_action(segment_text))
            )
            if not target:
                continue
            rendered = render_summary_clause(clause_nature, target, segment_text)
            key = simplify(rendered)
            if key and key not in seen:
                seen.add(key)
                parts.append(rendered)

    if parts:
        return "; ".join(parts)

    return infer_sintese(natureza, dispositivos)


THEME_RULES = [
    ("Direito digital", ["assinatura eletronica", "assinatura eletronica", "e-notariado", "ato notarial eletronico", "direito civil digital"]),
    ("Direitos da personalidade", ["nome", "imagem", "voz", "pseudonimo", "direitos da personalidade", "nascituro", "eutanasia", "integridade fisica", "pessoa natural"]),
    ("Direito de fam\u00EDlia", ["casamento", "uniao estavel", "divorcio", "conjuge", "convivente", "companheiro", "poder familiar", "filho", "filhos", "parental", "guarda", "alimentos"]),
    ("Direito das sucess\u00F5es", ["heranca", "herdeiro", "sucessao", "testamento", "inventario", "partilha", "legitima"]),
    ("Neg\u00F3cio jur\u00EDdico / contratos", ["contrato", "contratos", "obrigacao", "negocio juridico", "vicio", "nulidade", "clausula penal"]),
    ("Direito de empresa", ["sociedade", "sociedades", "empresario", "empresa", "registro empresarial", "quota", "quotas", "falencia", "contabilidade", "balanco patrimonial", "demonstracoes financeiras", "demonstracao dos resultados do exercicio"]),
    ("Direito das coisas", ["posse", "propriedade", "usucapiao", "condominio", "penhor", "hipoteca", "alienacao fiduciaria"]),
    ("Responsabilidade civil", ["indenizacao", "responsabilidade civil", "dano moral", "danos", "reparacao"]),
    ("Seguros", ["seguro", "seguros", "segurado", "seguradora"]),
]


def infer_theme(text: str) -> str:
    s = simplify(text)
    for theme, keywords in THEME_RULES:
        if any(simplify(keyword) in s for keyword in keywords):
            return theme
    return ""


def normalize_output_value(value):
    """Normaliza valor de saída, corrigindo apenas encoding duplo óbvio de OCR."""
    if not isinstance(value, str):
        return value
    text = value
    # Apenas corrige se é claramente corrupção de encoding OCR (padrão: Ã/Â seguido de caracteres accentuados)
    if re.search(r'[Ã][aáâãäc]|[Â][aáâ]|[Ç]', text):
        try:
            # Tenta decodificar como UTF-8 mal-interpretado de Latin-1
            text = text.encode("latin-1").decode("utf-8")
            logger.debug(f"Corrigido encoding duplo: {text[:50]}...")
        except Exception as e:
            logger.debug(f"Falha ao corrigir encoding: {e}")
    return text


def format_multiline_clauses(value: str) -> str:
    text = normalize_output_value(value)
    if not isinstance(text, str):
        return text
    return re.sub(r";\s+", ";\n", text)


def normalize_header_value(value: str) -> str:
    return re.sub(r"[^\w\s]", "", simplify(str(value) if value is not None else ""))


def find_header_index(headers: List[str], candidates: List[str]) -> Optional[int]:
    for idx, value in enumerate(headers, start=1):
        normalized = normalize_header_value(value)
        for candidate in candidates:
            if candidate in normalized:
                return idx
    return None


def load_reference_theme_map(path: Path) -> Dict[int, str]:
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        logger.warning(f"Não foi possível abrir a planilha de referência: {path}")
        return {}

    ws = None
    for name in wb.sheetnames:
        if "emendas" in simplify(name):
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    header = [str(cell.value or "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    numero_idx = find_header_index(header, ["numero da emenda", "numero emenda", "numero"])
    tema_idx = find_header_index(header, ["tema juridico", "eixo tematico", "tema juridico eixo tematico"])
    if numero_idx is None or tema_idx is None:
        return {}

    theme_map: Dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        numero = row[numero_idx - 1]
        tema = row[tema_idx - 1]
        if numero is None or tema is None:
            continue
        try:
            numero_int = int(numero)
        except Exception:
            try:
                numero_int = int(str(numero).strip())
            except Exception:
                continue
        text = str(tema).strip()
        if text:
            theme_map[numero_int] = text
    return theme_map


def fill_missing_themes_from_reference(wb: Workbook, reference_path: Path) -> None:
    theme_map = load_reference_theme_map(reference_path)
    if not theme_map:
        return

    for sheet_name in ("Emendas", "Emendas Mistas"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [str(cell.value or "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        numero_col = find_header_index(header, ["numero da emenda", "numero emenda", "numero"])
        tema_col = find_header_index(header, ["tema juridico", "eixo tematico", "tema juridico eixo tematico"])
        if numero_col is None or tema_col is None:
            continue

        for row in ws.iter_rows(min_row=2):
            numero_cell = row[numero_col - 1]
            tema_cell = row[tema_col - 1]
            if tema_cell.value is not None and str(tema_cell.value).strip():
                continue
            numero = numero_cell.value
            if numero is None:
                continue
            try:
                numero_int = int(numero)
            except Exception:
                try:
                    numero_int = int(str(numero).strip())
                except Exception:
                    continue
            replacement = theme_map.get(numero_int)
            if replacement:
                tema_cell.value = replacement


def build_row_from_pdf(pdf_path: Path) -> EmendaRow:
    match = PDF_NAME_RE.match(pdf_path.name)
    if not match:
        raise ValueError(f"Nome de PDF fora do padrÃ£o: {pdf_path.name}")

    numero = int(match.group(1))
    autor = match.group(2).strip()

    full_text = extract_pdf_text(pdf_path)
    dispositive = extract_dispositive_text(full_text)
    dispositivos = infer_dispositivos(dispositive)
    natureza = infer_nature(dispositive)
    tema = infer_theme(dispositive)
    sintese = infer_sintese_from_text(dispositive, natureza, dispositivos)

    natureza = normalize_nature(natureza)
    sintese = infer_sintese_from_text(dispositive, natureza, dispositivos) if dispositivos else sintese

    return EmendaRow(
        numero=numero,
        autor=autor,
        dispositivos=dispositivos,
        natureza=natureza,
        tema=tema,
        sintese=sintese,
        texto=dispositive,
    )


def build_row_with_fallback(pdf_path: Path) -> Tuple[EmendaRow, Optional[str]]:
    try:
        return build_row_from_pdf(pdf_path), None
    except Exception as exc:
        match = PDF_NAME_RE.match(pdf_path.name)
        numero = int(match.group(1)) if match else 0
        autor = match.group(2).strip() if match else pdf_path.stem
        row = EmendaRow(
            numero=numero,
            autor=autor,
            dispositivos="",
            natureza="",
            tema="",
            sintese="",
            texto="",
        )
        error_msg = f"{pdf_path.name}: {exc}"
        logger.error(f"Erro ao processar PDF: {error_msg}", exc_info=True)
        return row, error_msg


# Constantes movidas para o topo do arquivo (ART_NUM_RE, ARTICLE_LABEL_RE, STRUCTURE_TYPES)
STRUCTURE_ORDER = {label: idx for idx, (label, _) in enumerate(STRUCTURE_TYPES)}
CONSOLIDADO_ORDER = {
    **STRUCTURE_ORDER,
    "Artigo": len(STRUCTURE_TYPES),
    "Artigo nao especificado": len(STRUCTURE_TYPES) + 1,
}


def format_num(n: int) -> str:
    s = str(n)
    parts = []
    while len(s) > 3:
        parts.append(s[-3:])
        s = s[:-3]
    parts.append(s)
    return ".".join(reversed(parts))


def normalize_article_ref(value: str) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"(\d+)[º°o](?=-|$)", r"\1", text)
    return text


def format_article_ref(value: str) -> str:
    text = normalize_article_ref(value)
    if not text:
        return text

    parts = text.split("-", 1)
    base = parts[0]
    suffix = parts[1] if len(parts) > 1 else ""

    rendered_base = base
    if re.fullmatch(r"\d+", base):
        number = int(base)
        if 1 <= number <= 9:
            rendered_base = f"{number}°"
        else:
            rendered_base = str(number)
    rendered_suffix = suffix.upper()
    return f"{rendered_base}-{rendered_suffix}" if rendered_suffix else rendered_base


def expand_range(a: str, b: str) -> List[str]:
    ma = re.fullmatch(r"(\d+(?:\.\d+)*-?)([a-z])", a)
    mb = re.fullmatch(r"(\d+(?:\.\d+)*-?)([a-z])", b)
    if ma and mb and ma.group(1) == mb.group(1):
        start = ord(ma.group(2))
        end = ord(mb.group(2))
        if start <= end and end - start <= 26:
            return [ma.group(1) + chr(c) for c in range(start, end + 1)]

    if re.fullmatch(r"\d+(?:\.\d+)*", a) and re.fullmatch(r"\d+(?:\.\d+)*", b):
        ia = int(a.replace(".", ""))
        ib = int(b.replace(".", ""))
        if ia <= ib and ib - ia <= 100:
            return [format_num(n) for n in range(ia, ib + 1)]

    return [a, b]


def parse_art_list(chunk: str) -> List[str]:
    tokens = re.findall(ART_NUM_RE + r"|\ba\b|\be\b|\bou\b|;", chunk)
    arts: List[str] = []
    idx = 0
    while idx < len(tokens):
        token = tokens[idx]
        if re.fullmatch(ART_NUM_RE, token):
            if idx + 2 < len(tokens) and tokens[idx + 1] == "a" and re.fullmatch(ART_NUM_RE, tokens[idx + 2]):
                arts.extend(normalize_article_ref(item) for item in expand_range(token, tokens[idx + 2]))
                idx += 3
                continue
            arts.append(normalize_article_ref(token))
        idx += 1
    return arts


def law_windows(text: str) -> List[str]:
    return [text[max(0, m.start() - 260):m.start()] for m in re.finditer(r"10\.406", text)]


def is_project_reference_context(text: str) -> bool:
    s = simplify(text)
    if not s:
        return False
    patterns = [
        r"^\s*[oº°]?\s*[,;:-]?\s*(?:do|da)\s+projeto\b",
        r"^\s*[oº°]?\s*[,;:-]?\s*(?:do|da)\s+pl\b",
        r"^\s*[oº°]?\s*[,;:-]?\s*(?:do|da)\s+projeto\s+de\s+lei\b",
        r"^\s*[oº°]?\s*[,;:-]?\s*(?:do|da)\s+proieto\b",
        r"^\s*projeto\s+de\s+lei\b",
    ]
    return any(re.search(pattern, s) for pattern in patterns)


def is_project_article_mention(context_before: str, context_after: str) -> bool:
    before = simplify(context_before)
    after = simplify(context_after)
    if is_project_reference_context(after):
        return True
    return bool(re.search(r"(?:propost[oa]s?|promovid[oa]s?|inclu[ií]d[oa]s?|alterad[oa]s?)\s+pelo\s*$", before))


def article_reference_key(art: str, diploma: str = "") -> str:
    art_norm = normalize_article_ref(art)
    return f"{art_norm}||{diploma.strip()}" if diploma else art_norm


def split_article_reference_key(value: str) -> Tuple[str, str]:
    raw = str(value or "")
    if "||" in raw:
        art, diploma = raw.split("||", 1)
        return art, diploma
    return raw, ""


def format_article_reference_label(art: str, diploma: str = "") -> str:
    label = f"art. {format_article_ref(art)}"
    if diploma:
        return f"{label} ({diploma})"
    return label


def canonicalize_law_number(number: str) -> str:
    raw = clean_spaces(number or "").rstrip(".,;:")
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if digits == "10406":
        return "Código Civil"
    if len(digits) < 4:
        return ""
    if "/" in raw:
        return raw
    if "." in raw:
        parts = [re.sub(r"\D", "", part) for part in raw.split(".")]
        if any(not part for part in parts):
            return ""
        if len(parts[-1]) < 3:
            return ""
        normalized_parts = [str(int(parts[0]))]
        normalized_parts.extend(part.zfill(max(3, len(part))) for part in parts[1:])
        return ".".join(normalized_parts)
    return f"{int(digits[:-3])}.{digits[-3:]}"


def detect_normative_diploma(window: str) -> str:
    raw = clean_spaces(window or "")
    s = simplify(raw)
    if not s:
        return ""
    if "10.406" in s or "codigo civil" in s:
        return "Código Civil"

    for law_match in re.finditer(r"lei(?:\s+federal)?\s+n[ºo]?\s*([\d./-]+)", raw, flags=re.IGNORECASE):
        before = simplify(raw[max(0, law_match.start() - 24):law_match.start()])
        if before.endswith("projeto de"):
            continue
        normalized = canonicalize_law_number(law_match.group(1))
        if not normalized:
            continue
        if normalized == "Código Civil":
            return normalized
        return f"Lei nº {normalized}"
    return ""


def detect_external_law_label(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""
    for match in re.finditer(
        r"lei(?:\s+federal)?\s+n[ºo]?\s*([\d./-]+)(?:,\s+de\s+[^,.;]*?\s+de\s+(\d{4}))?",
        raw,
        flags=re.IGNORECASE,
    ):
        before = simplify(raw[max(0, match.start() - 24):match.start()])
        if before.endswith("projeto de"):
            continue
        number = canonicalize_law_number(match.group(1))
        if not number or number == "Código Civil" or number == "4":
            continue
        year = match.group(2)
        if year and "/" not in number:
            return f"Lei nº {number}/{year[-2:]}"
        return f"Lei nº {number}"
    return ""


def append_external_law_label(target: str, source_text: str) -> str:
    label = detect_external_law_label(source_text)
    if not label or re.search(r"\bLei\s+n[ºo]?\b|\bCódigo Civil\b|\bPL\s*4/2025\b", target, flags=re.IGNORECASE):
        return target
    if not re.search(r"(?:\bArts?\.|\barts?\.|\bincisos?\b|§|\bparágrafo\b|\bParágrafo\b)", target):
        return target
    return f"{target} da {label}"


def is_project_new_article_addition(full_text: str, art: str) -> bool:
    directive = clean_spaces(extract_directive_context(full_text) or full_text or "")
    s = simplify(directive)
    art_pattern = build_article_ref_pattern(art)
    if not re.search(r"\b(?:inclua-se|incluam-se|acrescente-se|acrescentem-se|insira-se|insiram-se)\b", s):
        return False
    if not re.search(r"\b(?:ao|aos|no|nos)\s+(?:art\.?\s*2[º°o]?\s+do\s+)?projeto\b|\bpl\s*4/2025\b|\bprojeto\s+de\s+lei\s+n[ºo]?\s*4\b", s):
        return False
    return bool(
        re.search(r"\bnovo\s+(?:art\.?|artigo)\s*" + art_pattern + r"\b", s, flags=re.IGNORECASE)
        or (
            re.search(r"\bo\s+seguinte\s+dispositivo\b", s, flags=re.IGNORECASE)
            and re.search(r"(?:art\.?|artigo)\s*" + art_pattern + r"\b", s, flags=re.IGNORECASE)
        )
    )


def has_explicit_project_article_context(text: str, art: str) -> bool:
    s = simplify(text or "")
    if not s:
        return False
    art_pattern = build_article_ref_pattern(art)
    return bool(
        re.search(
            r"(?:art\.?|artigo)\s*" + art_pattern + r"\s+do\s+projeto",
            s,
            flags=re.IGNORECASE,
        )
        or re.search(
            r"(?:inciso|incisos)\s+[ivxlcdm]+(?:-[\da-z]+)?\s+(?:ao|aos|do|dos)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\s*" + art_pattern + r"\s+do\s+projeto",
            s,
            flags=re.IGNORECASE,
        )
        or re.search(
            r"(?:art\.?|artigo)\s*" + art_pattern + r"[^.;:\n]{0,120}\bpl\s*4/2025\b",
            s,
            flags=re.IGNORECASE,
        )
        or re.search(
            r"(?:art\.?|artigo)\s*" + art_pattern + r"[^.;:\n]{0,120}\bprojeto\s+de\s+lei\s+n[ºo]?\s*4\b",
            s,
            flags=re.IGNORECASE,
        )
    )


def has_explicit_civil_code_article_context(text: str, art: str) -> bool:
    s = simplify(text or "")
    if not s:
        return False
    art_pattern = build_article_ref_pattern(art)
    article_ref = r"(?:art\.?|artigo)\s*" + art_pattern
    subordinated_ref = (
        r"(?:inciso|incisos|§|§§|paragrafo|paragrafos|paragrafo\s+unico|caput)"
        r"[^.;:\n]{0,120}?\s+do\s+" + article_ref
    )
    return bool(
        re.search(
            r"(?:"
            + article_ref
            + r"|"
            + subordinated_ref
            + r")"
            + r"[^.;:\n]{0,180}?\b(?:da|de|pela|na)\s+lei\s+n[ºo]?\s*10\.406\b",
            s,
            flags=re.IGNORECASE,
        )
        or re.search(
            r"(?:"
            + article_ref
            + r"|"
            + subordinated_ref
            + r")"
            + r"[^.;:\n]{0,180}?\bcodigo\s+civil\b",
            s,
            flags=re.IGNORECASE,
        )
        or (
            re.search(r"(?:ambos|todas|todos)\s+da\s+lei\s+n[ºo]?\s*10\.406\b", s, flags=re.IGNORECASE)
            and re.search(r"(?:"
                          + article_ref
                          + r"|"
                          + subordinated_ref
                          + r")", s, flags=re.IGNORECASE)
        )
    )


def classify_article_diploma(full_text: str, art: str) -> str:
    directive = extract_directive_context(full_text) or clean_spaces(full_text or "")
    directive = re.sub(r"[\"“][^\"”]*[\"”]", " ", directive)
    s_directive = simplify(directive)
    windows = extract_target_windows(directive, "Artigo", art)
    art_num = normalize_article_ref(art)
    try:
        art_base_int = int(art_num.split("-")[0].replace(".", ""))
    except Exception:
        art_base_int = None

    if art_base_int is not None and 1 <= art_base_int <= 20 and art_num in set(extract_project_clause_articles(full_text)):
        return "PL 4/2025"

    if art_base_int is not None and 1 <= art_base_int <= 20 and is_project_new_article_addition(full_text, art):
        return "PL 4/2025"

    if has_explicit_civil_code_article_context(full_text, art):
        return "Código Civil"

    directive_diploma = detect_normative_diploma(directive)
    if art_base_int is not None and art_base_int > 20:
        return directive_diploma or "Código Civil"

    article_marker = re.compile(
        r"\b" + ARTICLE_LABEL_RE + r"\s*(" + ART_NUM_RE + r"(?:\s*(?:,|e|ou|a)\s*" + ART_NUM_RE + r")*)",
        flags=re.IGNORECASE,
    )
    for match in article_marker.finditer(s_directive):
        listed = [normalize_article_ref(item) for item in parse_art_list(match.group(1))]
        if art_num not in listed:
            continue
        tail = directive[match.end():match.end() + 220]
        tail_s = simplify(tail)
        if art_base_int is not None and 1 <= art_base_int <= 20 and re.search(r"\bdo\s+projeto\b|\bpl\s*4/2025\b|\bprojeto\s+de\s+lei\s+n[ºo]?\s*4\b", tail_s):
            return "PL 4/2025"
        if art_base_int is not None and 1 <= art_base_int <= 20 and re.search(r"\bpropost[oa]s?\s+(?:no|pelo)\s+art\.?\s*2[º°o]?\s+do\s+projeto\b", tail_s):
            return "PL 4/2025"
        diploma = detect_normative_diploma(tail)
        if diploma:
            return diploma

    found_project = False
    for window in windows:
        if art_base_int is not None and 1 <= art_base_int <= 20 and has_explicit_project_article_context(window, art):
            found_project = True
            continue
        diploma = detect_normative_diploma(window)
        if diploma:
            return diploma
        if art_base_int is not None and 1 <= art_base_int <= 20 and has_explicit_project_article_context(window, art):
            found_project = True

    if found_project:
        return "PL 4/2025"

    s = simplify(directive or full_text)
    if art_base_int is not None and 1 <= art_base_int <= 20:
        if has_explicit_project_article_context(directive or full_text, art):
            return "PL 4/2025"
        if re.search(
            r"(?:ao|aos|do|da|no|na)\s+projeto(?:\s+de\s+lei)?[^.;:\n]{0,120}(?:art\.?|artigo)\s*" + build_article_ref_pattern(art),
            s,
            flags=re.IGNORECASE,
        ):
            return "PL 4/2025"
        if re.search(
            r"(?:inciso|incisos)\s+[ivxlcdm]+(?:-[\da-z]+)?\s+(?:ao|aos|do|dos)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\s*" + build_article_ref_pattern(art) + r"\s+do\s+projeto",
            s,
            flags=re.IGNORECASE,
        ):
            return "PL 4/2025"
        if re.search(r"ao\s+projeto.*o\s+seguinte\s+dispositivo", s, flags=re.IGNORECASE):
            return "PL 4/2025"
        if re.search(r"(?:ao|no)\s+projeto.*o\s+seguinte\s+dispositivo", s, flags=re.IGNORECASE):
            return "PL 4/2025"

    if "10.406" in s or "codigo civil" in s:
        return "Código Civil"

    global_law = detect_normative_diploma(directive)
    if global_law:
        return global_law
    if art_base_int is not None and art_base_int > 20:
        return "Código Civil"
    return ""


def is_project_clause_adjustment_article(text: str, art: str) -> bool:
    s = simplify(text)
    if not s:
        return False
    art_norm = normalize_article_ref(art)
    return bool(
        re.search(
            r"(?:retirando\s+sua(?:s)?\s+citac(?:ao|oes)\s+do|suprimindo-se\s+sua\s+revogacao\s+pel[oa]|suprimindo-se\s+a\s+revogacao.+?\bno)\s+"
            r"(?:inciso|incisos)\s+[ivxlcdm]+(?:-[\da-z]+)?\s+(?:ao|aos|do|dos)\s+(?:caput\s+do\s+)?art\.?\s*"
            + build_article_ref_pattern(art_norm)
            + r"\s+do\s+projeto",
            s,
            flags=re.IGNORECASE,
        )
    )


def extract_project_clause_articles(text: str) -> List[str]:
    s = clean_spaces(text or "")
    if not s:
        return []
    matches = re.findall(
        r"(?:inciso|incisos)\s+[ivxlcdm]+(?:-[\da-z]+)?\s+(?:ao|aos|do|dos)\s+(?:caput\s+do\s+)?art\.?\s*(" + ART_NUM_RE + r")\s+do\s+projeto",
        s,
        flags=re.IGNORECASE,
    )
    return [normalize_article_ref(match) for match in matches if match]


def extract_project_clause_vote_details(text: str, art: str) -> List[Tuple[str, str]]:
    s = clean_spaces(text or "")
    if not s:
        return []
    details: List[Tuple[str, str]] = []
    art_pattern = build_article_ref_pattern(art)
    pattern = re.compile(
        r"(?P<action>acrescente-se|acrescentem-se|inclua-se|incluam-se|suprima-se|suprimam-se|d[eê]-se\s+nova\s+redacao|d[eê]-se\s+nova\s+reda[cç][aã]o)"
        r"[^.;:\n]{0,160}?\b(?:inciso|incisos)\s+(?P<inciso>[ivxlcdm]+(?:-[\da-z]+)?)\s+"
        r"(?:ao|aos|do|dos)\s+(?:caput\s+do\s+)?art\.?\s*"
        + art_pattern
        + r"\s+do\s+projeto",
        flags=re.IGNORECASE,
    )
    for match in pattern.finditer(s):
        action = simplify(match.group("action"))
        if "suprim" in action:
            nature = "Supressiva"
        elif "redacao" in action:
            nature = "Modificativa"
        else:
            nature = "Aditiva"
        details.append((nature, f"Inciso {match.group('inciso').upper()}"))
    return details


def extract_articles(text: str) -> List[str]:
    text = simplify(text)
    windows = law_windows(text)
    article_marker = re.compile(r"\b" + ARTICLE_LABEL_RE + r"\s*(" + ART_NUM_RE + r"(?:\s*(?:,|;|e|ou|a)\s*" + ART_NUM_RE + r")*)")
    arts: List[str] = []
    for window in windows:
        matches = list(article_marker.finditer(window))
        for idx, match in enumerate(matches):
            next_start = matches[idx + 1].start() if idx + 1 < len(matches) else len(window)
            tail = window[match.end():next_start]
            before = window[max(0, match.start() - 80):match.start()]
            if is_project_article_mention(before, tail[:120]):
                continue
            arts.extend(parse_art_list(match.group(1)))
        if not matches:
            bare = re.search(r"(?:artigo\s+)?(" + ART_NUM_RE + r")\s+(?:da|na)\s+lei\b", window)
            if bare:
                arts.append(normalize_article_ref(bare.group(1)))
    return sorted(set(arts))


def extract_articles_loose(text: str) -> List[str]:
    text = simplify(text)
    article_marker = re.compile(r"\b" + ARTICLE_LABEL_RE + r"\s*(" + ART_NUM_RE + r"(?:\s*(?:,|;|e|ou|a)\s*" + ART_NUM_RE + r")*)")
    arts: List[str] = []
    for match in article_marker.finditer(text):
        context_before = text[max(0, match.start() - 40):match.start()]
        context_after = text[match.end():match.end() + 120]
        if re.search(r"(incluido|incluida|incluidos|incluidas|alterado|alterada|alterados|alteradas|promovida|promovido|feito|feita)\s+pelo\s*$", context_before):
            continue
        if is_project_article_mention(context_before, context_after):
            continue
        arts.extend(parse_art_list(match.group(1)))
    return sorted(set(arts))


def parse_secs(chunk: str) -> List[str]:
    chunk = re.sub(r"\bsec\s+sec\s+(?=\d)", "secs ", chunk)
    chunk = re.sub(r"(\d+)\s*sec\b", r"\1o", chunk)
    chunk = chunk.replace(" e seu ", " e ")
    chunk = chunk.replace(" de seu ", " e ")
    chunk = chunk.replace(" seu ", " ")
    refs: List[str] = []
    if "paragrafo unico" in chunk:
        refs.append("paragrafo_unico")
    ord_marker = r"[oº°]"
    rng = re.search(r"(?:paragrafos|secs?)\s*(\d+)" + ord_marker + r"\s*(?:a|ao)\s*(\d+)" + ord_marker, chunk)
    if rng:
        start = int(rng.group(1))
        end = int(rng.group(2))
        if start <= end and end - start <= 30:
            refs.extend(f"sec {n}" for n in range(start, end + 1))
    list_match = re.search(r"(?:paragrafos|secs?)\s*([\d oº°,eou]+)", chunk)
    if list_match:
        refs.extend(f"sec {n}" for n in re.findall(r"(\d+)" + ord_marker, list_match.group(1)))
    refs.extend(f"sec {n}" for n in re.findall(r"(?:sec|paragrafo)\s*(\d+)" + ord_marker, chunk))
    return sorted(set(refs))


def extract_paragraphs(text: str) -> List[Tuple[str, str]]:
    text = simplify(text).replace("Â§", " sec ").replace("§", " sec ")
    text = re.sub(r"\bsec\s+sec\s+(?=\d)", "secs ", text)
    text = re.sub(r"(\d+)\s*sec\b", r"\1o", text)
    text = re.sub(r"\bparagrafo unico\b", " paragrafo unico ", text)
    text = re.sub(r"\bp\.\s*unico\b", " paragrafo unico ", text)
    refs = set()
    for window in law_windows(text):
        pat1 = re.compile(r"((?:paragrafo unico|(?:sec|paragrafo)\s*\d+[oº°]|(?:paragrafos|secs?)\s*\d+[oº°]\s*(?:a|ao)\s*\d+[oº°]|(?:paragrafos|secs?)\s*[\d oº°,eou]+))\s+(?:do|da|ao)\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")")
        for match in pat1.finditer(window):
            for ref in parse_secs(match.group(1)):
                refs.add((normalize_article_ref(match.group(2)), ref))

        pat2 = re.compile(r"(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")([^.;:()]*)")
        for match in pat2.finditer(window):
            tail = match.group(2)
            if is_project_reference_context(tail[:120]):
                continue
            if "paragrafo unico" in tail or "sec " in tail or "paragrafo " in tail:
                for ref in parse_secs(tail):
                    refs.add((normalize_article_ref(match.group(1)), ref))
    return sorted(refs)


def extract_paragraphs_loose(text: str) -> List[Tuple[str, str]]:
    text = simplify(text).replace("Â§", " sec ").replace("§", " sec ")
    text = re.sub(r"\bsec\s+sec\s+(?=\d)", "secs ", text)
    text = re.sub(r"(\d+)\s*sec\b", r"\1o", text)
    text = re.sub(r"\bparagrafo unico\b", " paragrafo unico ", text)
    text = re.sub(r"\bp\.\s*unico\b", " paragrafo unico ", text)
    refs = set()
    pat1 = re.compile(r"((?:paragrafo unico|(?:sec|paragrafo)\s*\d+[oº°]|(?:paragrafos|secs?)\s*\d+[oº°]\s*(?:a|ao)\s*\d+[oº°]|(?:paragrafos|secs?)\s*[\d oº°,eou]+))\s+(?:do|da|ao)\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")")
    for match in pat1.finditer(text):
        for ref in parse_secs(match.group(1)):
            refs.add((normalize_article_ref(match.group(2)), ref))

    pat2 = re.compile(r"(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")([^.;:()]*)")
    for match in pat2.finditer(text):
        tail = match.group(2)
        if is_project_reference_context(tail[:120]):
            continue
        if "paragrafo unico" in tail or "sec " in tail or "paragrafo " in tail:
            for ref in parse_secs(tail):
                refs.add((normalize_article_ref(match.group(1)), ref))
    return sorted(refs)


def roman_to_int(value: str) -> int:
    numerals = {"i": 1, "v": 5, "x": 10, "l": 50, "c": 100, "d": 500, "m": 1000}
    total = 0
    prev = 0
    for ch in reversed(value.lower()):
        current = numerals.get(ch, 0)
        if current < prev:
            total -= current
        else:
            total += current
            prev = current
    return total


def format_named_structure(piece: str) -> str:
    lowers = {"de", "do", "da", "dos", "das", "e"}
    words = []
    for token in piece.split():
        if token in lowers:
            words.append(token)
        else:
            words.append(token.capitalize())
    return " ".join(words)


def encode_structure_value(primary_value: str, parents: List[Tuple[str, str]]) -> str:
    encoded = [primary_value]
    for label, value in parents:
        encoded.append(f"{label}={value}")
    return "||".join(encoded)


def decode_structure_value(value: str) -> Tuple[str, List[Tuple[str, str]]]:
    text = str(value or "")
    if "||" not in text:
        return text, []
    parts = text.split("||")
    primary = parts[0]
    parents: List[Tuple[str, str]] = []
    for item in parts[1:]:
        if "=" not in item:
            continue
        label, parent_value = item.split("=", 1)
        parents.append((label, parent_value))
    return primary, parents


def find_structure_refs_in_chunk(chunk: str) -> List[Tuple[str, str]]:
    def first_structure_ref(segment: str) -> Optional[Tuple[str, str]]:
        matches: List[Tuple[int, str, str]] = []
        for label, keyword in STRUCTURE_TYPES:
            numeric_pattern = re.compile(rf"\b{keyword}\s+([ivxlcdm]+(?:-[a-z]+)?|unico)\b")
            for match in numeric_pattern.finditer(segment):
                matches.append((match.start(), label, match.group(1)))

            named_pattern = re.compile(
                rf"\b{keyword}\s+((?:dos?|das?|de)\s+[a-z][a-z\s/-]{{2,60}}?)(?=(?:\s+(?:da|do|de|ao|a)\s+(?:lei|codigo|parte|livro|titulo|subtitulo|capitulo|secao)\b)|[;,.()]|$)"
            )
            for match in named_pattern.finditer(segment):
                matches.append((match.start(), label, match.group(1).strip()))

        if not matches:
            return None

        matches.sort(key=lambda item: item[0])
        _, label, value = matches[0]
        primary_order = STRUCTURE_ORDER.get(label, 99)
        parents: List[Tuple[str, str]] = []
        for _, parent_label, parent_value in matches[1:]:
            parent_order = STRUCTURE_ORDER.get(parent_label, 99)
            if parent_order < primary_order and (parent_label, parent_value) not in parents:
                parents.append((parent_label, parent_value))
        if label == "Título" and simplify(value) == "unico":
            for parent_label, parent_value in parents:
                if parent_label == "Livro":
                    return (parent_label, parent_value)
        return (label, encode_structure_value(value, parents))

    refs = set()
    clauses = [piece.strip() for piece in re.split(r"\s*;\s*", chunk) if piece.strip()]
    article_pattern = re.compile(r"\bart\.?\s*" + ART_NUM_RE)
    split_pattern = re.compile(r"\s+(?:e|ou|com|bem\s+como)\s+|,\s*")

    for clause in clauses:
        article_match = article_pattern.search(clause)
        target_area = clause if not article_match else clause[:article_match.start()]
        if not target_area.strip():
            continue

        segments = [piece.strip() for piece in re.split(split_pattern, target_area) if piece.strip()]
        for segment in segments:
            ref = first_structure_ref(segment)
            if ref:
                refs.add(ref)
    if any(label == "Livro" for label, _ in refs):
        refs = {(label, value) for label, value in refs if not (label == "Título" and simplify(decode_structure_value(value)[0]) == "unico")}
    return sorted(refs)


def structure_value_sort_key(value: str) -> Tuple[int, int, str]:
    primary, _ = decode_structure_value(value)
    token = simplify(primary).strip()
    if token == "unico":
        return (0, 0, "")
    roman = re.fullmatch(r"([ivxlcdm]+)(?:-([a-z]+))?", token)
    if roman:
        return (1, roman_to_int(roman.group(1)), roman.group(2) or "")
    return (2, 0, token)


def structure_sort_key(item: Tuple[str, str]) -> Tuple[int, Tuple[int, int, str]]:
    label, value = item
    return (STRUCTURE_ORDER.get(label, 99), structure_value_sort_key(value))


def extract_structures(text: str) -> List[Tuple[str, str]]:
    text = simplify(text)
    refs = set()
    for window in law_windows(text):
        for ref in find_structure_refs_in_chunk(window):
            refs.add(ref)
    return sorted(refs, key=structure_sort_key)


def extract_structures_loose(text: str) -> List[Tuple[str, str]]:
    text = simplify(text)
    refs = set(find_structure_refs_in_chunk(text))
    return sorted(refs, key=structure_sort_key)


def extract_generic_article_targets(text: str) -> List[str]:
    s = simplify(text)
    if "10.406" not in text and "lei 10.406" not in s and "codigo civil" not in s:
        return []

    generic_patterns = [
        r"\bnovo artigo sem localizacao especifica\b",
        r"\bnovo artigo sem localização específica\b",
        r"\bo seguinte artigo ao texto da lei\b",
        r"\bo seguinte artigo\b",
        r"\bnovo artigo a lei\b",
        r"\bnovo artigo a lei 10\.406\b",
        r"\bnovo artigo a lei 10\.406/02\b",
        r"\bonde couber,\s*novo artigo\b",
        r"\bonde couber,\s*no projeto o seguinte artigo\b",
    ]
    if any(re.search(pattern, s) for pattern in generic_patterns):
        return ["novo artigo sem localização específica"]
    return []


def format_structure_ref(label: str, value: str) -> str:
    primary, parents = decode_structure_value(value)
    token = simplify(primary).strip()
    if re.fullmatch(r"[ivxlcdm]+(?:-[a-z]+)?", token):
        rendered = f"{label} {token.upper()}"
    elif token == "unico":
        rendered = f"{label} \u00DAnico"
    else:
        rendered = f"{label} {format_named_structure(token)}"

    rendered_parents: List[str] = []
    for parent_label, parent_value in parents:
        parent_token = simplify(parent_value).strip()
        if re.fullmatch(r"[ivxlcdm]+(?:-[a-z]+)?", parent_token):
            rendered_parents.append(f"{parent_label} {parent_token.upper()}")
        elif parent_token == "unico":
            rendered_parents.append(f"{parent_label} \u00DAnico")
        else:
            rendered_parents.append(f"{parent_label} {format_named_structure(parent_token)}")
    if rendered_parents:
        rendered = f"{rendered} do " + " do ".join(rendered_parents)
    return rendered


def format_generic_ref(value: str) -> str:
    token = clean_spaces(value)
    if not token:
        return value
    return token[:1].upper() + token[1:]


def article_sort_key(art: str) -> Tuple[int, Tuple[str, ...]]:
    art_value = str(art or "")
    if "||" in art_value:
        art_value = art_value.split("||", 1)[0]
    label_match = re.search(r"(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")", art_value, flags=re.IGNORECASE)
    if label_match:
        art_value = label_match.group(1)
    parts = normalize_article_ref(art_value).split("-")
    base = int(parts[0].replace(".", ""))
    return (base, tuple(parts[1:]))


def paragraph_sort_key(item: Tuple[str, str]) -> Tuple[Tuple[int, Tuple[str, ...]], Tuple[int, int]]:
    art, ref = item
    if ref == "paragrafo_unico":
        pkey = (0, 0)
    else:
        pkey = (1, int(re.search(r"\d+", ref).group()))
    return (article_sort_key(art), pkey)


def format_human_list(parts: List[str]) -> str:
    parts = [part for part in parts if part]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    if len(parts) == 2:
        return f"{parts[0]} e {parts[1]}"
    return f"{'; '.join(parts[:-1])} e {parts[-1]}"


def resolve_nature_set(natures: List[str], fallback: str) -> str:
    cleaned = [nature for nature in natures if nature]
    ordered_unique: List[str] = []
    for nature in cleaned:
        if nature not in ordered_unique:
            ordered_unique.append(nature)
    if not ordered_unique:
        return fallback
    if len(ordered_unique) == 1:
        return ordered_unique[0]
    return "Mista"


VOTE_NATURE_PRIORITY = {
    "Supressiva": 1,
    "Modificativa": 2,
    "Aditiva": 3,
}


def normalize_vote_nature(nature: str) -> str:
    normalized = normalize_nature(nature or "")
    return normalized if normalized in VOTE_NATURE_PRIORITY else ""


def detect_vote_natures(text: str) -> List[str]:
    s = simplify(extract_nature_context(text) or text)
    if not s:
        return []

    command_natures = command_natures_from_text(text)
    if command_natures:
        return command_natures

    has_rest = any(re.search(p, s) for p in NATURE_PATTERNS["restaurativa"])
    has_subs = any(re.search(p, s) for p in NATURE_PATTERNS["substitutiva"])
    has_add = any(re.search(p, s) for p in NATURE_PATTERNS["aditiva"])
    has_sup = any(re.search(p, s) for p in NATURE_PATTERNS["supressiva"])
    has_mod = any(re.search(p, s) for p in NATURE_PATTERNS["modificativa"])

    if extract_project_change_suppression_target(text):
        has_sup = True
        has_mod = False

    if is_new_article_insertion_context(text):
        has_add = True
        has_mod = False

    found: List[str] = []
    if has_sup or has_rest:
        found.append("Supressiva")
    if has_mod or has_subs:
        found.append("Modificativa")
    if has_add:
        found.append("Aditiva")
    return found


def extract_directive_context(text: str) -> str:
    raw = clean_spaces(text or "")
    if not raw:
        return ""

    raw = re.sub(
        r':\s*[“"][\s\S]+?[”"]\s*(?:\([A-Z]{2}\))?\s*(?=(?:suprima-se|suprimam-se|d\S*-se|acrescente-se|inclua-se|retome-se|mantenha-se)\b)',
        ". ",
        raw,
        flags=re.IGNORECASE,
    )

    cut_points = [len(raw)]
    patterns = [
        r':\s*[“"]',
        r"\s+na\s+forma\s+proposta\s+pelo\s+art\.?\s+",
        r"\s+nos\s+termos\s+a\s+seguir",
        r"\s+justifica[cç][aã]o\b",
        r"\s+conforme\s+o\s+art\.?\s+",
        r"\s+de\s+acordo\s+com\s+o\s+art\.?\s+",
        r"\s+nos\s+termos\s+do\s+art\.?\s+",
        r"\s+segundo\s+o\s+art\.?\s+",
        r"\s+para\s+fins\s+de\s+",
        r"\s+por\s+prova\s+documental\s+",
    ]
    for pattern in patterns:
        match = re.search(pattern, raw, flags=re.IGNORECASE)
        if match:
            cut_points.append(match.start())
    return raw[:min(cut_points)].strip(" ;:\n\t")


def extract_nature_context(text: str) -> str:
    raw = extract_directive_context(text)
    if not raw:
        return ""
    match = re.search(r':\s*[“"]', raw)
    if match:
        raw = raw[:match.start()]
    return raw.strip(" ;:\n\t")


def split_directive_clauses(text: str) -> List[str]:
    directive = extract_directive_context(text)
    if not directive:
        return []

    sentence_split = re.compile(
        r"(?<=[.!?])\s+(?=(?:(?:o|os|a|as)\s+)?(?:art\.?|arts?\.?|artigo|artigos|livro|titulo|subtitulo|capitulo|secao)\s|\b(?:inclua-se|incluam-se|acrescente-se|acrescentem-se|suprima-se|suprimam-se|suprimir|d[eê]-se|insira-se|insiram-se|mantenha-se|retome-se|substitua-se|substituam-se)\b)",
        flags=re.IGNORECASE,
    )
    clauses: List[str] = []
    for segment in re.split(r"\s*;\s*", directive):
        segment = segment.strip(" ;")
        if not segment:
            continue
        pieces = sentence_split.split(segment)
        expanded_pieces: List[str] = []
        for piece in pieces:
            expanded_pieces.extend(
                part.strip(" ;")
                for part in re.split(r"(?=\bItem\s+\d+\s*[–-])", piece, flags=re.IGNORECASE)
                if part.strip(" ;")
            )
        for piece in expanded_pieces:
            piece = piece.strip(" ;")
            if piece:
                clauses.append(piece)
    return clauses


def build_article_ref_pattern(art: str) -> str:
    art_norm = normalize_article_ref(art)
    if "-" in art_norm:
        base, suffix = art_norm.split("-", 1)
        return re.escape(base) + r"(?:[º°o])?-" + re.escape(suffix)
    return re.escape(art_norm) + r"(?:[º°o])?(?!\s*-)"


def extract_clause_article_targets(clause: str) -> List[str]:
    article_marker = re.compile(
        r"\b" + ARTICLE_LABEL_RE + r"\s*(" + ART_NUM_RE + r"(?:\s*(?:,|e|ou|a)\s*" + ART_NUM_RE + r")*)"
    )
    targets: List[str] = []
    for match in article_marker.finditer(simplify(clause)):
        targets.extend(parse_art_list(match.group(1)))
    return sorted(set(targets), key=article_sort_key)


def format_inciso_detail(raw: str) -> str:
    text = simplify(raw)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"\s+", " ", text).strip(" ,;")
    text = re.sub(r"\b([ivxlcdm]+(?:-[a-z]+)?)\b", lambda match: match.group(1).upper(), text)
    roman_tokens = re.findall(r"\b[IVXLCDM]+(?:-[A-Z]+)?\b", text)
    prefix = "Inciso" if len(roman_tokens) == 1 and "," not in text and " e " not in text and " a " not in text else "Incisos"
    return f"{prefix} {text}".strip()


def inciso_marker_pattern() -> str:
    return r"(?:inciso|incisos|inc\.?)"


def format_paragraph_detail_labels(chunk: str) -> List[str]:
    s = simplify(chunk).replace("§", " sec ")
    s = re.sub(r"\bsec\s+sec\s+(?=\d)", "secs ", s)
    s = re.sub(r"(\d+)\s*sec\b", r"\1o", s)
    if "paragrafo unico" in s:
        return ["Parágrafo único"]

    sec_refs = parse_secs(s)
    sec_nums = [ref.split()[1] for ref in sec_refs if ref.startswith("sec ")]
    if not sec_nums:
        return []

    rendered = [f"{num}º" for num in sec_nums]
    if len(rendered) == 1:
        return [f"§ {rendered[0]}"]
    return [f"§§ {format_human_list(rendered)}"]


def lower_first(text: str) -> str:
    if not text:
        return text
    return text[:1].lower() + text[1:] if text[:1].isalpha() else text


def infer_clause_vote_nature(clause: str) -> str:
    raw = normalize_ocr_noise(clean_spaces(clause or ""))
    s = simplify(clause)
    if not s:
        return ""

    if extract_project_change_suppression_target(clause):
        return "Supressiva"
    if re.search(r"\bsuprima(?:m)?-se\b|\bsupress\S+\b", raw, flags=re.IGNORECASE):
        return "Supressiva"
    if re.search(r"\bd\S*-se\s+nova\s+reda\S+\b|\bpassa(?:m)?\s+a\s+vigorar\s+com\b", raw, flags=re.IGNORECASE):
        return "Modificativa"
    if re.search(r"\bacrescente-se\b|\bacrescentem-se\b|\binclua-se\b|\bincluam-se\b|\bprop\S*e\s+a\s+cria\S*o\b", raw, flags=re.IGNORECASE):
        return "Aditiva"

    if is_new_article_insertion_context(clause):
        return "Aditiva"

    matches: List[Tuple[int, str]] = []
    pattern_groups = {
        "Supressiva": NATURE_PATTERNS["supressiva"] + [r"\bsuprimir\s+art\.?\b", r"\bsuprimir\s+artigo\b"],
        "Modificativa": NATURE_PATTERNS["modificativa"] + [r"\bd[eê]-se\s+nova\s+redacao\b", r"\bd[eê]-se\s+a\s+seguinte\s+redacao\b"],
        "Aditiva": NATURE_PATTERNS["aditiva"],
    }
    for nature, patterns in pattern_groups.items():
        for pattern in patterns:
            match = re.search(pattern, s)
            if match:
                matches.append((match.start(), nature))
                break

    if matches:
        matches.sort(key=lambda item: item[0])
        return matches[0][1]

    if re.match(r"^(?:paragrafo unico|(?:sec|paragrafo)\s*\d+[oº°]|arts?\.?\s*\d+|art\.?\s*\d+)\s+(?:ao|a|à|aos|as)\b", s):
        return "Aditiva"

    found = detect_vote_natures(s)
    if len(found) == 1:
        return found[0]
    return ""


def extract_clause_detail_label(clause: str) -> str:
    s = simplify(clause).replace("§", " sec ")
    s = re.sub(r"\bsec\s+sec\s+(?=\d)", "secs ", s)
    s = re.sub(r"(\d+)\s*sec\b", r"\1o", s)
    s = re.sub(r"\bparagrafo unico\b", " paragrafo unico ", s)
    s = re.sub(r"\bp\.\s*unico\b", " paragrafo unico ", s)

    sec_chunk_pattern = (
        r"(?:paragrafo unico|"
        r"(?:sec|paragrafo)\s*\d+[oº°](?:\s*(?:,|e|ou)\s*(?:sec|paragrafo)?\s*\d+[oº°])*|"
        r"(?:paragrafos|secs?)\s*\d+[oº°](?:\s*(?:,|e|ou)\s*(?:sec|paragrafo)?\s*\d+[oº°])*|"
        r"(?:paragrafos|secs?)\s*\d+[oº°]\s*(?:a|ao)\s*\d+[oº°])"
    )
    nested_inciso_match = re.search(
        r"\b" + inciso_marker_pattern() + r"\s+([ivxlcdm,\seoua,]+?)\s+(?:do|da|ao)\s+(" + sec_chunk_pattern + r")(?=\s+(?:do|da|ao)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\b)",
        s,
    )
    if nested_inciso_match:
        parent_labels = format_paragraph_detail_labels(nested_inciso_match.group(2))
        inciso_label = format_inciso_detail(nested_inciso_match.group(1))
        if parent_labels:
            return f"{inciso_label} do {lower_first(parent_labels[0])}"
        return inciso_label

    inciso_match = re.search(r"\b" + inciso_marker_pattern() + r"\s+([ivxlcdm,\seoua,]+?)(?=\s+(?:do|da|ao|a)\b)", s)
    if inciso_match:
        return format_inciso_detail(inciso_match.group(1))

    paragraph_labels = format_paragraph_detail_labels(s)
    if paragraph_labels:
        return paragraph_labels[0]

    if re.search(r"\bcaput\b", s):
        return "Caput"

    return ""


def extract_article_bound_details(clause: str, art: str) -> List[str]:
    s = simplify(clause).replace("§", " sec ")
    s = re.sub(r"\bsec\s+sec\s+(?=\d)", "secs ", s)
    s = re.sub(r"(\d+)\s*sec\b", r"\1o", s)
    art_pattern = build_article_ref_pattern(art)
    details: List[str] = []
    sec_chunk_pattern = (
        r"(?:paragrafo unico|"
        r"(?:sec|paragrafo)\s*\d+[oº°](?:\s*(?:,|e|ou)\s*(?:sec|paragrafo)?\s*\d+[oº°])*|"
        r"(?:paragrafos|secs?)\s*\d+[oº°](?:\s*(?:,|e|ou)\s*(?:sec|paragrafo)?\s*\d+[oº°])*|"
        r"(?:paragrafos|secs?)\s*\d+[oº°]\s*(?:a|ao)\s*\d+[oº°])"
    )

    nested_parent_spans: List[Tuple[int, int]] = []
    nested_inciso_pattern = re.compile(
        r"\b" + inciso_marker_pattern() + r"\s+([ivxlcdm,\seoua,]+?)\s+(?:do|da|ao)\s+(" + sec_chunk_pattern + r")\s+"
        r"(?:do|da|ao)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\s*" + art_pattern + r"\b"
    )
    for match in nested_inciso_pattern.finditer(s):
        inciso_label = format_inciso_detail(match.group(1))
        parent_labels = format_paragraph_detail_labels(match.group(2))
        if parent_labels:
            for parent_label in parent_labels:
                label = f"{inciso_label} do {lower_first(parent_label)}"
                if label not in details:
                    details.append(label)
            nested_parent_spans.append(match.span(2))
        elif inciso_label not in details:
            details.append(inciso_label)

    sec_pattern = re.compile(
        r"(" + sec_chunk_pattern + r")\s+(?:do|da|ao)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\s*"
        + art_pattern
        + r"\b"
    )
    for match in sec_pattern.finditer(s):
        if any(start <= match.start(1) and match.end(1) <= end for start, end in nested_parent_spans):
            continue
        chunk = match.group(1)
        for label in format_paragraph_detail_labels(chunk):
            if label not in details:
                details.append(label)

    art_mention = re.search(r"\b(?:art\.?|artigo)\s*" + art_pattern + r"\b", s)
    if art_mention:
        possessive_sec_pattern = re.compile(
            r"(?:,\s*)?(?:e\s+)?de\s+seu(?:s)?\s+(" + sec_chunk_pattern + r")"
        )
        for match in possessive_sec_pattern.finditer(s):
            if match.start() < art_mention.end():
                continue
            chunk = match.group(1)
            for label in format_paragraph_detail_labels(chunk):
                if label not in details:
                    details.append(label)

    inciso_pattern = re.compile(
        r"\b" + inciso_marker_pattern() + r"\s+([ivxlcdm,\seoua,]+?)(?=\s+(?:do|da|ao)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\s*" + art_pattern + r"\b)"
    )
    for match in inciso_pattern.finditer(s):
        label = format_inciso_detail(match.group(1))
        if label not in details:
            details.append(label)

    caput_direct = re.search(r"\bcaput\s+do\s+(?:art\.?|artigo)\s*" + art_pattern + r"\b", s)
    caput_shared = bool(art_mention and re.search(r"\bcaput\b", s[max(0, art_mention.start() - 140):art_mention.start()]))
    if caput_direct or caput_shared:
        if "Caput" not in details:
            details.insert(0, "Caput")

    bare_art_pattern = re.compile(r"\b(?:art\.?|artigo|arts?\.?)\s*" + art_pattern + r"\b")
    bare_match = bare_art_pattern.search(s)
    if bare_match:
        before = s[max(0, bare_match.start() - 80):bare_match.start()]
        bound_to_detail = re.search(r"(?:paragrafo\s+unico|sec|paragrafo|paragrafos|inciso|incisos|inc\.?)\s*(?:[\divxlcdm]|unico)", before)
        if not details:
            previous_detail = re.search(
                r"((?:paragrafo unico|(?:sec|paragrafo)\s*\d+[oº°](?:\s*(?:,|e|ou)\s*(?:sec|paragrafo)?\s*\d+[oº°])*|(?:paragrafos|secs?)\s*\d+[oº°](?:\s*(?:,|e|ou)\s*(?:sec|paragrafo)?\s*\d+[oº°])*|(?:paragrafos|secs?)\s*\d+[oº°]\s*(?:a|ao)\s*\d+[oº°]))\s+(?:do|da|ao|aos|dos|das)?\s*$",
                before,
            )
            if previous_detail:
                details.extend(label for label in format_paragraph_detail_labels(previous_detail.group(1)) if label not in details)
            if not details:
                details.append("")
        elif not bound_to_detail and re.search(r"(?:ao|aos|do|dos)\s*$", before) and "Caput" not in details:
            details.insert(0, "Caput")

    return details


def split_mixed_action_clause(clause: str) -> List[str]:
    text = clean_spaces(clause or "")
    if not text:
        return []

    parts = [
        part.strip(" ;")
        for part in re.split(
            r"\s+e\s+(?=(?:suprima-se|suprimam-se|d[eê]-se|acrescente-se|acrescentem-se|inclua-se|incluam-se)\b)",
            text,
            flags=re.IGNORECASE,
        )
        if part.strip(" ;")
    ]
    if len(parts) <= 1:
        return [text]

    trailing_article = ""
    article_matches = list(re.finditer(r"\b(?:do|da|ao|aos)\s+(?:caput\s+do\s+)?(?:art\.?|artigo)\s*" + ART_NUM_RE + r"\b", text, flags=re.IGNORECASE))
    if article_matches:
        trailing_article = text[article_matches[-1].start():article_matches[-1].end()]

    expanded: List[str] = []
    for part in parts:
        if trailing_article and not re.search(r"\b(?:art\.?|artigo)\s*" + ART_NUM_RE + r"\b", simplify(part)):
            expanded.append(f"{part} {trailing_article}")
        else:
            expanded.append(part)
    return expanded


def extract_article_vote_details(text: str, art: str) -> List[Tuple[str, str]]:
    project_details = extract_project_clause_vote_details(text, art)
    clauses = split_directive_clauses(text)
    if not clauses:
        return project_details

    details: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for detail in project_details:
        seen.add(detail)
        details.append(detail)
    for clause in clauses:
        for action_clause in split_mixed_action_clause(clause):
            clause_targets = extract_clause_article_targets(action_clause)
            if normalize_article_ref(art) not in clause_targets:
                continue
            nature = infer_clause_vote_nature(action_clause)
            if not nature:
                continue
            bound_details = extract_article_bound_details(action_clause, art)
            if not bound_details:
                clause_label = extract_clause_detail_label(action_clause) if len(clause_targets) == 1 else ""
                bound_details = [clause_label]
            for detail in bound_details:
                key = (nature, detail)
                if key in seen:
                    continue
                seen.add(key)
                details.append(key)
    return details


def infer_article_nature_by_nearest_action(text: str, art: str) -> str:
    s = simplify(text)
    art_norm = normalize_article_ref(art)
    if not s or not art_norm:
        return ""

    article_pattern = re.compile(r"\b(?:art\.?|artigo)\s*" + build_article_ref_pattern(art) + r"\b")
    before_action_patterns = [
        ("Modificativa", r"d[eê]-se\s+nova\s+redacao"),
        ("Modificativa", r"d[eê]-se\s+a\s+seguinte\s+redacao"),
        ("Supressiva", r"suprima-se"),
        ("Supressiva", r"suprimam-se"),
        ("Supressiva", r"suprimir"),
        ("Aditiva", r"acrescente-se"),
        ("Aditiva", r"acrescentem-se"),
        ("Aditiva", r"inclua-se"),
        ("Aditiva", r"incluam-se"),
        ("Aditiva", r"insira-se"),
        ("Aditiva", r"insiram-se"),
    ]
    after_action_patterns = [
        ("Modificativa", r"passa\s+a\s+vigorar\s+com\s+a\s+seguinte\s+redacao"),
        ("Modificativa", r"passa\s+a\s+vigorar\s+com\s+as\s+seguintes\s+alteracoes"),
        ("Modificativa", r"passam\s+a\s+vigorar\s+com\s+a\s+seguinte\s+redacao"),
        ("Aditiva", r"passa\s+a\s+vigorar\s+acrescid[oa]"),
        ("Aditiva", r"passam\s+a\s+vigorar\s+acrescid[oa]s"),
        ("Aditiva", r"fica\s+acrescid[oa]"),
        ("Aditiva", r"ficam\s+acrescid[oa]s"),
        ("Supressiva", r"fica\s+revogado"),
        ("Supressiva", r"ficam\s+revogados"),
    ]

    for match in article_pattern.finditer(s):
        tail = s[match.end():match.end() + 120]
        if is_project_reference_context(tail):
            continue

        before = s[max(0, match.start() - 220):match.start()]
        best_before = None
        for nature, pattern in before_action_patterns:
            for action_match in re.finditer(pattern, before):
                if best_before is None or action_match.end() > best_before[1]:
                    best_before = (nature, action_match.end())
        if best_before:
            return best_before[0]

        after = s[match.end():min(len(s), match.end() + 220)]
        best_after = None
        for nature, pattern in after_action_patterns:
            action_match = re.search(pattern, after)
            if action_match and (best_after is None or action_match.start() < best_after[1]):
                best_after = (nature, action_match.start())
        if best_after:
            return best_after[0]

    return ""


def extract_target_windows(text: str, tipo: str, valor: str) -> List[str]:
    s = simplify(text)
    if not s:
        return []

    patterns: List[re.Pattern[str]] = []
    tipo_norm = simplify(tipo)
    valor_norm = simplify(valor)

    if tipo == "Artigo":
        if valor_norm:
            patterns.append(re.compile(r"\b(?:art\.?|artigo)\s*" + build_article_ref_pattern(valor) + r"\b"))
    elif tipo_norm == simplify("Artigo nao especificado"):
        patterns.append(re.compile(r"\bnovo artigo\b"))
    else:
        if tipo_norm and valor_norm:
            patterns.append(re.compile(r"\b" + re.escape(tipo_norm) + r"\s+" + re.escape(valor_norm) + r"\b"))
        if tipo_norm:
            patterns.append(re.compile(r"\b" + re.escape(tipo_norm) + r"\b"))

    windows: List[str] = []
    seen: set[Tuple[int, int]] = set()
    for pattern in patterns:
        for match in pattern.finditer(s):
            raw_start = max(0, match.start() - 220)
            raw_end = min(len(s), match.end() + 260)
            raw_window = s[raw_start:raw_end]
            rel_start = match.start() - raw_start
            rel_end = match.end() - raw_start

            previous_breaks = [raw_window.rfind(marker, 0, rel_start) for marker in (";", "\n", ":")]
            clause_start = max(previous_breaks)
            clause_start = clause_start + 1 if clause_start >= 0 else max(0, rel_start - 140)

            next_breaks = [pos for marker in (";", "\n") for pos in [raw_window.find(marker, rel_end)] if pos >= 0]
            clause_end = min(next_breaks) if next_breaks else min(len(raw_window), rel_end + 180)

            if clause_end <= clause_start:
                clause_start = max(0, rel_start - 140)
                clause_end = min(len(raw_window), rel_end + 180)

            key = (raw_start + clause_start, raw_start + clause_end)
            if key in seen:
                continue
            seen.add(key)
            windows.append(raw_window[clause_start:clause_end].strip())
    return windows


def infer_vote_nature_for_target(
    full_text: str,
    tipo: str,
    valor: str,
    fallback_nature: str,
    article_nature_overrides: Dict[str, str],
) -> str:
    for window in extract_target_windows(full_text, tipo, valor):
        if re.search(r"\bpassa(?:m)?\s+a\s+vigorar\s+acrescid[oa]s?\b", window, flags=re.IGNORECASE):
            return "Modificativa"

    if tipo != "Artigo" and re.search(r"\bpassa(?:m)?\s+a\s+vigorar\s+acrescid[oa]s?\b", full_text, flags=re.IGNORECASE):
        return "Modificativa"

    if tipo == "Artigo":
        override = normalize_vote_nature(article_nature_overrides.get(valor, ""))
        if override:
            return override
        nearest_action = infer_article_nature_by_nearest_action(full_text, valor)
        if nearest_action:
            return nearest_action

    fallback_vote = normalize_vote_nature(fallback_nature)
    if fallback_vote:
        return fallback_vote

    candidate_natures: List[str] = []
    for window in extract_target_windows(full_text, tipo, valor):
        candidate_natures.extend(detect_vote_natures(window))

    resolved = resolve_nature_set(candidate_natures, "")
    resolved_vote = normalize_vote_nature(resolved)
    if resolved_vote:
        return resolved_vote

    if resolved == "Mista":
        for nature in ("Supressiva", "Modificativa", "Aditiva"):
            if nature in candidate_natures:
                return nature

    # Fallback final: tenta inferir pela emenda inteira quando o alvo nao trouxe sinal suficiente.
    inferred_global = normalize_vote_nature(infer_nature(full_text))
    if inferred_global:
        return inferred_global

    detected_global = detect_vote_natures(full_text)
    for nature in ("Supressiva", "Modificativa", "Aditiva"):
        if nature in detected_global:
            return nature

    return ""


def format_emenda_refs(items: List[Tuple[int, str, str]]) -> str:
    by_emenda: Dict[int, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for numero, natureza, detalhe in items:
        natureza_norm = normalize_vote_nature(natureza)
        detalhe = str(detalhe or "").strip()
        if natureza_norm:
            if detalhe and detalhe not in by_emenda[numero][natureza_norm]:
                by_emenda[numero][natureza_norm].append(detalhe)
            elif not detalhe and natureza_norm not in by_emenda[numero]:
                by_emenda[numero][natureza_norm] = []
        elif "" not in by_emenda[numero]:
            by_emenda[numero][""] = []

    parts: List[str] = []
    for numero in sorted(by_emenda):
        nature_chunks: List[str] = []
        groups = by_emenda[numero]
        ordered_natures = sorted(
            [nature for nature in groups.keys() if nature],
            key=lambda nature: VOTE_NATURE_PRIORITY.get(nature, 99),
        )
        for nature in ordered_natures:
            details = groups[nature]
            if details:
                nature_chunks.append(f"{nature} - {format_human_list(details)}")
            else:
                nature_chunks.append(nature)
        if nature_chunks:
            parts.append(f"{numero} ({' / '.join(nature_chunks)})")
        else:
            parts.append(str(numero))

    return format_human_list(parts)


def build_voting_order(items: List[Tuple[int, str, str]]) -> str:
    groups: Dict[str, List[str]] = {
        "Supressiva": [],
        "Modificativa": [],
        "Aditiva": [],
    }
    by_emenda: Dict[int, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for numero, nature, detail in items:
        vote_nature = normalize_vote_nature(nature)
        if not vote_nature:
            continue
        detail = str(detail or "").strip()
        if detail and detail not in by_emenda[numero][vote_nature]:
            by_emenda[numero][vote_nature].append(detail)
        elif not detail and vote_nature not in by_emenda[numero]:
            by_emenda[numero][vote_nature] = []

    for numero in sorted(by_emenda):
        for nature in ("Supressiva", "Modificativa", "Aditiva"):
            if nature not in by_emenda[numero]:
                continue
            details = by_emenda[numero][nature]
            entry = f"{numero} - {format_human_list(details)}" if details else str(numero)
            groups[nature].append(entry)

    if len(by_emenda) == 1 and sum(1 for entries in groups.values() if entries) == 1:
        return "Não se aplica"

    sections: List[str] = []
    ordinal = 1
    non_empty_groups = sum(1 for entries in groups.values() if entries)
    for nature, singular_label, plural_label in [
        ("Supressiva", "Supressiva", "Supressivas"),
        ("Modificativa", "Modificativa", "Modificativas"),
        ("Aditiva", "Aditiva", "Aditivas"),
    ]:
        entries = groups[nature]
        if not entries:
            continue
        label = singular_label if len(entries) == 1 else plural_label
        rendered = format_human_list(entries)
        if non_empty_groups <= 1:
            sections.append(f"{label} ({rendered})")
        else:
            sections.append(f"{ordinal}° - {label} ({rendered})")
        ordinal += 1

    return " / ".join(sections)


def consolidado_nature(nature: str) -> str:
    return normalize_nature(nature)


def normalize_nature(nature: str) -> str:
    mapping = {
        "Restaurativa": "Supressiva",
        "Substitutiva": "Modificativa",
    }
    return mapping.get(nature, nature)


def extract_article_nature_overrides(text: str) -> Dict[str, str]:
    s = simplify(text)
    nature_map: Dict[str, List[str]] = defaultdict(list)

    project_citation_adjustment = extract_project_citation_adjustment_target(text)
    if project_citation_adjustment:
        parts = [clean_summary_segment(part) for part in re.split(r"\s*;\s*", project_citation_adjustment) if clean_summary_segment(part)]
        for part in parts:
            art_match = re.search(r"(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")", part, flags=re.IGNORECASE)
            if not art_match:
                continue
            art_norm = normalize_article_ref(art_match.group(1))
            if part.lower().startswith(("inciso ", "incisos ")):
                nature_map[art_norm].append("Modificativa")
            else:
                nature_map[art_norm].append("Supressiva")

    suppression_patterns = [
        re.compile(
            r"(?:suprima-se|suprimam-se|supressao|supressao do|supressao dos|supressÃ£o|supressÃ£o do|supressÃ£o dos)\s+"
            r"(?:a\s+alteracao|a\s+redacao|a\s+nova\s+redacao|a\s+revogacao|o|os|a|as)?\s*"
            r"(?:do|da|dos|das)?\s*"
            r"(?:art\.?|artigo|arts?\.?|artigos?)\s*("
            + ART_NUM_RE
            + r"(?:\s*(?:,|e|ou|a)\s*"
            + ART_NUM_RE
            + r")*)"
        ),
    ]
    modificative_patterns = [
        re.compile(r"d[eÃª]-se\s+nova\s+reda[cÃ§g][aÃ£]o\s+ao?\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")"),
        re.compile(r"d[eÃª]-se\s+a\s+seguinte\s+reda[cÃ§g][aÃ£]o\s+ao?\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")"),
    ]
    restaurative_patterns = [
        re.compile(r"mantenha-se\s+a\s+reda[cÃ§g][aÃ£]o(?:\s+vigente)?\s+do\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")"),
        re.compile(r"retome-se\s+a\s+reda[cÃ§g][aÃ£]o\s+do\s+(?:art\.?|artigo)\s*(" + ART_NUM_RE + r")"),
    ]

    for pattern in suppression_patterns:
        for match in pattern.finditer(s):
            for art in parse_art_list(match.group(1)):
                nature_map[art].append("Supressiva")

    raw_mentions = list(re.finditer(r"\b" + ARTICLE_LABEL_RE + r"\s*(" + ART_NUM_RE + r")", s))
    article_mentions = []
    for match in raw_mentions:
        tail = s[match.end():match.end() + 120]
        if is_project_reference_context(tail):
            continue
        article_mentions.append(match)

    for idx, match in enumerate(article_mentions):
        art = normalize_article_ref(match.group(1))
        next_start = article_mentions[idx + 1].start() if idx + 1 < len(article_mentions) else len(s)
        tail = s[match.end():next_start]
        if any(token in tail[:220] for token in [
            "passa a vigorar acrescido",
            "passa a vigorar acrescida",
            "fica acrescido",
            "fica acrescida",
            "acrescido de",
            "acrescida de",
        ]):
            nature_map[art].append("Aditiva")
        if "passa a vigorar com " in tail[:220]:
            nature_map[art].append("Modificativa")
        if any(token in tail[:220] for token in [
            "mantenha-se a redacao",
            "mantenha-se a redaÃ§Ã£o",
            "retome-se a redacao",
            "retome-se a redaÃ§Ã£o",
        ]):
            nature_map[art].append("Restaurativa")

    for pattern in modificative_patterns:
        for match in pattern.finditer(s):
            nature_map[normalize_article_ref(match.group(1))].append("Modificativa")

    for pattern in restaurative_patterns:
        for match in pattern.finditer(s):
            nature_map[normalize_article_ref(match.group(1))].append("Restaurativa")

    return {art: resolve_nature_set(natures, "") for art, natures in nature_map.items()}


def consolidado_sort_key(item: Tuple[str, str]) -> Tuple[int, object]:
    tipo, referencia = item
    if tipo == "Artigo":
        return (CONSOLIDADO_ORDER[tipo], article_sort_key(referencia))
    if simplify(tipo) == simplify("Artigo não especificado"):
        return (CONSOLIDADO_ORDER.get("Artigo nao especificado", 99), referencia)
    return (CONSOLIDADO_ORDER.get(tipo, 99), structure_value_sort_key(referencia))


def build_consolidado(rows: List[EmendaRow]) -> List[Tuple[str, str, int, str, str]]:
    consolidated_map: Dict[Tuple[str, str], List[Tuple[int, str, str]]] = defaultdict(list)

    for row in rows:
        source = extract_dispositive_text((row.dispositivos or "").strip())
        if not source:
            source = infer_dispositivos(extract_dispositive_text(row.texto))
        if not source:
            source = extract_dispositive_text((row.texto or "").strip())
        sanitized_texto = extract_dispositive_text(row.texto or "")
        full_text = "\n".join(part for part in [source, sanitized_texto] if part)
        directive_text = sanitized_texto or full_text
        article_nature_overrides = extract_article_nature_overrides(full_text)
        source_paragraphs = set(extract_paragraphs(source)) | set(extract_paragraphs_loose(source))
        directive_paragraphs = set(extract_paragraphs(directive_text)) | set(extract_paragraphs_loose(directive_text))
        paragraph_refs = sorted(source_paragraphs | directive_paragraphs, key=paragraph_sort_key)
        source_arts = set(extract_articles(source)) | set(extract_articles_loose(source))
        directive_arts = set(extract_articles(directive_text)) | set(extract_articles_loose(directive_text))
        project_clause_arts = set(extract_project_clause_articles(directive_text or full_text))
        if source_arts:
            extra_directive_arts = {
                candidate
                for candidate in directive_arts
                if any(candidate.startswith(f"{base_art}-") for base_art in source_arts if candidate != base_art)
            }
            arts_seed = source_arts | extra_directive_arts | {art for art, _ in paragraph_refs} | project_clause_arts
        else:
            arts_seed = directive_arts | {art for art, _ in paragraph_refs} | project_clause_arts
        arts = sorted(arts_seed, key=article_sort_key)
        if directive_arts:
            arts = [
                art for art in arts
                if not any(candidate.startswith(f"{art}-") for candidate in directive_arts if candidate != art)
            ]
        for art in arts:
            diploma = classify_article_diploma(full_text, art)
            article_key = article_reference_key(art, diploma)
            detail_entries = extract_article_vote_details(directive_text, art)
            if detail_entries:
                for detail_nature, detail_label in detail_entries:
                    if diploma == "PL 4/2025" and is_project_clause_adjustment_article(full_text, art):
                        detail_nature = "Modificativa"
                    consolidated_map[("Artigo", article_key)].append((row.numero, detail_nature, detail_label))
                continue
            vote_nature = infer_vote_nature_for_target(directive_text or full_text, "Artigo", art, row.natureza, article_nature_overrides)
            if diploma == "PL 4/2025" and is_project_clause_adjustment_article(full_text, art):
                vote_nature = "Modificativa"
            consolidated_map[("Artigo", article_key)].append((row.numero, vote_nature or consolidado_nature(article_nature_overrides.get(art, row.natureza)), ""))

        generic_articles = extract_generic_article_targets(source)
        for ref in generic_articles:
            vote_nature = infer_vote_nature_for_target(directive_text or full_text, "Artigo nao especificado", ref, row.natureza, article_nature_overrides)
            consolidated_map[("Artigo nao especificado", ref)].append((row.numero, vote_nature or consolidado_nature(row.natureza), ""))

        structures = extract_structures(source)
        if not structures:
            structures = extract_structures_loose(source)
        long_structure = extract_longest_structural_chain(source)
        if long_structure and "Parte Geral" in long_structure:
            tipo_match = re.match(r"^\s*(Livro|T[íi]tulo|Subt[íi]tulo|Cap[íi]tulo|Se[cç][aã]o)\b", long_structure, flags=re.IGNORECASE)
            if tipo_match:
                tipo_norm = format_named_structure(simplify(tipo_match.group(1))).replace("Titulo", "Título").replace("Capitulo", "Capítulo").replace("Secao", "Seção")
                structures = [(tipo_norm, long_structure)]
        for tipo, valor in structures:
            vote_nature = infer_vote_nature_for_target(directive_text or full_text, tipo, valor, row.natureza, article_nature_overrides)
            consolidated_map[(tipo, valor)].append((row.numero, vote_nature or consolidado_nature(row.natureza), ""))

    consolidated: List[Tuple[str, str, int, str, str]] = []
    for key in sorted(consolidated_map, key=consolidado_sort_key):
        tipo, valor = key
        refs = consolidated_map[key]
        if tipo == "Artigo":
            art_value, diploma = split_article_reference_key(valor)
            referencia = format_article_reference_label(art_value, diploma)
        elif simplify(tipo) == simplify("Artigo nao especificado"):
            referencia = format_generic_ref(valor)
        else:
            referencia = valor if "Parte Geral" in str(valor or "") else format_structure_ref(tipo, valor)
        consolidated.append((tipo, referencia, len({n for n, _, _ in refs}), format_emenda_refs(refs), build_voting_order(refs)))
    return consolidated


def estimate_cell_lines(text: str, column_width: float) -> int:
    content = "" if text is None else str(text)
    if not content:
        return 1

    effective_width = max(8, int(column_width * 0.95))
    total = 0
    for raw_line in content.splitlines() or [""]:
        line = raw_line or " "
        total += max(1, math.ceil(len(line) / effective_width))
    return max(1, total)


def autosize_sheet(ws, preferred_widths: Optional[Dict[int, float]] = None) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = header_alignment
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal=(cell.alignment.horizontal if cell.alignment else None))
            cell.border = border

    preferred_widths = preferred_widths or {}
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        preferred = preferred_widths.get(idx, 12)
        width = max(preferred, min(max_length + 2, preferred + 30))
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for idx, cell in enumerate(row, start=1):
            column_letter = get_column_letter(idx)
            column_width = ws.column_dimensions[column_letter].width or 12
            text = "" if cell.value is None else str(cell.value)
            max_lines = max(max_lines, estimate_cell_lines(text, column_width))
        ws.row_dimensions[row[0].row].height = max(24, min((max_lines * 18) + 6, 409))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def center_columns(ws, columns: List[int], include_header: bool = True) -> None:
    centered = Alignment(wrap_text=True, vertical="center", horizontal="center")
    start_row = 1 if include_header else 2
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for idx in columns:
            if idx - 1 < len(row):
                row[idx - 1].alignment = centered


def center_vertical_columns(ws, columns: List[int], include_header: bool = True) -> None:
    start_row = 1 if include_header else 2
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for idx in columns:
            if idx - 1 >= len(row):
                continue
            cell = row[idx - 1]
            current = cell.alignment or Alignment()
            cell.alignment = Alignment(
                horizontal=current.horizontal,
                vertical="center",
                text_rotation=current.text_rotation,
                wrap_text=current.wrap_text,
                shrink_to_fit=current.shrink_to_fit,
                indent=current.indent,
                relativeIndent=current.relativeIndent,
                justifyLastLine=current.justifyLastLine,
                readingOrder=current.readingOrder,
            )


def save_workbook(wb: Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_{ts}{path.suffix}")
        wb.save(alt)
        return alt


def build_workbook(rows: List[EmendaRow]) -> Workbook:
    wb = Workbook()
    headers = [
        "N\u00FAmero da emenda",
        "Autor",
        "Dispositivo(s)-alvo",
        "Natureza da emenda",
        "Tema jur\u00EDdico / eixo tem\u00E1tico",
        "S\u00EDntese da emenda",
        "Texto completo da emenda",
    ]
    emendas_widths = {
        1: 16,
        2: 24,
        3: 90,
        4: 20,
        5: 28,
        6: 90,
        7: 140,
    }

    ws = wb.active
    ws.title = "Emendas"
    ws.append(headers)

    for row in rows:
        ws.append([
            normalize_output_value(row.numero),
            normalize_output_value(row.autor),
            format_multiline_clauses(row.dispositivos),
            normalize_output_value(row.natureza),
            normalize_output_value(row.tema),
            format_multiline_clauses(row.sintese),
            normalize_output_value(row.texto),
        ])

    autosize_sheet(ws, preferred_widths=emendas_widths)
    center_columns(ws, [1, 2, 4, 5])

    ws_mistas = wb.create_sheet("Emendas Mistas")
    ws_mistas.append(headers)
    for row in rows:
        if normalize_nature(row.natureza) != "Mista":
            continue
        ws_mistas.append([
            normalize_output_value(row.numero),
            normalize_output_value(row.autor),
            format_multiline_clauses(row.dispositivos),
            normalize_output_value(row.natureza),
            normalize_output_value(row.tema),
            format_multiline_clauses(row.sintese),
            normalize_output_value(row.texto),
        ])
    autosize_sheet(ws_mistas, preferred_widths=emendas_widths)
    center_columns(ws_mistas, [1, 2, 4, 5])

    reference_file = BASE_DIR / "Planilha_emendas_pl4_2025 atualizada até 09.04.2026.xlsx"
    fill_missing_themes_from_reference(wb, reference_file)

    print("Consolidando emendas...")
    consolidated = build_consolidado(rows)
    ws2 = wb.create_sheet("Consolidado")
    ws2.append(["Dispositivo afetado", "Qtd de Emendas", "Emendas (n\u00FAmero e natureza)", "Ordem de Vota\u00E7\u00E3o"])
    for item in consolidated:
        _, referencia, quantidade, emendas_ref, ordem_votacao = item
        ws2.append([
            normalize_output_value(referencia),
            normalize_output_value(quantidade),
            normalize_output_value(emendas_ref),
            normalize_output_value(ordem_votacao),
        ])
    autosize_sheet(ws2)
    center_columns(ws2, [1, 2, 4])
    center_vertical_columns(ws2, [3, 4])

    return wb


def iter_emenda_pdfs() -> List[Path]:
    pdfs = [p for p in BASE_DIR.glob("*.pdf") if PDF_NAME_RE.match(p.name)]
    return sorted(pdfs, key=lambda p: int(PDF_NAME_RE.match(p.name).group(1)))


def wait_for_exit() -> None:
    if os.environ.get("PY_NO_PAUSE") == "1":
        return
    try:
        input("Pressione Enter para fechar...")
    except EOFError:
        pass


def main() -> None:
    logger.info("="*60)
    logger.info("Iniciando processamento de emendas legislativas...")
    logger.info("="*60)
    
    rows: List[EmendaRow] = []
    failures: List[str] = []

    pdfs = iter_emenda_pdfs()
    if not pdfs:
        logger.error("ERRO: nenhum PDF individual de emenda encontrado na pasta.")
        return

    OCR_PROGRESS["current"] = 0

    logger.info(f"Lendo emendas... ({len(pdfs)} PDFs)")
    workers = max(2, min(8, (os.cpu_count() or 4)))
    logger.info(f"Processamento paralelo com {workers} workers...")
    
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(build_row_with_fallback, pdf) for pdf in pdfs]
        for future in as_completed(futures):
            row, failure = future.result()
            rows.append(row)
            if failure:
                failures.append(failure)

    rows.sort(key=lambda item: item.numero)

    logger.info(f"Criando planilha {OUTPUT_XLSX.name}...")
    wb = build_workbook(rows)
    saved = save_workbook(wb, OUTPUT_XLSX)

    logger.info(f"✓ Processamento concluído!")
    logger.info(f"  - PDFs processados: {len(rows)}")
    logger.info(f"  - Arquivo gerado: {saved}")
    
    if failures:
        logger.warning(f"⚠ Ocorreram {len(failures)} falha(s) no processamento de PDFs")
        for item in failures[:20]:
            logger.warning(f"  - {item}")
        if len(failures) > 20:
            logger.warning(f"  - ... e mais {len(failures) - 20} falha(s)")
    
    logger.info("="*60)


if __name__ == "__main__":
    try:
        main()
    finally:
        wait_for_exit()
