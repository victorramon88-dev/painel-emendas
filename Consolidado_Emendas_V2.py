import importlib.util
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
INPUT_XLSX = BASE_DIR / "Emendas.xlsx"
OUTPUT_XLSX = BASE_DIR / "Consolidado Emendas V2.xlsx"
SOURCE_SCRIPT = BASE_DIR / "Consolidado_Emendas.py"


def load_base_module():
    spec = importlib.util.spec_from_file_location("consolidado_base", SOURCE_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


base = load_base_module()


@dataclass
class EmendaV2Row:
    numero: int
    autor: str
    dispositivos: str
    natureza: str
    tema: str
    resumo: str
    objeto: str
    texto: str


def log(message: str) -> None:
    print(message, flush=True)


def find_header_index(headers: List[str], candidates: List[str]) -> Optional[int]:
    return base.find_header_index(headers, candidates)


def infer_theme_from_devices(dispositivos: str, texto: str) -> str:
    source = base.clean_spaces(dispositivos or "")
    source_s = base.simplify(source)
    long_structure = base.extract_longest_structural_chain(source)

    if "nao se aplica" in source_s:
        return "Não se aplica"
    if "novo artigo no projeto" in source_s:
        return "Não se aplica"
    if "direito civil digital" in source_s or "livro vi" in source_s or "2.027-" in source_s:
        return "DIREITO CIVIL DIGITAL"
    if "livro i-a" in source_s and "animais" in source_s:
        return "PARTE GERAL"

    articles = sorted(
        set(base.extract_articles_loose(source)) | set(base.extract_articles_loose(texto or "")),
        key=base.article_sort_key,
    )
    if not articles:
        if "parte geral" in base.simplify(long_structure):
            return "PARTE GERAL"
        return "Não se aplica"

    selected_art = articles[0]
    selected_diploma = base.classify_article_diploma(texto or source, selected_art)
    for candidate in articles:
        candidate_norm = base.normalize_article_ref(candidate)
        candidate_diploma = base.classify_article_diploma(texto or source, candidate)
        if candidate_diploma != "PL 4/2025":
            selected_art = candidate
            selected_diploma = candidate_diploma
            break
        try:
            candidate_num = int(candidate_norm.split("-")[0].replace(".", ""))
        except Exception:
            candidate_num = None
        if candidate_num != 20:
            selected_art = candidate
            selected_diploma = candidate_diploma
            break

    art_norm = base.normalize_article_ref(selected_art)
    if art_norm.startswith("2.027-"):
        return "DIREITO CIVIL DIGITAL"

    try:
        art_num = int(art_norm.split("-")[0].replace(".", ""))
    except Exception:
        return "Não se aplica"

    diploma = selected_diploma

    if diploma == "PL 4/2025":
        if 11 <= art_num <= 19:
            return "DISPOSIÇÕES FINAIS E TRANSITÓRIAS"
        if art_num == 20:
            return "CLÁUSULA DE REVOGAÇÃO"

    if 1 <= art_num <= 232:
        return "PARTE GERAL"
    if 233 <= art_num <= 420 or 887 <= art_num <= 926:
        return "DIREITO DAS OBRIGAÇÕES E TÍTULOS DE CRÉDITO"
    if 421 <= art_num <= 883:
        return "CONTRATOS E ATOS UNILATERAIS"
    if 884 <= art_num <= 886 or 927 <= art_num <= 954:
        return "RESPONSABILIDADE CIVIL E ENRIQUECIMENTO SEM CAUSA"
    if 966 <= art_num <= 1195:
        return "DIREITO DA EMPRESA"
    if 1196 <= art_num <= 1510:
        return "DIREITO DAS COISAS"
    if 1511 <= art_num <= 1783:
        return "DIREITO DE FAMÍLIA"
    if 1784 <= art_num <= 2027:
        return "DIREITO DAS SUCESSÕES"
    if 2028 <= art_num <= 9999:
        return "DISPOSIÇÕES FINAIS E TRANSITÓRIAS"
    return "Não se aplica"


def load_existing_rows(path: Path) -> List[EmendaV2Row]:
    log(f"Lendo planilha base: {path.name}")
    log("Copiando linhas existentes sem recalcular campos...")
    wb = load_workbook(path, data_only=False, rich_text=True)
    ws = wb["Emendas"] if "Emendas" in wb.sheetnames else wb[wb.sheetnames[0]]

    header = [str(cell.value or "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    numero_idx = find_header_index(header, ["numero da emenda", "numero emenda", "numero"])
    autor_idx = find_header_index(header, ["autor"])
    dispositivos_idx = find_header_index(header, ["dispositivo alvo", "dispositivos alvo", "dispositivoalvo", "dispositivosalvo", "dispositivos alterados"])
    natureza_idx = find_header_index(header, ["natureza da emenda", "natureza"])
    tema_idx = find_header_index(header, ["tema juridico", "eixo tematico"])
    resumo_idx = find_header_index(header, ["resumo da emenda", "resumo"])
    objeto_idx = find_header_index(header, ["objeto da emenda", "sintese da emenda", "sintese"])
    texto_idx = find_header_index(header, ["texto completo da emenda", "texto completo"])

    rows: List[EmendaV2Row] = []
    for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        numero = values[numero_idx - 1] if numero_idx else None
        if numero is None:
            continue
        try:
            numero_int = int(str(numero).strip())
        except Exception:
            continue

        row = EmendaV2Row(
            numero=numero_int,
            autor=str(values[autor_idx - 1] or "").strip() if autor_idx else "",
            dispositivos=str(values[dispositivos_idx - 1] or "").strip() if dispositivos_idx else "",
            natureza=base.normalize_nature(str(values[natureza_idx - 1] or "").strip()) if natureza_idx else "",
            tema=str(values[tema_idx - 1] or "").strip() if tema_idx else "",
            resumo=str(values[resumo_idx - 1] or "").strip() if resumo_idx else "",
            objeto=str(values[objeto_idx - 1] or "").strip() if objeto_idx else "",
            texto=str(values[texto_idx - 1] or "").strip() if texto_idx else "",
        )
        rows.append(row)

    log(f"Emendas carregadas da planilha base: {len(rows)}")
    return sorted(rows, key=lambda row: row.numero)


def iter_new_pdf_paths(existing_numbers: set[int]) -> List[Path]:
    pdfs = [p for p in BASE_DIR.glob("*.pdf") if base.PDF_NAME_RE.match(p.name)]
    pending: List[Path] = []
    for pdf in pdfs:
        match = base.PDF_NAME_RE.match(pdf.name)
        if not match:
            continue
        numero = int(match.group(1))
        if numero in existing_numbers:
            continue
        pending.append(pdf)
    return sorted(pending, key=lambda p: int(base.PDF_NAME_RE.match(p.name).group(1)))


def build_new_row_from_pdf(pdf_path: Path) -> EmendaV2Row:
    log(f"Processando novo PDF: {pdf_path.name}")
    parsed = base.build_row_from_pdf(pdf_path)
    tema = infer_theme_from_devices(parsed.dispositivos, parsed.texto)
    objeto = parsed.sintese
    resumo = parsed.sintese
    return EmendaV2Row(
        numero=parsed.numero,
        autor=parsed.autor,
        dispositivos=parsed.dispositivos,
        natureza=parsed.natureza,
        tema=tema,
        resumo=resumo,
        objeto=objeto,
        texto=parsed.texto,
    )


def to_base_rows(rows: List[EmendaV2Row]) -> List[base.EmendaRow]:
    return [
        base.EmendaRow(
            numero=row.numero,
            autor=row.autor,
            dispositivos=row.dispositivos,
            natureza=row.natureza,
            tema=row.tema,
            sintese=row.objeto,
            texto=row.texto,
        )
        for row in rows
    ]


def append_v2_rows(ws, rows: List[EmendaV2Row]) -> None:
    for row in rows:
        ws.append([
            base.normalize_output_value(row.numero),
            base.normalize_output_value(row.autor),
            base.format_multiline_clauses(row.dispositivos),
            base.normalize_output_value(row.natureza),
            base.normalize_output_value(row.tema),
            base.normalize_output_value(row.resumo),
            base.format_multiline_clauses(row.objeto),
            base.normalize_output_value(row.texto),
        ])


SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace("", SPREADSHEET_NS)
ET.register_namespace("", CONTENT_TYPES_NS)
ET.register_namespace("", RELS_NS)


def excel_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def inline_string_cell(row_number: int, column_number: int, value) -> ET.Element:
    cell = ET.Element(f"{{{SPREADSHEET_NS}}}c", {
        "r": f"{excel_column_name(column_number)}{row_number}",
        "t": "inlineStr",
    })
    inline = ET.SubElement(cell, f"{{{SPREADSHEET_NS}}}is")
    text = ET.SubElement(inline, f"{{{SPREADSHEET_NS}}}t")
    text_value = "" if value is None else str(value)
    text.text = text_value
    if text_value != text_value.strip() or "\n" in text_value:
        text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    return cell


def numeric_cell(row_number: int, column_number: int, value) -> ET.Element:
    cell = ET.Element(f"{{{SPREADSHEET_NS}}}c", {"r": f"{excel_column_name(column_number)}{row_number}"})
    number = ET.SubElement(cell, f"{{{SPREADSHEET_NS}}}v")
    number.text = str(value)
    return cell


def values_for_new_row(row: EmendaV2Row) -> List[str]:
    return [
        base.normalize_output_value(row.numero),
        base.normalize_output_value(row.autor),
        base.format_multiline_clauses(row.dispositivos),
        base.normalize_output_value(row.natureza),
        base.normalize_output_value(row.tema),
        base.normalize_output_value(row.resumo),
        base.format_multiline_clauses(row.objeto),
        base.normalize_output_value(row.texto),
    ]


def append_new_rows_to_sheet_xml(sheet_xml: bytes, source_path: Path, new_rows: List[EmendaV2Row]) -> bytes:
    if not new_rows:
        return sheet_xml

    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(f"{{{SPREADSHEET_NS}}}sheetData")
    if sheet_data is None:
        return sheet_xml

    source_wb = load_workbook(source_path, data_only=False, rich_text=False, read_only=True)
    source_ws = source_wb["Emendas"] if "Emendas" in source_wb.sheetnames else source_wb[source_wb.sheetnames[0]]
    next_row = source_ws.max_row + 1
    source_wb.close()

    for new_row in sorted(new_rows, key=lambda item: item.numero):
        row_el = ET.SubElement(sheet_data, f"{{{SPREADSHEET_NS}}}row", {"r": str(next_row)})
        for column_number, value in enumerate(values_for_new_row(new_row), start=1):
            if column_number == 1:
                row_el.append(numeric_cell(next_row, column_number, value))
            else:
                row_el.append(inline_string_cell(next_row, column_number, value))
        next_row += 1

    dimension = root.find(f"{{{SPREADSHEET_NS}}}dimension")
    if dimension is not None:
        dimension.set("ref", f"A1:H{next_row - 1}")

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def ensure_shared_strings_content_type(content_types_xml: bytes) -> bytes:
    root = ET.fromstring(content_types_xml)
    part_name = "/xl/sharedStrings.xml"
    for override in root.findall(f"{{{CONTENT_TYPES_NS}}}Override"):
        if override.get("PartName") == part_name:
            return content_types_xml
    ET.SubElement(root, f"{{{CONTENT_TYPES_NS}}}Override", {
        "PartName": part_name,
        "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
    })
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def ensure_shared_strings_relationship(rels_xml: bytes) -> bytes:
    root = ET.fromstring(rels_xml)
    shared_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"
    for rel in root.findall(f"{{{RELS_NS}}}Relationship"):
        if rel.get("Type") == shared_type:
            rel.set("Target", "sharedStrings.xml")
            return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    used_ids = {rel.get("Id", "") for rel in root.findall(f"{{{RELS_NS}}}Relationship")}
    next_id = 1
    while f"rId{next_id}" in used_ids:
        next_id += 1
    ET.SubElement(root, f"{{{RELS_NS}}}Relationship", {
        "Id": f"rId{next_id}",
        "Type": shared_type,
        "Target": "sharedStrings.xml",
    })
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def restore_source_emendas_xml(target_path: Path, source_path: Path, new_rows: List[EmendaV2Row]) -> None:
    temp_path = target_path.with_name(f"{target_path.stem}.tmp{target_path.suffix}")
    with zipfile.ZipFile(source_path, "r") as source_zip:
        source_sheet = source_zip.read("xl/worksheets/sheet1.xml")
        source_sheet = append_new_rows_to_sheet_xml(source_sheet, source_path, new_rows)
        source_shared_strings = source_zip.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in source_zip.namelist() else None

    replaced = {"xl/worksheets/sheet1.xml", "[Content_Types].xml", "xl/_rels/workbook.xml.rels"}
    if source_shared_strings is not None:
        replaced.add("xl/sharedStrings.xml")

    with zipfile.ZipFile(target_path, "r") as in_zip, zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as out_zip:
        for item in in_zip.infolist():
            if item.filename in replaced:
                continue
            out_zip.writestr(item, in_zip.read(item.filename))
        out_zip.writestr("xl/worksheets/sheet1.xml", source_sheet)
        if source_shared_strings is not None:
            out_zip.writestr("xl/sharedStrings.xml", source_shared_strings)
        out_zip.writestr("xl/_rels/workbook.xml.rels", ensure_shared_strings_relationship(in_zip.read("xl/_rels/workbook.xml.rels")))
        out_zip.writestr("[Content_Types].xml", ensure_shared_strings_content_type(in_zip.read("[Content_Types].xml")))

    shutil.move(str(temp_path), str(target_path))

def build_workbook_v2(existing_rows: List[EmendaV2Row], new_rows: List[EmendaV2Row], source_path: Path = INPUT_XLSX):
    all_rows = existing_rows + sorted(new_rows, key=lambda row: row.numero)
    log("Montando workbook V2")

    headers = [
        "Número da emenda",
        "Autor",
        "Dispositivo(s)-alvo",
        "Natureza da emenda",
        "Tema jurídico / eixo temático",
        "Resumo da Emenda",
        "Objeto da emenda",
        "Texto completo da emenda",
    ]
    widths = {1: 16, 2: 24, 3: 90, 4: 20, 5: 34, 6: 72, 7: 90, 8: 140}

    wb = load_workbook(source_path, data_only=False, rich_text=False)
    ws = wb["Emendas"] if "Emendas" in wb.sheetnames else wb[wb.sheetnames[0]]
    for sheet_name in list(wb.sheetnames):
        if wb[sheet_name] is not ws:
            del wb[sheet_name]
    if ws.title != "Emendas":
        ws.title = "Emendas"
    if new_rows:
        append_v2_rows(ws, sorted(new_rows, key=lambda row: row.numero))
        base.autosize_sheet(ws, preferred_widths=widths)
        base.center_columns(ws, [1, 2, 4, 5])

    ws_mistas = wb.create_sheet("Emendas Mistas")
    ws_mistas.append(headers)
    append_v2_rows(ws_mistas, [row for row in all_rows if base.normalize_nature(row.natureza) == "Mista"])
    base.autosize_sheet(ws_mistas, preferred_widths=widths)
    base.center_columns(ws_mistas, [1, 2, 4, 5])

    log("Consolidando referências")
    consolidated = base.build_consolidado(to_base_rows(all_rows))
    ws_cons = wb.create_sheet("Consolidado")
    ws_cons.append(["Dispositivo afetado", "Qtd de Emendas", "Emendas (número e natureza)", "Ordem de Votação"])
    for _, referencia, quantidade, emendas_ref, ordem_votacao in consolidated:
        ws_cons.append([
            base.normalize_output_value(referencia),
            base.normalize_output_value(quantidade),
            base.normalize_output_value(emendas_ref),
            base.normalize_output_value(ordem_votacao),
        ])
    base.autosize_sheet(ws_cons)
    base.center_columns(ws_cons, [1, 2, 4])
    base.center_vertical_columns(ws_cons, [3, 4])
    return wb


def main() -> None:
    existing_rows = load_existing_rows(INPUT_XLSX)
    existing_numbers = {row.numero for row in existing_rows}

    new_pdf_paths = iter_new_pdf_paths(existing_numbers)
    log(f"Novos PDFs encontrados: {len(new_pdf_paths)}")
    new_rows = [build_new_row_from_pdf(pdf) for pdf in new_pdf_paths]

    wb = build_workbook_v2(existing_rows, new_rows, INPUT_XLSX)
    saved_path = base.save_workbook(wb, OUTPUT_XLSX)
    restore_source_emendas_xml(saved_path, INPUT_XLSX, new_rows)
    log(f"Planilha salva em: {saved_path}")
    base.wait_for_exit()


if __name__ == "__main__":
    main()










